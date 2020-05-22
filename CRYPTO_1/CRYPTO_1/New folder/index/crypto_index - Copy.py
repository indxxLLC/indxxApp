#%reset -f

import openpyxl
import xlsxwriter
import csv
from datetime import date
import pandas as pd
from datetime import datetime as dt

path = "C:\\CRYPTO_1\\New folder\\updated\\Index_Levels.xlsx"
ind_wb_obj = openpyxl.load_workbook(path) 
ind_sheet_obj = ind_wb_obj.active

def date_convert(datetim):
    return str(datetim.strftime('%Y-%m-%d'))

def date_convert_1(datetim):
    return datetim.date()

'''#################### PARAMETRS ####################'''
                        
Req_dates = "daily" #If daily, picks the latest available date. else the custome dates
req_sheets= "latest"
start_date = dt.strftime(date(2019,7,16), "%Y-%m-%d")

'''#################### PARAMETRS ####################'''
                        
#Run Checks----------------------------------------------------------------------------------
if Req_dates == "daily":
    index_min_row = index_max_row1 = ind_sheet_obj.max_row    
    diff_check = 0
    path = "C:\\CRYPTO_1\\New folder\\updated\\Weights.xlsx"
    weights_wb_obj = openpyxl.load_workbook(path) 
    weights_sheet_obj = weights_wb_obj.active
    weights_max_row1 = weights_min_row = weights_sheet_obj.max_row
else:
    path1 = "C:\\CRYPTO_1\\New folder\\updated\\Index_Levels.xlsx"
    last_date = pd.read_excel(path1,parse_dates = ["Dates"])
    index_min_row = last_date[last_date['Dates'] == start_date].index[-1]
    index_max_row1 = last_date.index[-1]
    index_values_count = index_max_row1-index_min_row
    
    
    path1 = "C:\\CRYPTO_1\\New folder\\updated\\Weights.xlsx"
    if req_sheets == "latest":
        last_date = pd.read_excel(path1,None,parse_dates = ["Date"])
    else:
        last_date = pd.read_excel(path1,sheet_name = req_sheets,parse_dates = ["Date"])
        
    latest_sheet = list(last_date.keys()); sheet = latest_sheet[-1]
    req_sheet = pd.read_excel(path,sheet_name = sheet,parse_dates = ["Date"])
    
    req_sheet['Date'] = req_sheet['Date'].apply(date_convert_1)
    req_sheet['Date'] = req_sheet['Date'].apply(date_convert)
    
    weights_min_row = req_sheet[req_sheet['Date'] == start_date].index[-1]
    weights_max_row1 = req_sheet.index[-1]
    weights_values_count = weights_max_row1-weights_min_row
    
    diff_check = weights_values_count-index_values_count


if diff_check ==0:
    INDEX_ADJUSTER = index_max_row1 - weights_max_row1
    for ind_max_row in range(index_min_row,index_max_row1+1):
        if Req_dates != "daily":
            cell_obj_date = ind_sheet_obj.cell(row = ind_max_row+2, column = 1)
            cell_obj_index_val = ind_sheet_obj.cell(row = ind_max_row+2, column = 3)
            index_date = cell_obj_date.value ; index_value = cell_obj_index_val.value
        else:
            cell_obj_date = ind_sheet_obj.cell(row = ind_max_row, column = 1)
            cell_obj_index_val = ind_sheet_obj.cell(row = ind_max_row, column = 3)
            index_date = cell_obj_date.value ; index_value = cell_obj_index_val.value            
        
        workbook = xlsxwriter.Workbook('C:/inetpub/wwwroot/crypto/output/INFC10_'+str(index_date.strftime("%Y-%m-%d"))+'.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', 'Date') 
        worksheet.write('B1', str(index_date)) 
        worksheet.write('A2', 'Index Value') 
        worksheet.write('B2', index_value)
        
        worksheet.write('A5', 'Symbol') 
        worksheet.write('B5', 'Consituent Name')
        worksheet.write('C5', 'Weights')
        
        
        
        path = "C:\\CRYPTO_1\\New folder\\updated\\Weights.xlsx"
        wb_obj = openpyxl.load_workbook(path) 
          
        # Get workbook active sheet object 
        # from the active attribute 
        sheet_obj = wb_obj.active
        if Req_dates != "daily":
            max_row = ind_max_row-INDEX_ADJUSTER+2
        else:
            max_row = ind_max_row-INDEX_ADJUSTER  #sheet_obj.max_row 
        max_col = sheet_obj.max_column
        csv_path = "C:\\CRYPTO_1\\New folder\\Files\\Historical_Data\\Universe_Masterlist\\CryptoCoin_Masterlist2019-05-29.csv"
        
        
        col = 6
        for i in range(2, max_col+1):
            cell_obj_curr = sheet_obj.cell(row = 1, column = i)
            cell_obj_value = sheet_obj.cell(row = max_row, column = i)
            currency = cell_obj_curr.value.replace("PR ", "")
            weight_value = cell_obj_value.value
            reader = csv.reader(open(csv_path))
#            print(currency)
            for raw in reader:
                if currency == raw[2] and currency != 'XRP':
                    worksheet.write('A'+str(col), raw[1])
                elif currency == raw[2] and currency == 'XRP':
                    currency = 'Ripple'
                    worksheet.write('A'+str(col), raw[1])
            worksheet.write('B'+str(col), currency)
            worksheet.write('C'+str(col), weight_value)
            col = col+1
        	#print(cell_obj_curr.value)
        	#print(cell_obj_value.value)
            
        workbook.close()
    