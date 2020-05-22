import openpyxl
import xlwt
from datetime import datetime, timedelta

"""
Note: Run to particular date. You need to set 'ind_max_row' value and the below set how 
"""
path = "E:\\CRYPTO_1\\New folder\\updated\\Index_Levels.xlsx"
ind_wb_obj = openpyxl.load_workbook(path) 
ind_sheet_obj = ind_wb_obj.active
ind_max_row = ind_sheet_obj.max_row			#Update : max_row-1 to pick one day before index value
cell_obj_date = ind_sheet_obj.cell(row = ind_max_row, column = 1)
cell_obj_index_val = ind_sheet_obj.cell(row = ind_max_row, column = 3)
index_date = cell_obj_date.value
index_value = cell_obj_index_val.value


#today = datetime.today()
# YY-mm-dd
index_date = index_date.strftime("%Y-%m-%d")

date_N_days_ago = (datetime.now() - timedelta(days=1)) #Update : days=2 to get one day before   #.strftime('%Y-%m-%d %H:%M:%S')

style = xlwt.XFStyle()
style.num_format_str = 'YYYY-mm-dd'
sheetname = 'INFC10-values-'+str(index_date)+'.xls'
workbook = xlwt.Workbook()
worksheet = workbook.add_sheet(sheetname)
worksheet.write(0,0, 'Index')
worksheet.write(0,1, 'Value') 
worksheet.write(0,2, 'Date')
worksheet.write(1,0, '.INFC10') 
worksheet.write(1,1, index_value)
worksheet.write(1,2, date_N_days_ago, style)
workbook.save('C:/inetpub/wwwroot/crypto/output/'+sheetname)
print(sheetname+" file created successfully.")


