# -*- coding: utf-8 -*-
"""
Created on Tue Jun 18 12:46:46 2019

@author: anubh
"""
from openpyxl import load_workbook
import pandas as pd
import csv
import pandas as pd
import numpy as np
import datetime as dt
from datetime import date, timedelta, datetime
import pandas
from pandas import read_csv,read_excel, DataFrame, date_range, Series
from dateutil.relativedelta import relativedelta
from numpy import nan, inf, dot



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
      
    
def workdays(d, end, excluded=(6, 7)):
    days = []
    while d.date() <= end.date():
        days.append(d)
        d += dt.timedelta(days=1)
    return days
Calc_Date=dt.datetime.today()-timedelta(1)

RUN_MODE = "daily" #if daily then it runs for everyday else custom run 

if RUN_MODE == 'daily':
    Fact_Dates=workdays(Calc_Date , Calc_Date)
else:
    CALC_START_DATE =  dt.datetime(2019,7,13) # dt.datetime(YYYY,month,day) ex:month:7 for July date = 8
    Fact_Dates=workdays(CALC_START_DATE , Calc_Date)
    
#Fact_Dates=workdays(dt.datetime(2019,7,5) , dt.datetime(2019,7,8))
#Fact_Dates=workdays(Calc_Date , Calc_Date)

Fact_Dates_1=[]
for p in Fact_Dates:
    Fact_Dates_1.append(p.date())
    
start_date=dt.date(year = 2019, month = 5, day = 1)

for dates in Fact_Dates_1:
    pr_coins : list = list()
    temp_dataframe: DataFrame = DataFrame(index = range(len(Fact_Dates_1)))
    temp_dataframe_1: DataFrame = DataFrame(index = range(len(Fact_Dates_1)))
    if dates==dates.replace(day=1):
        portfolio : DataFrame =read_csv(filepath_or_buffer = "C:\\CRYPTO_1\\New folder\\Files\\Historical_Data\\Poortfolios\\CryptoIndex_new1_2019_05_31.csv" ,encoding = "utf-8")
    else:
        portfolio : DataFrame =read_csv("C:\\CRYPTO_1\\New folder\\Files\\Historical_Data\\Poortfolios\\CryptoIndex_2019_06_27.csv",encoding="utf-8")
    portfolio["Weights"] = portfolio["Weights"].map(lambda wt : wt / 100)
    
    
    
    for index_pf in portfolio.index :
        coin_history_data: DataFrame = read_excel("C:\\CRYPTO_1\\New folder\\Files\\Historical_Data\\" +
                                                    portfolio.at[index_pf, "Coin"].title() + ".xlsx",sheet_name='Sheet1',
                                                parse_dates = ["Date"], dayfirst = True)
        coin_history_data["Date"] = coin_history_data["Date"].map(lambda dt : dt.date())
        temp_dataframe['Date']=dates
        temp_dataframe_1['Date']=dates
        Date_p1=dates+timedelta(1)
        Date_p2=dates+timedelta(2)
        
        if dates==dates.replace(day=1):
            base_price_date=(dates-relativedelta(months=1)).replace(day=1)
            base_price_date_p1=base_price_date+timedelta(1)
            base_price_date=base_price_date_p1
        else:
            base_price_date=dates.replace(day=2)
        base_price : float = coin_history_data[coin_history_data['Date']==base_price_date]["Current Price"].values[0]
        pr_coins.append("PR " + portfolio.at[index_pf, "Coin"])
        a = coin_history_data [coin_history_data["Date"]==Date_p1]["Current Price"].values[0]
        b=(a-base_price)/base_price
        temp_dataframe[ portfolio.at[index_pf, "Coin"]]=a
        temp_dataframe["PR "+ portfolio.at[index_pf, "Coin"]]=b
        temp_dataframe_1["PR "+ portfolio.at[index_pf, "Coin"]]=1+b
        #month_first_date=dates.replace(day=1)
        #b2=temp_dataframe_1["PR "+ portfolio.at[index_pf, "Coin"]].tail(1).values[0]

        
    if dates==start_date:
        if dates==dates.replace(day=1):
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",temp_dataframe.tail(1) , sheet_name='Prices'+str(dates), header=True, index=False)
            #append_df_to_excel("D:\\New folder\\updated\\Weights.xlsx",temp_dataframe_1.tail(1) , sheet_name='Weights_1'+str(dates), header=True, index=False)
        else:
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",temp_dataframe.tail(1) , sheet_name='Prices'+str(dates), header=True, index=False)
            #append_df_to_excel("D:\\New folder\\updated\\Price_Returns.xlsx",temp_dataframe_1.tail(1) , sheet_name='Weights_1'+str(dates), header=True, index=False)
    else:
        if dates==dates.replace(day=1):
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",temp_dataframe.tail(1) , sheet_name='Prices'+str(dates), header=True, index=False)
            #append_df_to_excel("D:\\New folder\\updated\\Weights.xlsx",temp_dataframe_1.tail(1) , sheet_name='Weights_1'+str(dates), header=True, index=False)
        elif dates==dates.replace(day=2):
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",temp_dataframe.tail(1) , sheet_name='Prices'+str(dates), header=True, index=False)
            #append_df_to_excel("D:\\New folder\\updated\\Weights.xlsx",temp_dataframe_1.tail(1) , sheet_name='Weights'+str(dates), header=True, index=False)
        else:
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",temp_dataframe.tail(1) , sheet_name='Prices'+str(dates.replace(day=2)), header=False, index=False)
            #append_df_to_excel("D:\\New folder\\updated\\Weights.xlsx",temp_dataframe_1.tail(1) , sheet_name='Weights'+str(dates.replace(day=2)), header=False, index=False)
    
    
    first_day_from_fact_date=dates.replace(day=1)
    
    if dates==first_day_from_fact_date:
        base_index_level_date=(first_day_from_fact_date-relativedelta(months=1)).replace(day=1)
    else:
        base_index_level_date=first_day_from_fact_date
    
    
    
    
    if dates==start_date:
        base_index_value : float = 78341.3563
        Weighted_Returns = dot(portfolio["Weights"].values, temp_dataframe[temp_dataframe["Date"] == dates][pr_coins].values.T)[0]
        df = pd.DataFrame([[dates,Weighted_Returns, base_index_value]])
        append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Index_Levels.xlsx",df , sheet_name='Prices', header=["Dates","Weighted Returns","Index_level"], index=False)
       
    else:
        price_returns=pd.read_excel("C:\\CRYPTO_1\\New folder\\updated\\Index_Levels.xlsx",sheet_name='Prices',parse_dates = ["Dates"])
        if dates==dates.replace(day=1):
            temp_dataframe=pd.read_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",sheet_name='Prices'+str(dates),parse_dates = ["Date"])
        else:
            temp_dataframe=pd.read_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",sheet_name='Prices'+str(dates.replace(day=2)),parse_dates = ["Date"])
        date_m1=dates-timedelta(days=1)
        base_index_value=price_returns[price_returns['Dates']==pd.Timestamp(base_index_level_date)]["Index_level"].values[0]
        Weighted_Returns = dot(portfolio["Weights"].values, temp_dataframe[temp_dataframe["Date"] == pd.Timestamp(dates)][pr_coins].values.T)[0]
        
        Index_Value = (1 + Weighted_Returns) * base_index_value
        df1=pd.DataFrame([[dates,Weighted_Returns, Index_Value]])
        append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Index_Levels.xlsx",df1 , sheet_name='Prices', header=False, index=False)
       
    temp_dataframe_3: DataFrame = DataFrame(index = range(len(Fact_Dates_1)))
    for index_pf1 in portfolio.index:
        temp_dataframe_3['Date']=dates
        
        if dates==dates.replace(day=1):
            df=pd.read_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",sheet_name='Prices'+str(dates),parse_dates = ["Date"])
        else:
            df=pd.read_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",sheet_name='Prices'+str(dates.replace(day=2)),parse_dates = ["Date"])
        
        
        
        
        if dates==dates.replace(day=2):
            temp_dataframe_3["PR "+ portfolio.at[index_pf1, "Coin"]]=portfolio.at[index_pf1,'Weights']
        else:
            date_tm1=dates-timedelta(1)
            last_date_prev_month=date.today().replace(day=1)-timedelta(days=1)
            if date_tm1==last_date_prev_month:
                df=pd.read_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",sheet_name='Prices'+str(last_date_prev_month.replace(day=2)),parse_dates = ["Date"])
            f=np.array(portfolio['Weights'])[index_pf1]
            p=1+df[df["Date"] == pd.Timestamp(date_tm1)][pr_coins].values.T
            Weighted_Returns_1=Weighted_Returns = dot(portfolio["Weights"].values, 1+df[df["Date"] == pd.Timestamp(date_tm1)][pr_coins].values.T)[0]
            temp_dataframe_3["PR "+ portfolio.at[index_pf1, "Coin"]]=f*((p[index_pf1][0])/Weighted_Returns_1)
            
    if dates==start_date:
        if dates==dates.replace(day=1):
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Weights.xlsx",temp_dataframe_3.tail(1) , sheet_name='Weights'+str(dates), header=True, index=False)
        else:
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Price_Returns.xlsx",temp_dataframe_3.tail(1) , sheet_name='Weights'+str(dates), header=True, index=False)
    else:
        if dates==dates.replace(day=1):
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Weights.xlsx",temp_dataframe_3.tail(1) , sheet_name='Weights'+str(dates), header=True, index=False)
        elif dates==dates.replace(day=2):
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Weights.xlsx",temp_dataframe_3.tail(1) , sheet_name='Weights'+str(dates), header=True, index=False)
        else:
            append_df_to_excel("C:\\CRYPTO_1\\New folder\\updated\\Weights.xlsx",temp_dataframe_3.tail(1) , sheet_name='Weights'+str(dates.replace(day=2)), header=False, index=False)
    

''' if dates==month_first_date:
            temp_dataframe_1["PR_W "+ portfolio.at[index_pf, "Coin"]]=0.1
        else:
            date_m1=dates-timedelta(1)
            
            temp_dataframe_1["PR_W "+ portfolio.at[index_pf, "Coin"]]=0.1  '''