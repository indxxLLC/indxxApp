# calendar/views.py
"""
 * Version : 1.0
 * Project: calendar Automation
 * Copyright : Indxx Capital Management
 * Author: Pavan Rajput
 * Created Date: 08-04-2019
 * Modified Date: dd-mm-yyyy
 * Licensed under : Self
"""
# Create your views here.
from django.contrib.auth.models import User
from django.core.paginator import Paginator,EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from django.shortcuts import render, render_to_response
from django.template import loader
from api.models import Calendarlist, Registerindex, Ruleslist
from django.utils.translation import ugettext as _
from mycalendar.effectivedate_rule1 import EffectiveDate_Calculate
from mycalendar.AnnouncementDate_rule import AnnouncementDate_Calculate
from mycalendar.SelectionDate2_rule import SelectionDate_Calculate
from mycalendar.FreezeDate_rule import FreezeDate_Calculate
from datetime import date, datetime, timedelta
from mycalendar.Preliminarycommdate_rule import PreliminaryCommDate_Calculate
from mycalendar.CompletionDate_cycle1_rule import CompletionDate_Calculate
from mycalendar.Committee_commdate_rule import Committe_comm_date_Calculate
from mycalendar.SelectionDate1_rule import selection_date_1_Calculate
from django.db.models import Q

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pdb;


def index(request):
	templateName = "index.html"
	clientList = Calendarlist.objects.filter(category = 'Client_Name')
	
	indexlist = Registerindex.objects.all().order_by('id').reverse()
	page = request.GET.get('page', 1)
	paginator = Paginator(indexlist, 20)
	try:
		indexdata = paginator.page(page)
	except PageNotAnInteger:
		indexdata = paginator.page(1)
	except EmptyPage:
		indexdata = paginator.page(paginator.num_pages)
		
	context = {
		'indexlist': indexdata,
		'clientList': clientList,
	}
		
	return render(request, templateName, context)
"""
if request.method == 'POST':
search_result = request.POST['search2']
if(search_result!=""):
#templatename = "SearchIndex.html"
lookups = Q(Index_Name__icontains = search_result) | Q(Contract_Type__icontains = search_result) | Q(Reconstitution__icontains = search_result) | Q(Calc_Agent__icontains = search_result) | Q(Ind_Sty__icontains = search_result) | Q(Contract_Type__icontains = search_result) | Q(Type_of_Ind__icontains = search_result)
results= Registerindex.objects.filter(lookups).distinct().reverse()

if(len(results) == 0):
obj = Calendarlist.objects.filter(description = search_result)
code = ""
for ind in obj:
	code = int(ind.code)
	if(code!=""):
		results = Registerindex.objects.filter(Client_Name_id = code)
	else:
		templateName = "Nosearch.html"


context={
	'search_text': search_result,
	'results': results,
	'clientList': clientList,
}
return render(request, templateName, context)

else:
templateName = "Nosearch.html"
context={
	'search_text': search_result,
	'results': '',
	'clientList': clientList,
}

else:
"""

	
def mycalendar(request):
    templates = loader.get_template("index.html")
    return HttpResponse(templates.render());	

def addindex(request):
	template = "addindex.html"
	client_name = Calendarlist.objects.filter(category = 'Client_Name')
	ind_Sty = Calendarlist.objects.filter(category = 'Ind_Sty')
	ind_Version = Calendarlist.objects.filter(category = 'Ind_Ver')
	cal = Calendarlist.objects.filter(category = 'Cal')
	cal_Agent = Calendarlist.objects.filter(category = 'Cal_Agent')
	contract_Type = Calendarlist.objects.filter(category = 'Contract_Type')
	type_of_Ind = Calendarlist.objects.filter(category = 'Type_of_Ind')
	prod_Stat = Calendarlist.objects.filter(category = 'Prod_Stat')
	etf_Launched = Calendarlist.objects.filter(category = 'ETF_Launched')
	reconstitution = Calendarlist.objects.filter(category = 'Reconstitution')
	rebalance = Calendarlist.objects.filter(category = 'Rebalance')
	review = Calendarlist.objects.filter(category = 'Review')
	theme_Review = Calendarlist.objects.filter(category = 'Theme_Review')
	comp_Date = Calendarlist.objects.filter(category = 'Comp_Date')
	selec_Date_Cyc_1 = Calendarlist.objects.filter(category = 'Selec_Date_Cyc_1')
	selec_Date_Cyc_2 = Calendarlist.objects.filter(category = 'Selec_Date_Cyc_2')
	weights_Share_Freeze = Calendarlist.objects.filter(category = 'Weights_Share_Freeze_Date')
	public_Announcement = Calendarlist.objects.filter(category = 'Public_Announcement_Date')
	client_Comm = Calendarlist.objects.filter(category = 'Client_Comm_Date')
	ind_Cmte_Comm_Date = Calendarlist.objects.filter(category = 'Ind_Cmte_Comm_Date')
	effec_Date = Calendarlist.objects.filter(category = 'Effec_Date')
	prelim_Comm_Date = Calendarlist.objects.filter(category = 'Prelim_Comm_Date')
	comm_to_Calc_Agent = Calendarlist.objects.filter(category = 'Comm_to_Calc_Agent')

	
	context = {
		'client_name':client_name,
		'ind_Sty':ind_Sty,
		'ind_Version':ind_Version,
		'cal':cal,
		'cal_Agent':cal_Agent,
		'contract_Type':contract_Type,
		'type_of_Ind':type_of_Ind,
		'prod_Stat':prod_Stat,
		'etf_Launched':etf_Launched,
		'reconstitution':reconstitution,
		'rebalance':rebalance,
		'review':review,
		'theme_Review':theme_Review,
		'comp_Date':comp_Date,
		'selec_Date_Cyc_1':selec_Date_Cyc_1,
		'selec_Date_Cyc_2':selec_Date_Cyc_2,
		'weights_Share_Freeze':weights_Share_Freeze,
		'public_Announcement':public_Announcement,
		'client_Comm':client_Comm,
		'ind_Cmte_Comm_Date':ind_Cmte_Comm_Date,
		'effec_Date':effec_Date,
		'prelim_Comm_Date':prelim_Comm_Date,
		'comm_to_Calc_Agent':comm_to_Calc_Agent,
		}
		
	return render(request, template, context)



def post_index(request):
	if request.method == 'POST':
		if ((request.POST['mnl_cmp_date'] != "" and request.POST['cmp_Date'] != "") or (
                request.POST['mnl_ind_Comm_date'] != "" and request.POST['ind_Comm_Date'] != "")
                or (request.POST['mnl_pb_announce_date'] != "" and request.POST['public_Announcement'] != "") or (
                        request.POST['mnl_client_comm_date'] != "" and request.POST['client_Comm'] != "")
                or (request.POST['mnl_eff_date'] != "" and request.POST['effective_date'] != "") or (
                        request.POST['mnl_prelim_date'] != "" and request.POST['pre_comm_date'] != "")
                or (request.POST['mnl_comm_cal_date'] != "" and request.POST['comm_to_Calc_Agent'] != "") or (
                        request.POST['mnl_sel1_date'] != "" and request.POST['selec_Date_Cyc_1'] != "") or
                (request.POST['mnl_sel2_date'] != "" and request.POST['selec_Date_Cyc_2'] != "") or (
                        request.POST['mnl_freeze_date'] != "" and request.POST['weights_Share_Freeze'] != "")):
			template = "500.html"
			context = {"message": "not saved"}
			return render(request, template, context)
		else:
			if(request.POST['cmp_Date'] != "" or request.POST['ind_Comm_Date'] != "" or request.POST['client_Comm'] != "" or request.POST['public_Announcement'] != "" or request.POST['effective_date'] != "" or 
					request.POST['pre_comm_date'] != "" or request.POST['comm_to_Calc_Agent'] != "" or request.POST['selec_Date_Cyc_2'] != "" or request.POST['selec_Date_Cyc_1'] != "" or 
					request.POST['weights_Share_Freeze'] != ""):
				if((request.POST['reconstitution']=="Annual") or (request.POST['reconstitution']=="Monthly")):
					num_month = 1
					gap=1
				elif(request.POST['reconstitution']== "Quarterly"):
					num_month = 4
					gap = 3
				else:
					num_month = 2
					gap = 6

				reconst_Effective_Date = ""
				reconst_Selection_Date_Cycle_2 = ""
				reconst_Weights_Share_Freeze_Date = ""
				reconst_Public_Announcement_Date = ""
				reconst_Client_Comm_Date = ""
				reconst_Prelim_Comm_Date = ""
				reconst_Completion_Date = ""
				reconst_Ind_Cmte_Comm_Date = ""
				reconst_Selection_Date_Cycle_1 = ""
				rev_effective = ""
				rev_selection2 = ""
				rev_freeze = ""
				rev_announce = ""
				rev_clientcomm = ""
				rev_prelim = ""
				rev_completion = ""
				rev_commiteecomm = ""
				rev_selection1 = ""
				re_effective = ""
				re_selection2 = ""
				re_freeze = ""
				re_announce = ""
				re_clientcomm = ""
				re_prelim = ""
				re_completion = ""
				re_commiteecomm = ""
				re_selection1 = ""

				months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
							  October=10, November=11, December=12)
				month1 = months[request.POST['reconst_month']]
				month1 = int(month1)
				cal_month = month1
				year = date.today().year
				for i in range(0,num_month):
					if(month1>12):
						month1 = month1%12;
					if(month1<cal_month):
						year = year+1
					 
					eff_rule = int(request.POST['effective_date'])
					sel_rule = int(request.POST['selec_Date_Cyc_2'])
					fre_rule = int(request.POST['weights_Share_Freeze'])
					ann_rule = int(request.POST['public_Announcement'])
					client_comm_rule = int(request.POST['client_Comm'])
					prelim_comm_rule = int(request.POST['pre_comm_date'])
					#prelim_comm_rule = int(request.POST['pre_comm_date'])
					color_1 = 1
					date1 = EffectiveDate_Calculate(month1, year, eff_rule)
					eff_day = int(date1[3:5])
					eff_mon = int(date1[0:2])
					eff_year = int(date1[6:8])
					sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule)
					sel_day = int(sel_date[3:5])
					sel_month = int(sel_date[0:2])
					sel_year = int(sel_date[6:8])
					freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule)
					freeze_day = int(freeze_date[3:5])
					freeze_mon = int(freeze_date[0:2])
					freeze_year = int(freeze_date[6:8])
					if (ann_rule == 1):
						announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule)
					else:
						announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule)
					announce_day = int(announce_date[3:5])
					announce_mon = int(announce_date[0:2])
					announce_year = int(announce_date[6:8])
					if (client_comm_rule == 1):
						client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule)
					else:
						client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule)
					Client_comm_day = int(client_comm_date[3:5])
					Client_comm_mon = int(client_comm_date[0:2])
					Client_comm_year = int(client_comm_date[6:8])
					Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day, announce_mon, announce_year)
					Prelim_comm_day = int(Prelim_comm_date[3:5])
					Prelim_comm_mon = int(Prelim_comm_date[0:2])
					Prelim_comm_year = int(Prelim_comm_date[6:8])
					Completion_date = CompletionDate_Calculate(sel_day, sel_month, sel_year)
					if (prelim_comm_rule == ""):
						Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon,
																		   Client_comm_year)
					else:
						Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day, Prelim_comm_mon,
																		   Prelim_comm_year)

					selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year)
					cal_month = month1
					month1= month1+gap
					reconst_Effective_Date = reconst_Effective_Date + date1 + ", "
					reconst_Selection_Date_Cycle_2 = reconst_Selection_Date_Cycle_2 + sel_date + ", "
					reconst_Weights_Share_Freeze_Date = reconst_Weights_Share_Freeze_Date + freeze_date + ", "
					reconst_Public_Announcement_Date = reconst_Public_Announcement_Date + announce_date + ", "
					reconst_Client_Comm_Date = reconst_Client_Comm_Date + client_comm_date + ", "
					reconst_Prelim_Comm_Date = reconst_Prelim_Comm_Date + Prelim_comm_date + ", "
					reconst_Completion_Date = reconst_Completion_Date + Completion_date + ", "
					reconst_Ind_Cmte_Comm_Date = reconst_Ind_Cmte_Comm_Date + Committee_comm_date + ", "
					reconst_Selection_Date_Cycle_1 = reconst_Selection_Date_Cycle_1+selection_date1+ ", "

				if (request.POST['review_month'] != ""):

					eff_rule_rev = int(request.POST['effective_date_review'])
					sel_rule_rev = int(request.POST['selec_Date_Cyc_2_review'])
					fre_rule_rev = int(request.POST['weights_Share_Freeze_review'])
					ann_rule_rev = int(request.POST['public_Announcement_review'])
					client_comm_rule_rev = int(request.POST['client_Comm_review'])
					prelim_comm_rule_rev = int(request.POST['pre_comm_date_review'])
					#prelim_comm_rule_rev = int(request.POST['pre_comm_date_review'])

					if(request.POST['review']=="Annual" or request.POST['review']=="Monthly"):
						num_month = 1
						gap=1
					elif(request.POST['review']== "Quarterly"):
						num_month = 4
						gap = 3
					else:
						num_month = 2
						gap = 6
					months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
							  October=10, November=11, December=12)
					month1 = months[request.POST['review_month']]
					month1 = int(month1)
					cal_month = month1
					year = date.today().year
					for i in range(0,num_month):
						if(month1>12):
						   month1 = month1%12;
						if(month1<cal_month):
						   year = year+1
						date1 = EffectiveDate_Calculate(month1, year, eff_rule_rev)
						eff_day = int(date1[3:5])
						eff_mon = int(date1[0:2])
						eff_year = int(date1[6:8])

						sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule_rev)
						sel_day = int(sel_date[3:5])
						sel_month = int(sel_date[0:2])
						sel_year = int(sel_date[6:8])

						freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule_rev)
						freeze_day = int(freeze_date[3:5])
						freeze_mon = int(freeze_date[0:2])
						freeze_year = int(freeze_date[6:8])

						if (ann_rule_rev == 1):
							announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rev)
						else:
							announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rev)
						announce_day = int(announce_date[3:5])
						announce_mon = int(announce_date[0:2])
						announce_year = int(announce_date[6:8])

						if (client_comm_rule_rev == 1):
							client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rev)
						else:
							client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rev)
						Client_comm_day = int(client_comm_date[3:5])
						Client_comm_mon = int(client_comm_date[0:2])
						Client_comm_year = int(client_comm_date[6:8])

						Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day, announce_mon, announce_year)
						Prelim_comm_day = int(Prelim_comm_date[3:5])
						Prelim_comm_mon = int(Prelim_comm_date[0:2])
						Prelim_comm_year = int(Prelim_comm_date[6:8])

						Completion_date = CompletionDate_Calculate(sel_day, sel_month, sel_year)
						if (prelim_comm_rule_rev == ""):
							Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon,
																			   Client_comm_year)
						else:
							Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day, Prelim_comm_mon,
																			   Prelim_comm_year)
						selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year)
						cal_month = month1
						month1= month1+gap
						rev_effective = rev_effective + date1 + ", "
						rev_selection2 = rev_selection2 + sel_date + ", "
						rev_freeze = rev_freeze + freeze_date + ", "
						rev_announce = rev_announce + announce_date + ", "
						rev_clientcomm = rev_clientcomm + client_comm_date + ", "
						rev_prelim = rev_prelim + Prelim_comm_date + ", "
						rev_completion = rev_completion + Completion_date + ", "
						rev_commiteecomm = rev_commiteecomm + Committee_comm_date + ", "
						rev_selection1 = rev_selection1 + selection_date1 + ", "

				if (request.POST['rebalance_month'] != ""):

					eff_rule_rebal = request.POST['effective_date_rebal']
					sel_rule_rebal = request.POST['selec_Date_Cyc_2_rebal']
					fre_rule_rebal = request.POST['weights_Share_Freeze_rebal']
					ann_rule_rebal = request.POST['public_Announcement_rebal']
					client_comm_rule_rebal = request.POST['client_Comm_rebal']
					prelim_comm_rule_rebal = request.POST['pre_comm_date_rebal']
					#prelim_comm_rule_rebal = int(request.POST['pre_comm_date_rebal'])
					if(request.POST['rebalance']=="Annual" or request.POST['rebalance']=="Monthly"):
						num_month = 1
						gap=1
					elif(request.POST['rebalance']== "Quarterly"):
						num_month = 4
						gap = 3
					else:
						num_month = 2
						gap = 6

					months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
							  October=10, November=11, December=12)
					month1 = months[request.POST['rebalance_month']]
					month1 = int(month1)
					cal_month = month1
					year = date.today().year
					for i in range(0,num_month):
						if(month1>12):
						   month1 = month1%12;
						if(month1<cal_month):
						   year = year+1

						date1 = EffectiveDate_Calculate(month1, year, eff_rule_rebal)
						eff_day = int(date1[3:5])
						eff_mon = int(date1[0:2])
						eff_year = int(date1[6:8])
						sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule_rebal)
						sel_day = int(sel_date[3:5])
						sel_month = int(sel_date[0:2])
						sel_year = int(sel_date[6:8])
						freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule_rebal)
						freeze_day = int(freeze_date[3:5])
						freeze_mon = int(freeze_date[0:2])
						freeze_year = int(freeze_date[6:8])

						if (ann_rule_rebal == 1):
							announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rebal)
						else:
							announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rebal)
						announce_day = int(announce_date[3:5])
						announce_mon = int(announce_date[0:2])
						announce_year = int(announce_date[6:8])

						if (client_comm_rule_rebal == 1):
							client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rebal)
						else:
							client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rebal)
						Client_comm_day = int(client_comm_date[3:5])
						Client_comm_mon = int(client_comm_date[0:2])
						Client_comm_year = int(client_comm_date[6:8])

						Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day, announce_mon, announce_year)
						Prelim_comm_day = int(Prelim_comm_date[3:5])
						Prelim_comm_mon = int(Prelim_comm_date[0:2])
						Prelim_comm_year = int(Prelim_comm_date[6:8])

						Completion_date = CompletionDate_Calculate(sel_day, sel_month, sel_year)

						if (prelim_comm_rule_rebal == ""):
							Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon,
																			   Client_comm_year)
						else:
							Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day, Prelim_comm_mon,
																			   Prelim_comm_year)
						selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year)
						cal_month = month1
						month1= month1+gap
						re_effective = re_effective + date1 + ", "
						re_selection2 = re_selection2 + sel_date + ", "
						re_freeze = re_freeze + freeze_date + ", "
						re_announce = re_announce + announce_date + ", "
						re_clientcomm = re_clientcomm + client_comm_date + ", "
						re_prelim = re_prelim + Prelim_comm_date + ", "
						re_completion = re_completion + Completion_date + ", "
						re_commiteecomm = re_commiteecomm + Committee_comm_date + ", "
						re_selection1 = re_selection1 + selection_date1 + ", "

				else:

					color_1 = 0

					reconst_Effective_Date = request.POST['mnl_eff_date'] if(request.POST['mnl_eff_date']=="") else datetime.datetime.strptime(request.POST['mnl_eff_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					reconst_Selection_Date_Cycle_2 = request.POST['mnl_sel2_date'] if(request.POST['mnl_sel2_date']=="") else datetime.datetime.strptime(request.POST['mnl_sel2_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					reconst_Weights_Share_Freeze_Date = request.POST['mnl_freeze_date'] if(request.POST['mnl_freeze_date']=="") else datetime.datetime.strptime(request.POST['mnl_freeze_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					reconst_Public_Announcement_Date = request.POST['mnl_pb_announce_date'] if(request.POST['mnl_pb_announce_date']=="") else datetime.datetime.strptime(request.POST['mnl_pb_announce_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					reconst_Client_Comm_Date = request.POST['mnl_client_comm_date'] if(request.POST['mnl_client_comm_date']=="") else datetime.datetime.strptime(request.POST['mnl_client_comm_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					reconst_Prelim_Comm_Date = request.POST['mnl_prelim_date'] if(request.POST['mnl_prelim_date']=="") else datetime.datetime.strptime(request.POST['mnl_prelim_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					reconst_Completion_Date = request.POST['mnl_cmp_date'] if(request.POST['mnl_cmp_date']=="") else datetime.datetime.strptime(request.POST['mnl_cmp_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					reconst_Ind_Cmte_Comm_Date = request.POST['mnl_ind_Comm_date'] if(request.POST['mnl_ind_Comm_date']=="") else datetime.datetime.strptime(request.POST['mnl_ind_Comm_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					reconst_Selection_Date_Cycle_1 =request.POST['mnl_sel1_date'] if(request.POST['mnl_sel1_date']=="") else datetime.datetime.strptime(request.POST['mnl_sel1_date'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_effective = request.POST['mnl_eff_date_review'] if(request.POST['mnl_eff_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_eff_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_selection1 = request.POST['mnl_sel1_date_review'] if(request.POST['mnl_sel1_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_sel1_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_announce = request.POST['mnl_pb_announce_date_review'] if(request.POST['mnl_pb_announce_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_pb_announce_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_clientcomm = request.POST['mnl_client_comm_date_review'] if(request.POST['mnl_client_comm_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_client_comm_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_commiteecomm = request.POST['mnl_ind_Comm_date_review'] if(request.POST['mnl_ind_Comm_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_ind_Comm_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_completion = request.POST['mnl_cmp_date_review'] if(request.POST['mnl_cmp_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_cmp_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_prelim = request.POST['mnl_prelim_date_review'] if(request.POST['mnl_prelim_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_prelim_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_freeze = request.POST['mnl_freeze_date_review'] if(request.POST['mnl_freeze_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_freeze_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					rev_selection2 = request.POST['mnl_sel2_date_review'] if(request.POST['mnl_sel2_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_sel2_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_announce = request.POST['mnl_pb_announce_date_rebal'] if(request.POST['mnl_pb_announce_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_pb_announce_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_clientcomm = request.POST['mnl_client_comm_date_rebal'] if(request.POST['mnl_client_comm_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_client_comm_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_commiteecomm = request.POST['mnl_ind_Comm_date_rebal'] if(request.POST['mnl_ind_Comm_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_ind_Comm_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_completion = request.POST['mnl_cmp_date_rebal'] if(request.POST['mnl_cmp_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_cmp_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_effective = request.POST['mnl_eff_date_rebal'] if(request.POST['mnl_eff_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_eff_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_freeze = request.POST['mnl_freeze_date_rebal'] if(request.POST['mnl_freeze_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_freeze_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_prelim = request.POST['mnl_prelim_date_rebal'] if(request.POST['mnl_prelim_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_prelim_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_selection1 = request.POST['mnl_sel1_date_rebal'] if(request.POST['mnl_sel1_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_sel1_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					re_selection2 = request.POST['mnl_sel2_date_rebal'] if(request.POST['mnl_sel2_date_rebal']=="") else datetime.datetime.strptime(request.POST['mnl_sel2_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%Y")
					index_data = Registerindex(
						Ident_ISIN=request.POST['isin_id'],
						Ident_Bloomberg=request.POST['bloomberg_id'],
						Ident_Reuters=request.POST['thomson_id'],
						Index_Name=request.POST['index_name'],
						Client_Name_id=request.POST['client_name'],
						Ind_Sty=request.POST['index_Style'],
						Ind_Ver=request.POST['ind_version'],
						Ind_Ver_ID=request.POST['index_version_id'],
						Calc=request.POST['calculation'],
						Calc_Agent=request.POST['cal_agent'],
						Data_Platform=request.POST['data_platform'],
						Data_Vendors=request.POST['data_vendors'],
						Contract_Type=request.POST['contract_Type'],
						Type_of_Ind=request.POST['type_index'],
						Product_Status=request.POST['prod_Status'],
						ETF_Launched=request.POST['etf_Launched'],
						Reconstitution=request.POST['reconstitution'],
						Rebalance=request.POST['rebalance'],
						Review=request.POST['review'],
						Theme_Review=request.POST['theme_Review'],
						Selection_Date_Cycle_1=reconst_Selection_Date_Cycle_1,
						Selection_Date_Cycle_1_rebal=re_selection1,
						Selection_Date_Cycle_1_review=rev_selection1,
						Completion_Date=reconst_Completion_Date,
						Completion_Date_rebal=re_completion,
						Completion_Date_review=rev_completion,
						Selection_Date_Cycle_2=reconst_Selection_Date_Cycle_2,
						Selection_Date_Cycle_2_rebal=re_selection2,
						Selection_Date_Cycle_2_review=rev_selection2,
						Ind_Cmte_Comm_Date=reconst_Ind_Cmte_Comm_Date,
						Ind_Cmte_Comm_Date_rebal=re_commiteecomm,
						Ind_Cmte_Comm_Date_review=rev_commiteecomm,
						Prelim_Comm_Date=reconst_Prelim_Comm_Date,
						Prelim_Comm_Date_rebal=re_prelim,
						Prelim_Comm_Date_review=rev_prelim,
						Weights_Share_Freeze_Date=reconst_Weights_Share_Freeze_Date,
						Weights_Share_Freeze_Date_rebal=re_freeze,
						Weights_Share_Freeze_Date_review=rev_freeze,
						Public_Announcement_Date=reconst_Public_Announcement_Date,
						Public_Announcement_Date_rebal=re_announce,
						Public_Announcement_Date_review=rev_announce,
						Client_Comm_Date=reconst_Client_Comm_Date,
						Client_Comm_Date_rebal=re_clientcomm,
						Client_Comm_Date_review=rev_clientcomm,
						Effective_Date=reconst_Effective_Date,
						Effective_Date_rebal=re_effective,
						Effective_Date_review=rev_effective,
						reconst_month=request.POST['reconst_month'],
						rebalance_month=request.POST['rebalance_month'],
						review_month=request.POST['review_month'],
						Comm_to_Calc_Agent=reconst_Client_Comm_Date,
						Comm_to_Calc_Agent_rebal=re_clientcomm,
						Comm_to_Calc_Agent_review=rev_clientcomm,
						color_code=color_1,
						live_date=request.POST['live_date'],
						etf_launch_date=request.POST['etf_date'],
						backtest_comp_date=request.POST['backtest_date'],
					)
			
				index_data.save()
				rule_data = Ruleslist(
					eff_rule = request.POST['effective_date'],
					eff_rule_rebal = request.POST['effective_date_rebal'],
					eff_rule_review = request.POST['effective_date_review'],
					sel_rule1 = request.POST['selec_Date_Cyc_1'],
					sel_rule1_rebal = request.POST['selec_Date_Cyc_1_rebal'],
					sel_rule1_review = request.POST['selec_Date_Cyc_1_review'],
					sel_rule2 = request.POST['selec_Date_Cyc_2'],
					sel_rule2_rebal = request.POST['selec_Date_Cyc_2_rebal'],
					sel_rule2_review = request.POST['selec_Date_Cyc_2_review'],
					announce_rule = request.POST['public_Announcement'],
					announce_rule_rebal = request.POST['public_Announcement_rebal'],
					announce_rule_review = request.POST['public_Announcement_review'],
					prelim_rule = request.POST['pre_comm_date'],
					prelim_rule_rebal = request.POST['pre_comm_date_rebal'],
					prelim_rule_review = request.POST['pre_comm_date_review'],
					clientcomm_rule = request.POST['client_Comm'],
					clientcomm_rule_rebal = request.POST['client_Comm_rebal'],
					clientcomm_rule_review = request.POST['client_Comm_review'],
					indcommittee_rule = request.POST['ind_Comm_Date'],
					indcommittee_rule_rebal = request.POST['ind_Comm_Date_rebal'],
					indcommittee_rule_review = request.POST['ind_Comm_Date_review'],
					freeze_rule = request.POST['weights_Share_Freeze'],
					freeze_rule_rebal = request.POST['weights_Share_Freeze_rebal'],
					freeze_rule_review = request.POST['weights_Share_Freeze_review'],
					comp_rule = request.POST['cmp_Date'],
					comp_rule_rebal = request.POST['cmp_Date_rebal'],
					comp_rule_review = request.POST['cmp_Date_review'],
					comm_cal_rule = request.POST['comm_to_Calc_Agent'],
					comm_cal_rule_rebal = request.POST['comm_to_Calc_Agent_rebal'],
					comm_cal_rule_review = request.POST['comm_to_Calc_Agent_review'],
					index_id = index_data.id,

				)
				rule_data.save()
				template = "postindex.html"
				context = {"message": "success"}
				return render(request, template, context)
	else:
		context = {"message": "index not saved!"}
		template = "index.html"
		return render(request, template, context)
	
		
def view_index(request, id):
    templateName = "ViewIndex.html"
    in_id = Registerindex.objects.filter(id = id)
    context = {
        'in_id' : in_id,
    }
    return render(request, templateName, context)
	
	
def search_index(request):
	if request.method == 'POST':
		search_result = request.POST['search2']
		if(search_result!=""):
			templatename = "SearchIndex.html"
			lookups = Q(Index_Name__icontains = search_result) | Q(Contract_Type__icontains = search_result) | Q(Reconstitution__icontains = search_result) | Q(Calc_Agent__icontains = search_result) | Q(Ind_Sty__icontains = search_result) | Q(Contract_Type__icontains = search_result) | Q(Type_of_Ind__icontains = search_result)
			results= Registerindex.objects.filter(lookups).distinct()

		if(len(results) == 0):
			obj = Calendarlist.objects.filter(description = search_result)
			code = ""
			for ind in obj:
				code = int(ind.code)
				if(code!=""):
					results = Registerindex.objects.filter(Client_Name_id = code)
				else:
					templatename = "SearchIndex.html"

		clientList = Calendarlist.objects.filter(category = 'Client_Name')
		page = request.GET.get('page',1)
		paginator = Paginator(results, 10)
		try:
			results = paginator.page(page)
		except PageNotAnInteger:
			results = paginator.page(1)
		except EmptyPage:
			results = paginator.page(paginator.num_pages)
		context={
			'search_text': search_result,
			'results': results,
			'clientList': clientList,
		}
		return render(request, templatename, context)

	else:
		templatename = "SearchIndex.html"
		return render(request, templatename)
		
		
def report(request):
	templateName = "report.html"
	client_name = Calendarlist.objects.filter(category = 'Client_Name')
	info = Registerindex.objects.all()
	page = request.GET.get('page',1)
	paginator = Paginator(info, 10)
	try:
		infodata = paginator.page(page)
	except PageNotAnInteger:
		infodata = paginator.page(1)
	except EmptyPage:
		infodata = paginator.page(paginator.num_pages)

	context = {
		'client_name': client_name,
		'info': infodata,
	}
	return render(request, templateName, context)
		
def back(request):
    templates = loader.get_template("home.html");
    return HttpResponse(templates.render());


def edit_index(request, id):
	templateName = "EditIndex.html"
	in_id = Registerindex.objects.filter(id=id)
	rule = Ruleslist.objects.filter(index = id)
	client_name = Calendarlist.objects.filter(category='Client_Name')
	ind_Sty = Calendarlist.objects.filter(category='Ind_Sty')
	ind_Version = Calendarlist.objects.filter(category='Ind_Ver')
	cal = Calendarlist.objects.filter(category='Cal')
	cal_Agent = Calendarlist.objects.filter(category='Cal_Agent')
	contract_Type = Calendarlist.objects.filter(category='Contract_Type')
	type_of_Ind = Calendarlist.objects.filter(category='Type_of_Ind')
	prod_Stat = Calendarlist.objects.filter(category='Prod_Stat')
	etf_Launched = Calendarlist.objects.filter(category='ETF_Launched')
	reconstitution = Calendarlist.objects.filter(category='Reconstitution')
	rebalance = Calendarlist.objects.filter(category='Rebalance')
	review = Calendarlist.objects.filter(category='Review')
	theme_Review = Calendarlist.objects.filter(category='Theme_Review')
	comp_Date = Calendarlist.objects.filter(category='Comp_Date')
	selec_Date_Cyc_1 = Calendarlist.objects.filter(category='Selec_Date_Cyc_1')
	selec_Date_Cyc_2 = Calendarlist.objects.filter(category='Selec_Date_Cyc_2')
	weights_Share_Freeze = Calendarlist.objects.filter(category='Weights_Share_Freeze_Date')
	public_Announcement = Calendarlist.objects.filter(category='Public_Announcement_Date')
	client_Comm = Calendarlist.objects.filter(category='Client_Comm_Date')
	ind_Cmte_Comm_Date = Calendarlist.objects.filter(category='Ind_Cmte_Comm_Date')
	effec_Date = Calendarlist.objects.filter(category='Effec_Date')
	prelim_Comm_Date = Calendarlist.objects.filter(category='Prelim_Comm_Date')
	comm_to_Calc_Agent = Calendarlist.objects.filter(category='Comm_to_Calc_Agent')
	context = {
		'in_id': in_id,
		'client_name': client_name,
		'ind_Sty': ind_Sty,
		'ind_Version': ind_Version,
		'cal': cal,
		'cal_Agent': cal_Agent,
		'contract_Type': contract_Type,
		'type_of_Ind': type_of_Ind,
		'prod_Stat': prod_Stat,
		'etf_Launched': etf_Launched,
		'reconstitution': reconstitution,
		'rebalance': rebalance,
		'review': review,
		'theme_Review': theme_Review,
		'comp_Date': comp_Date,
		'selec_Date_Cyc_1': selec_Date_Cyc_1,
		'selec_Date_Cyc_2': selec_Date_Cyc_2,
		'weights_Share_Freeze': weights_Share_Freeze,
		'public_Announcement': public_Announcement,
		'client_Comm': client_Comm,
		'ind_Cmte_Comm_Date': ind_Cmte_Comm_Date,
		'effec_Date': effec_Date,
		'prelim_Comm_Date': prelim_Comm_Date,
		'comm_to_Calc_Agent': comm_to_Calc_Agent,
		'rule':rule,
	}

	return render(request, templateName, context)



def update_index(request, id):
    in_id = Registerindex.objects.filter(id=id)
    index1 = Ruleslist.objects.filter(index = id)
    clientList = Calendarlist.objects.filter(category='Client_Name')
    for ind in in_id:
        if request.method == 'POST':
            if (request.POST['cmp_Date'] != "" or request.POST['ind_Comm_Date'] != "" or request.POST[
                'client_Comm'] != "" or request.POST['public_Announcement'] != "" or request.POST[
                'effective_date'] != "" or
                    request.POST['pre_comm_date'] != "" or request.POST['comm_to_Calc_Agent'] != "" or request.POST[
                        'selec_Date_Cyc_2'] != "" or request.POST['selec_Date_Cyc_1'] != "" or request.POST[
                        'weights_Share_Freeze'] != ""):
              if(request.POST['reconstitution']=="Annual" or request.POST['reconstitution']=="Monthly"):
                    num_month = 1
                    gap=1
              elif(request.POST['reconstitution']== "Quarterly"):
                    num_month = 4
                    gap = 3
              else:
                    num_month = 2
                    gap = 6

              reconst_Effective_Date = ""
              reconst_Selection_Date_Cycle_2 = ""
              reconst_Weights_Share_Freeze_Date = ""
              reconst_Public_Announcement_Date = ""
              reconst_Client_Comm_Date = ""
              reconst_Prelim_Comm_Date = ""
              reconst_Completion_Date = ""
              reconst_Ind_Cmte_Comm_Date = ""
              reconst_Selection_Date_Cycle_1 = ""
              rev_effective = ""
              rev_selection2 = ""
              rev_freeze = ""
              rev_announce = ""
              rev_clientcomm = ""
              rev_prelim = ""
              rev_completion = ""
              rev_commiteecomm = ""
              rev_selection1 = ""
              re_effective = ""
              re_selection2 = ""
              re_freeze = ""
              re_announce = ""
              re_clientcomm = ""
              re_prelim = ""
              re_completion = ""
              re_commiteecomm = ""
              re_selection1 = ""

              months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
                              October=10, November=11, December=12)
              month1 = months[request.POST['reconst_month']]
              month1 = int(month1)
              cal_month = month1
              year = date.today().year
              for i in range(0,num_month):
                if(month1>12):
                     month1 = month1%12;
                if(month1<cal_month):
                     year = year+1
                eff_rule = request.POST['effective_date']
                sel_rule = request.POST['selec_Date_Cyc_2']
                fre_rule = request.POST['weights_Share_Freeze']
                ann_rule = request.POST['public_Announcement']
                client_comm_rule = request.POST['client_Comm']
                prelim_comm_rule = request.POST['pre_comm_date']
				#prelim_comm_rule = int(request.POST['pre_comm_date'])
                color_1 = 1
                date1 = EffectiveDate_Calculate(month1, year, eff_rule)
                eff_day = int(date1[3:5])
                eff_mon = int(date1[0:2])
                eff_year = int(date1[6:8])
                sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule)
                sel_day = int(sel_date[3:5])
                sel_month = int(sel_date[0:2])
                sel_year = int(sel_date[6:8])
                freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule)
                freeze_day = int(freeze_date[3:5])
                freeze_mon = int(freeze_date[0:2])
                freeze_year = int(freeze_date[6:8])
                if (ann_rule == 1):
                    announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule)
                else:
                    announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule)
                announce_day = int(announce_date[3:5])
                announce_mon = int(announce_date[0:2])
                announce_year = int(announce_date[6:8])
                if (client_comm_rule == 1):
                    client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule)
                else:
                    client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule)
                Client_comm_day = int(client_comm_date[3:5])
                Client_comm_mon = int(client_comm_date[0:2])
                Client_comm_year = int(client_comm_date[6:8])
                Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day, announce_mon, announce_year)
                Prelim_comm_day = int(Prelim_comm_date[3:5])
                Prelim_comm_mon = int(Prelim_comm_date[0:2])
                Prelim_comm_year = int(Prelim_comm_date[6:8])
                Completion_date = CompletionDate_Calculate(sel_day, sel_month, sel_year)
                if (prelim_comm_rule == ""):
                    Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon,
                                                                       Client_comm_year)
                else:
                    Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day, Prelim_comm_mon,
                                                                       Prelim_comm_year)

                selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year)
                cal_month = month1
                month1= month1+gap
                reconst_Effective_Date = reconst_Effective_Date + date1 + ", "
                reconst_Selection_Date_Cycle_2 = reconst_Selection_Date_Cycle_2 + sel_date + ", "
                reconst_Weights_Share_Freeze_Date = reconst_Weights_Share_Freeze_Date + freeze_date + ", "
                reconst_Public_Announcement_Date = reconst_Public_Announcement_Date + announce_date + ", "
                reconst_Client_Comm_Date = reconst_Client_Comm_Date + client_comm_date + ", "
                reconst_Prelim_Comm_Date = reconst_Prelim_Comm_Date + Prelim_comm_date + ", "
                reconst_Completion_Date = reconst_Completion_Date + Completion_date + ", "
                reconst_Ind_Cmte_Comm_Date = reconst_Ind_Cmte_Comm_Date + Committee_comm_date + ", "
                reconst_Selection_Date_Cycle_1 = reconst_Selection_Date_Cycle_1+selection_date1+ ", "

              if (request.POST['review_month'] != ""):

                  eff_rule_rev = request.POST['effective_date_review']
                  sel_rule_rev = request.POST['selec_Date_Cyc_2_review']
                  fre_rule_rev = request.POST['weights_Share_Freeze_review']
                  ann_rule_rev = request.POST['public_Announcement_review']
                  client_comm_rule_rev = request.POST['client_Comm_review']
                  prelim_comm_rule_rev = request.POST['pre_comm_date_review']
				  #prelim_comm_rule_rev = int(request.POST['pre_comm_date_review'])

                  if(request.POST['review']=="Annual" or request.POST['review']=="Monthly"):
                    num_month = 1
                    gap=1
                  elif(request.POST['review']== "Quarterly"):
                    num_month = 4
                    gap = 3
                  else:
                    num_month = 2
                    gap = 6
                  months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
                              October=10, November=11, December=12)
                  month1 = months[request.POST['review_month']]
                  month1 = int(month1)
                  cal_month = month1
                  year = date.today().year
                  for i in range(0,num_month):
                    if(month1>12):
                       month1 = month1%12;
                    if(month1<cal_month):
                       year = year+1
                    date1 = EffectiveDate_Calculate(month1, year, eff_rule_rev)
                    eff_day = int(date1[3:5])
                    eff_mon = int(date1[0:2])
                    eff_year = int(date1[6:8])

                    sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule_rev)
                    sel_day = int(sel_date[3:5])
                    sel_month = int(sel_date[0:2])
                    sel_year = int(sel_date[6:8])

                    freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule_rev)
                    freeze_day = int(freeze_date[3:5])
                    freeze_mon = int(freeze_date[0:2])
                    freeze_year = int(freeze_date[6:8])

                    if (ann_rule_rev == 1):
                        announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rev)
                    else:
                        announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rev)
                    announce_day = int(announce_date[3:5])
                    announce_mon = int(announce_date[0:2])
                    announce_year = int(announce_date[6:8])

                    if (client_comm_rule_rev == 1):
                        client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rev)
                    else:
                        client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rev)
                    Client_comm_day = int(client_comm_date[3:5])
                    Client_comm_mon = int(client_comm_date[0:2])
                    Client_comm_year = int(client_comm_date[6:8])

                    Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day, announce_mon, announce_year)
                    Prelim_comm_day = int(Prelim_comm_date[3:5])
                    Prelim_comm_mon = int(Prelim_comm_date[0:2])
                    Prelim_comm_year = int(Prelim_comm_date[6:8])

                    Completion_date = CompletionDate_Calculate(sel_day, sel_month, sel_year)
                    if (prelim_comm_rule_rev == ""):
                        Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon,
                                                                           Client_comm_year)
                    else:
                        Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day, Prelim_comm_mon,
                                                                           Prelim_comm_year)
                    selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year)
                    cal_month = month1
                    month1= month1+gap
                    rev_effective = rev_effective + date1 + ", "
                    rev_selection2 = rev_selection2 + sel_date + ", "
                    rev_freeze = rev_freeze + freeze_date + ", "
                    rev_announce = rev_announce + announce_date + ", "
                    rev_clientcomm = rev_clientcomm + client_comm_date + ", "
                    rev_prelim = rev_prelim + Prelim_comm_date + ", "
                    rev_completion = rev_completion + Completion_date + ", "
                    rev_commiteecomm = rev_commiteecomm + Committee_comm_date + ", "
                    rev_selection1 = rev_selection1 + selection_date1 + ", "

              if (request.POST['rebalance_month'] != ""):

                  eff_rule_rebal = request.POST['effective_date_rebal']
                  sel_rule_rebal = request.POST['selec_Date_Cyc_2_rebal']
                  fre_rule_rebal = request.POST['weights_Share_Freeze_rebal']
                  ann_rule_rebal = request.POST['public_Announcement_rebal']
                  client_comm_rule_rebal = request.POST['client_Comm_rebal']
                  prelim_comm_rule_rebal = request.POST['pre_comm_date_rebal']
				  #prelim_comm_rule_rebal = int(request.POST['pre_comm_date_rebal'])
                  if(request.POST['rebalance']=="Annual" or request.POST['rebalance']=="Monthly"):
                    num_month = 1
                    gap=1
                  elif(request.POST['rebalance']== "Quarterly"):
                    num_month = 4
                    gap = 3
                  else:
                    num_month = 2
                    gap = 6

                  months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
                              October=10, November=11, December=12)
                  month1 = months[request.POST['rebalance_month']]
                  month1 = int(month1)
                  cal_month = month1
                  year = date.today().year
                  for i in range(0,num_month):
                    if(month1>12):
                       month1 = month1%12;
                    if(month1<cal_month):
                       year = year+1

                    date1 = EffectiveDate_Calculate(month1, year, eff_rule_rebal)
                    eff_day = int(date1[3:5])
                    eff_mon = int(date1[0:2])
                    eff_year = int(date1[6:8])
                    sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule_rebal)
                    sel_day = int(sel_date[3:5])
                    sel_month = int(sel_date[0:2])
                    sel_year = int(sel_date[6:8])
                    freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule_rebal)
                    freeze_day = int(freeze_date[3:5])
                    freeze_mon = int(freeze_date[0:2])
                    freeze_year = int(freeze_date[6:8])

                    if (ann_rule_rebal == 1):
                        announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rebal)
                    else:
                        announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rebal)
                    announce_day = int(announce_date[3:5])
                    announce_mon = int(announce_date[0:2])
                    announce_year = int(announce_date[6:8])

                    if (client_comm_rule_rebal == 1):
                        client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rebal)
                    else:
                        client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rebal)
                    Client_comm_day = int(client_comm_date[3:5])
                    Client_comm_mon = int(client_comm_date[0:2])
                    Client_comm_year = int(client_comm_date[6:8])

                    Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day, announce_mon, announce_year)
                    Prelim_comm_day = int(Prelim_comm_date[3:5])
                    Prelim_comm_mon = int(Prelim_comm_date[0:2])
                    Prelim_comm_year = int(Prelim_comm_date[6:8])

                    Completion_date = CompletionDate_Calculate(sel_day, sel_month, sel_year)

                    if (prelim_comm_rule_rebal == ""):
                        Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon,
                                                                           Client_comm_year)
                    else:
                        Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day, Prelim_comm_mon,
                                                                           Prelim_comm_year)
                    selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year)
                    cal_month = month1
                    month1= month1+gap
                    re_effective = re_effective + date1 + ", "
                    re_selection2 = re_selection2 + sel_date + ", "
                    re_freeze = re_freeze + freeze_date + ", "
                    re_announce = re_announce + announce_date + ", "
                    re_clientcomm = re_clientcomm + client_comm_date + ", "
                    re_prelim = re_prelim + Prelim_comm_date + ", "
                    re_completion = re_completion + Completion_date + ", "
                    re_commiteecomm = re_commiteecomm + Committee_comm_date + ", "
                    re_selection1 = re_selection1 + selection_date1 + ", "

            else:

                color_1 = 0
                reconst_Effective_Date = request.POST['mnl_eff_date']
                reconst_Selection_Date_Cycle_2 = request.POST['mnl_sel2_date']
                reconst_Weights_Share_Freeze_Date = request.POST['mnl_freeze_date']
                reconst_Public_Announcement_Date = request.POST['mnl_pb_announce_date']
                reconst_Client_Comm_Date = request.POST['mnl_client_comm_date']
                reconst_Prelim_Comm_Date = request.POST['mnl_prelim_date']
                reconst_Completion_Date = request.POST['mnl_cmp_date']
                reconst_Ind_Cmte_Comm_Date = request.POST['mnl_ind_Comm_date']
                reconst_Selection_Date_Cycle_1 = request.POST['mnl_sel1_date']
                rev_effective = request.POST['mnl_eff_date_review']
                rev_selection1 = request.POST['mnl_sel1_date_review']
                rev_announce = request.POST['mnl_pb_announce_date_review']
                rev_clientcomm = request.POST['mnl_client_comm_date_review']
                rev_commiteecomm = request.POST['mnl_ind_Comm_date_review']
                rev_completion = request.POST['mnl_cmp_date_review']
                rev_prelim = request.POST['mnl_prelim_date_review']
                rev_freeze = request.POST['mnl_freeze_date_review']
                rev_selection2 = request.POST['mnl_sel2_date_review']
                re_announce = request.POST['mnl_pb_announce_date_rebal']
                re_clientcomm = request.POST['mnl_client_comm_date_rebal']
                re_commiteecomm = request.POST['mnl_ind_Comm_date_rebal']
                re_completion = request.POST['mnl_cmp_date_rebal']
                re_effective = request.POST['mnl_eff_date_rebal']
                re_freeze = request.POST['mnl_freeze_date_rebal']
                re_prelim = request.POST['mnl_prelim_date_rebal']
                re_selection1 = request.POST['mnl_sel1_date_rebal']
                re_selection2 = request.POST['mnl_sel2_date_rebal']

            ind.Ident_ISIN=request.POST['isin_id']
            ind.Ident_Bloomberg=request.POST['bloomberg_id']
            ind.Ident_Reuters=request.POST['thomson_id']
            ind.Index_Name=request.POST['index_name']
            ind.Client_Name_id=request.POST['client_name']
            ind.Ind_Sty=request.POST['index_Style']
            ind.Ind_Ver=request.POST['ind_version']
            ind.Ind_Ver_ID=request.POST['index_version_id']
            ind.Calc=request.POST['calculation']
            ind.Calc_Agent=request.POST['cal_agent']
            ind.Data_Platform=request.POST['data_platform']
            ind.Data_Vendors=request.POST['data_vendors']
            ind.Contract_Type=request.POST['contract_Type']
            ind.Type_of_Ind=request.POST['type_index']
            ind.Product_Status=request.POST['prod_Status']
            ind.ETF_Launched=request.POST['etf_Launched']
            ind.Reconstitution=request.POST['reconstitution']
            ind.Rebalance=request.POST['rebalance']
            ind.Review=request.POST['review']
            ind.Theme_Review=request.POST['theme_Review']
            ind.Selection_Date_Cycle_1=reconst_Selection_Date_Cycle_1
            ind.Selection_Date_Cycle_1_rebal=re_selection1
            ind.Selection_Date_Cycle_1_review= rev_selection1
            ind.Completion_Date=reconst_Completion_Date
            ind.Completion_Date_rebal=re_completion
            ind.Completion_Date_review=rev_completion
            ind.Selection_Date_Cycle_2=reconst_Selection_Date_Cycle_2
            ind.Selection_Date_Cycle_2_rebal=re_selection2
            ind.Selection_Date_Cycle_2_review=rev_selection2
            ind.Ind_Cmte_Comm_Date=reconst_Ind_Cmte_Comm_Date
            ind.Ind_Cmte_Comm_Date_rebal=re_commiteecomm
            ind.Ind_Cmte_Comm_Date_review=rev_commiteecomm
            ind.Prelim_Comm_Date=reconst_Prelim_Comm_Date
            ind.Prelim_Comm_Date_rebal=re_prelim
            ind.Prelim_Comm_Date_review=rev_prelim
            ind.Weights_Share_Freeze_Date=reconst_Weights_Share_Freeze_Date
            ind.Weights_Share_Freeze_Date_rebal=re_freeze
            ind.Weights_Share_Freeze_Date_review=rev_freeze
            ind.Public_Announcement_Date=reconst_Public_Announcement_Date
            ind.Public_Announcement_Date_rebal= re_announce
            ind.Public_Announcement_Date_review=rev_announce
            ind.Client_Comm_Date=reconst_Client_Comm_Date
            ind.Client_Comm_Date_rebal=re_clientcomm
            ind.Client_Comm_Date_review=rev_clientcomm
            ind.Effective_Date=reconst_Effective_Date
            ind.Effective_Date_rebal=re_effective
            ind.Effective_Date_review=rev_effective
            ind.reconst_month=request.POST['reconst_month']
            ind.rebalance_month=request.POST['rebalance_month']
            ind.review_month=request.POST['review_month']
            ind.Comm_to_Calc_Agent=reconst_Client_Comm_Date
            ind.Comm_to_Calc_Agent_rebal=re_clientcomm
            ind.Comm_to_Calc_Agent_review=rev_clientcomm
            ind.color_code=color_1
            ind.live_date=request.POST['live_date']
            ind.etf_launch_date=request.POST['etf_date']
            ind.backtest_comp_date=request.POST['backtest_date']

            ind.save()

    for indrule in index1:
        if request.method == 'POST':
                indrule.eff_rule = request.POST['effective_date']
                indrule.eff_rule_rebal = request.POST['effective_date_rebal']
                indrule.eff_rule_review = request.POST['effective_date_review']
                indrule.sel_rule1 = request.POST['selec_Date_Cyc_1']
                indrule.sel_rule1_rebal = request.POST['selec_Date_Cyc_1_rebal']
                indrule.sel_rule1_review = request.POST['selec_Date_Cyc_1_review']
                indrule.sel_rule2 = request.POST['selec_Date_Cyc_2']
                indrule.sel_rule2_rebal = request.POST['selec_Date_Cyc_2_rebal']
                indrule.sel_rule2_review = int(request.POST['selec_Date_Cyc_2_review'])
                indrule.announce_rule = request.POST['public_Announcement']
                indrule.announce_rule_rebal = request.POST['public_Announcement_rebal']
                indrule.announce_rule_review = request.POST['public_Announcement_review']
                indrule.prelim_rule = request.POST['pre_comm_date']
                indrule.prelim_rule_rebal = request.POST['pre_comm_date_rebal']
                indrule.prelim_rule_review = request.POST['pre_comm_date_review']
                indrule.clientcomm_rule = request.POST['client_Comm']
                indrule.clientcomm_rule_rebal = request.POST['client_Comm_rebal']
                indrule.clientcomm_rule_review = request.POST['client_Comm_review']
                indrule.indcommittee_rule = request.POST['ind_Comm_Date']
                indrule.indcommittee_rule_rebal = request.POST['ind_Comm_Date_rebal']
                indrule.indcommittee_rule_review = request.POST['ind_Comm_Date_review']
                indrule.freeze_rule = request.POST['weights_Share_Freeze']
                indrule.freeze_rule_rebal = request.POST['weights_Share_Freeze_rebal']
                indrule.freeze_rule_review = request.POST['weights_Share_Freeze_review']
                indrule.comp_rule = request.POST['cmp_Date']
                indrule.comp_rule_rebal = request.POST['cmp_Date_rebal']
                indrule.comp_rule_review = request.POST['cmp_Date_review']
                indrule.comm_cal_rule = request.POST['comm_to_Calc_Agent']
                indrule.comm_cal_rule_rebal = request.POST['comm_to_Calc_Agent_rebal']
                indrule.comm_cal_rule_review = int(request.POST['comm_to_Calc_Agent_review'])
                indrule.save()

    templates = loader.get_template("DeleteIndex.html")
    return HttpResponse(templates.render())




def delete_index(request, id):
	Registerindex.objects.filter(id = id).delete()
	templates = loader.get_template("DeleteIndex.html")
	return HttpResponse(templates.render());
"""
def report_generate(request):
	client_name = Calendarlist.objects.filter(category = 'Client_Name')
	info = Registerindex.objects.all()
	if request.method == 'POST':
		cname1 = request.POST['Client_Code']
		month = request.POST['month1']
		month2 = request.POST['month2']
		sel_type = request.POST['sel_type']
		if(cname1 == "" and month =="" and month2 == ""):
			
			context = {
				'info' : info,
				'client_name' : client_name,

			}
			
		else:
			year = month[:4]
			
			#print(year)
			#pdb.set_trace()
			calendardata = Calendarlist.objects.all()
			if(month==""):
				year = date.today().year
			
			if(cname1 == ""):
				info = Registerindex.objects.all()
			else:
				info = Registerindex.objects.filter(Client_Name = cname1)
				
			informa=[]	
			months = dict(January=1, February=2, March=3, April = 4, May=5, June=6, July=7, August=8,September=9, October=10, November=11, December=12)
			for ind in info:
				month1 = months[ind.reconst_month]
				rulelist = Ruleslist.objects.get(index_id=ind.id)
				if(rulelist.eff_rule is not None):
					eff_rule = rulelist.eff_rule
					sel_rule = rulelist.sel_rule1
					sel_rule2 = rulelist.sel_rule2
					fre_rule = rulelist.freeze_rule
					ann_rule = rulelist.announce_rule
					client_comm_rule = rulelist.clientcomm_rule
					Prelim_comm_rule = rulelist.prelim_rule
					month1 = int(month1)
					year = int(year)
					
					if(month is not None):
						from_month = int(month[5:7])
						#to_month = int(month2[5:7])
						if(sel_type =="effective"):
							if eff_rule:
								date1 = EffectiveDate_Calculate(month1,year,eff_rule)
								eff_day = int(date1[3:5])
								eff_month = int(date1[0:2])
								#ind.Effective_Date = date1
								
							if from_month <= eff_month:
								if eff_rule:
									date1 = EffectiveDate_Calculate(month1,year,eff_rule)
									eff_day = int(date1[3:5])
									eff_month = int(date1[0:2])
									ind.Effective_Date = date1
									
								if sel_rule:
									sel_date = SelectionDate_Calculate(eff_day,month1,year,sel_rule)
									sel_day = int(sel_date[3:5])
									sel_month = int(sel_date[0:2])
									ind.Selection_Date_Cycle_2 = sel_date
																	
								if fre_rule:
									freeze_date = FreezeDate_Calculate(eff_day,month1,year,fre_rule)
									freeze_day = int(freeze_date[3:5])
									freeze_month = int(freeze_date[0:2])
									ind.Weights_Share_Freeze_Date = freeze_date
									
								if(ann_rule==1):
									announce_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
								else:
									announce_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
								announce_day = int(announce_date[3:5])
								announce_month = int(announce_date[0:2])
								ind.Public_Announcement_Date = announce_date

								if(client_comm_rule==1):
									client_comm_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
								else:
									client_comm_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
								Client_comm_day = int(client_comm_date[3:5])
								Client_comm_month = int(client_comm_date[0:2])
								ind.Client_Comm_Date = client_comm_date

								Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,month1,year)
								Prelim_comm_day = int(Prelim_comm_date[3:5])
								Prelim_comm_month = int(Prelim_comm_date[0:2])
								ind.Prelim_Comm_Date = Prelim_comm_date

								Completion_date = CompletionDate_Calculate(sel_day,month1,year)
								Completion_month = int(Completion_date[0:2])
								ind.Completion_Date = Completion_date
								if(Prelim_comm_rule==""):
									Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,month1,year)
								else:
									Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,month1,year)

								ind.Ind_Cmte_Comm_Date = Committee_comm_date
								Committee_comm_month = int(Committee_comm_date[0:2])

								selection_date1 = selection_date_1_Calculate(sel_day,sel_month,year)
								sel_1_month = int(selection_date1[0:2])
								ind.Selection_Date_Cycle_1 = selection_date1
								data = {}
								informa.append(ind)
					
						elif(sel_type =="selection"):
							if eff_rule:
								date1 = EffectiveDate_Calculate(month1,year,eff_rule)
								eff_day = int(date1[3:5])
								eff_month = int(date1[0:2])
								
									
							if sel_rule:
								sel_date = SelectionDate_Calculate(eff_day,month1,year,sel_rule)
								sel_day = int(sel_date[3:5])
								sel_month = int(sel_date[0:2])
								
								
							if(from_month <= sel_month):
								if eff_rule:
									date1 = EffectiveDate_Calculate(month1,year,eff_rule)
									eff_day = int(date1[3:5])
									eff_month = int(date1[0:2])
									ind.Effective_Date = date1
								
								if sel_rule:
									sel_date = SelectionDate_Calculate(eff_day,month1,year,sel_rule)
									sel_day = int(sel_date[3:5])
									sel_month = int(sel_date[0:2])
									ind.Selection_Date_Cycle_2 = sel_date
									
								if fre_rule:
									freeze_date = FreezeDate_Calculate(eff_day,month1,year,fre_rule)
									freeze_day = int(freeze_date[3:5])
									freeze_month = int(freeze_date[0:2])
									ind.Weights_Share_Freeze_Date = freeze_date
									
								if(ann_rule==1):
									announce_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
								else:
									announce_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
								announce_day = int(announce_date[3:5])
								announce_month = int(announce_date[0:2])
								ind.Public_Announcement_Date = announce_date

								if(client_comm_rule==1):
									client_comm_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
								else:
									client_comm_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
								Client_comm_day = int(client_comm_date[3:5])
								Client_comm_month = int(client_comm_date[0:2])
								ind.Client_Comm_Date = client_comm_date

								Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,month1,year)
								Prelim_comm_day = int(Prelim_comm_date[3:5])
								Prelim_comm_month = int(Prelim_comm_date[0:2])
								ind.Prelim_Comm_Date = Prelim_comm_date

								Completion_date = CompletionDate_Calculate(sel_day,month1,year)
								Completion_month = int(Completion_date[0:2])
								ind.Completion_Date = Completion_date
								if(Prelim_comm_rule==""):
									Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,month1,year)
								else:
									Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,month1,year)

								ind.Ind_Cmte_Comm_Date = Committee_comm_date
								Committee_comm_month = int(Committee_comm_date[0:2])

								selection_date1 = selection_date_1_Calculate(sel_day,sel_month,year)
								sel_1_month = int(selection_date1[0:2])
								ind.Selection_Date_Cycle_1 = selection_date1
								data = {}
								informa.append(ind)
							else:
								informa=[]
					
						elif(sel_type =="icom"):
							if eff_rule:
								date1 = EffectiveDate_Calculate(month1,year,eff_rule)
								eff_day = int(date1[3:5])
								eff_month = int(date1[0:2])
								ind.Effective_Date = date1
								
							if sel_rule:
								sel_date = SelectionDate_Calculate(eff_day,month1,year,sel_rule)
								sel_day = int(sel_date[3:5])
								sel_month = int(sel_date[0:2])
								ind.Selection_Date_Cycle_2 = sel_date
								
							if fre_rule:
								freeze_date = FreezeDate_Calculate(eff_day,month1,year,fre_rule)
								freeze_day = int(freeze_date[3:5])
								freeze_month = int(freeze_date[0:2])
								ind.Weights_Share_Freeze_Date = freeze_date
								
							if(ann_rule==1):
								announce_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
							else:
								announce_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
							announce_day = int(announce_date[3:5])
							announce_month = int(announce_date[0:2])
							ind.Public_Announcement_Date = announce_date

							if(client_comm_rule==1):
								client_comm_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
							else:
								client_comm_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
							Client_comm_day = int(client_comm_date[3:5])
							Client_comm_month = int(client_comm_date[0:2])
							ind.Client_Comm_Date = client_comm_date

							Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,month1,year)
							Prelim_comm_day = int(Prelim_comm_date[3:5])
							Prelim_comm_month = int(Prelim_comm_date[0:2])
							ind.Prelim_Comm_Date = Prelim_comm_date

							Completion_date = CompletionDate_Calculate(sel_day,month1,year)
							Completion_month = int(Completion_date[0:2])
							ind.Completion_Date = Completion_date
							if(Prelim_comm_rule==""):
								Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,month1,year)
							else:
								Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,month1,year)

							ind.Ind_Cmte_Comm_Date = Committee_comm_date
							Committee_comm_month = int(Committee_comm_date[0:2])
							
							if from_month <= Committee_comm_month:
								if eff_rule:
									date1 = EffectiveDate_Calculate(month1,year,eff_rule)
									eff_day = int(date1[3:5])
									eff_month = int(date1[0:2])
									ind.Effective_Date = date1
									
								if sel_rule:
									sel_date = SelectionDate_Calculate(eff_day,month1,year,sel_rule)
									sel_day = int(sel_date[3:5])
									sel_month = int(sel_date[0:2])
									ind.Selection_Date_Cycle_2 = sel_date
									
								if fre_rule:
									freeze_date = FreezeDate_Calculate(eff_day,month1,year,fre_rule)
									freeze_day = int(freeze_date[3:5])
									freeze_month = int(freeze_date[0:2])
									ind.Weights_Share_Freeze_Date = freeze_date
									
								if(ann_rule==1):
									announce_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
								else:
									announce_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
								announce_day = int(announce_date[3:5])
								announce_month = int(announce_date[0:2])
								ind.Public_Announcement_Date = announce_date

								if(client_comm_rule==1):
									client_comm_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
								else:
									client_comm_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
								Client_comm_day = int(client_comm_date[3:5])
								Client_comm_month = int(client_comm_date[0:2])
								ind.Client_Comm_Date = client_comm_date

								Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,month1,year)
								Prelim_comm_day = int(Prelim_comm_date[3:5])
								Prelim_comm_month = int(Prelim_comm_date[0:2])
								ind.Prelim_Comm_Date = Prelim_comm_date

								Completion_date = CompletionDate_Calculate(sel_day,month1,year)
								Completion_month = int(Completion_date[0:2])
								ind.Completion_Date = Completion_date
								if(Prelim_comm_rule==""):
									Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,month1,year)
								else:
									Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,month1,year)

								ind.Ind_Cmte_Comm_Date = Committee_comm_date
								Committee_comm_month = int(Committee_comm_date[0:2])

								selection_date1 = selection_date_1_Calculate(sel_day,sel_month,year)
								sel_1_month = int(selection_date1[0:2])
								ind.Selection_Date_Cycle_1 = selection_date1
								data = {}
								informa.append(ind)
					
					else:
						if eff_rule:
							date1 = EffectiveDate_Calculate(month1,year,eff_rule)
							eff_day = int(date1[3:5])
							eff_month = int(date1[0:2])
							ind.Effective_Date = date1
								
								
						if sel_rule:
							sel_date = SelectionDate_Calculate(eff_day,month1,year,sel_rule)
							sel_day = int(sel_date[3:5])
							sel_month = int(sel_date[0:2])
							ind.Selection_Date_Cycle_1 = sel_date
							
							
						if fre_rule:
							freeze_date = FreezeDate_Calculate(eff_day,month1,year,fre_rule)
							freeze_day = int(freeze_date[3:5])
							freeze_month = int(freeze_date[0:2])
							ind.Weights_Share_Freeze_Date = freeze_date
							
						if(ann_rule==1):
							announce_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
						else:
							announce_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
						announce_day = int(announce_date[3:5])
						announce_month = int(announce_date[0:2])
						ind.Public_Announcement_Date = announce_date

						if(client_comm_rule==1):
							client_comm_date = AnnouncementDate_Calculate(eff_day,month1,year,ann_rule)
						else:
							client_comm_date = AnnouncementDate_Calculate(freeze_day,month1,year,ann_rule)
						Client_comm_day = int(client_comm_date[3:5])
						Client_comm_month = int(client_comm_date[0:2])
						ind.Client_Comm_Date = client_comm_date

						Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,month1,year)
						Prelim_comm_day = int(Prelim_comm_date[3:5])
						Prelim_comm_month = int(Prelim_comm_date[0:2])
						ind.Prelim_Comm_Date = Prelim_comm_date

						Completion_date = CompletionDate_Calculate(sel_day,month1,year)
						Completion_month = int(Completion_date[0:2])
						ind.Completion_Date = Completion_date
						if(Prelim_comm_rule==""):
							Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,month1,year)
						else:
							Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,month1,year)

						ind.Ind_Cmte_Comm_Date = Committee_comm_date
						Committee_comm_month = int(Committee_comm_date[0:2])

						selection_date1 = selection_date_1_Calculate(sel_day,sel_month,year)
						sel_1_month = int(selection_date1[0:2])
						ind.Selection_Date_Cycle_1 = selection_date1
						data = {}
						informa.append(ind)
					
					
					
				else:
					informa.append(ind)
					
			context = {
				'info' : informa,
				'client_name' : client_name,
				'client_code' : cname1,
				'from_MY' : month,
				'to_MY' : month2,
				'sel_type' : sel_type,


			}	
				

	
	else:
		context = {
			'info' : info,
			'client_name' : client_name,
			
		}
	templateName = "reportGenerate.html"
	return render(request, templateName, context)
"""

def report_generate(request):
	client_name = Calendarlist.objects.filter(category = 'Client_Name')
	info = Registerindex.objects.all()
	if request.method == 'POST':
		cname1 = request.POST['Client_Code']
		month = request.POST['month1']
		month2 = request.POST['month2']
		sel_type = request.POST['sel_type']
		if(cname1 == "" and month =="" and month2 == ""):

			context = {
				'info' : info,
				'client_name' : client_name,

			}

		else:
			year = month[:4]

			#print(year)
			#pdb.set_trace()
			calendardata = Calendarlist.objects.all()
			#client_name = Calendarlist.objects.filter(category = 'Client_Name')
			if(month==""):
				year = date.today().year

			if(cname1 == ""):
				info = Registerindex.objects.all()
			else:
				info = Registerindex.objects.filter(Client_Name = cname1)

			informa=[]
			months = dict(January=1, February=2, March=3, April = 4, May=5, June=6, July=7, August=8,September=9, October=10, November=11, December=12)
			for ind in info:
				
				#print(type(ind.reconst_month))
				#pdb.set_trace()
				if(month=="" and month2==""):
					informa.append(ind)
				elif(ind.reconst_month is None or ind.rebalance_month is None):
					informa.append(ind)
				else:
					month1 = months[ind.reconst_month]
					month1_rebal = months[ind.rebalance_month]
					
					rulelist = Ruleslist.objects.get(index_id=ind.id)
					if(rulelist.eff_rule is not None):
						eff_rule = rulelist.eff_rule
						sel_rule = rulelist.sel_rule1
						sel_rule2 = rulelist.sel_rule2
						fre_rule = rulelist.freeze_rule
						ann_rule = rulelist.announce_rule
						client_comm_rule = rulelist.clientcomm_rule
						Prelim_comm_rule = rulelist.prelim_rule

						month1 = int(month1)
						year = int(year)
						
					if(rulelist.eff_rule_rebal is not None):
						eff_rule_rebal = rulelist.eff_rule_rebal
						sel_rule_rebal = rulelist.sel_rule1_rebal
						sel_rule2_rebal = rulelist.sel_rule2_rebal
						fre_rule_rebal = rulelist.freeze_rule_rebal
						ann_rule_rebal = rulelist.announce_rule_rebal
						client_comm_rule_rebal = rulelist.clientcomm_rule_rebal
						Prelim_comm_rule_rebal = rulelist.prelim_rule_rebal

						month1_rebal = int(month1_rebal)
						year = int(year)

						if(month!=""):
							from_month = int(month[5:7])
							#to_month = int(month2[5:7])
							if(sel_type =="effective"):
								if eff_rule:
									date1 = EffectiveDate_Calculate(month1,year,eff_rule)
									eff_day = int(date1[3:5])
									eff_month = int(date1[0:2])
									#ind.Effective_Date = date1
								if eff_rule_rebal:
									date1_rebal = EffectiveDate_Calculate(month1_rebal,year,eff_rule_rebal)
									eff_day_rebal = int(date1_rebal[3:5])
									eff_month_rebal = int(date1_rebal[0:2])

								if from_month <= eff_month:
									if eff_rule:
										date1 = EffectiveDate_Calculate(month1,year,eff_rule)
										eff_day = int(date1[3:5])
										eff_month = int(date1[0:2])
										eff_year = int(date1[6:8])
										ind.Effective_Date_rec = date1

									if sel_rule2:
										sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,sel_rule2)
										sel_day = int(sel_date[3:5])
										sel_month = int(sel_date[0:2])
										sel_year = int(sel_date[6:8])
										ind.Selection_Date_Cycle_2_rec = sel_date

									if fre_rule:
										freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,fre_rule)
										freeze_day = int(freeze_date[3:5])
										freeze_month = int(freeze_date[0:2])
										freeze_year = int(freeze_date[6:8])
										ind.Weights_Share_Freeze_Date_rec = freeze_date

									if(ann_rule==1):
										announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,ann_rule)
									else:
										announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,ann_rule)
									announce_day = int(announce_date[3:5])
									announce_month = int(announce_date[0:2])
									announce_year = int(announce_date[6:8])
									ind.Public_Announcement_Date_rec = announce_date

									if(client_comm_rule==1):
										client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,client_comm_rule)
									else:
										client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,client_comm_rule)
									Client_comm_day = int(client_comm_date[3:5])
									Client_comm_month = int(client_comm_date[0:2])
									Client_comm_year = int(client_comm_date[6:8])
									ind.Client_Comm_Date_rec = client_comm_date

									Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year)
									Prelim_comm_day = int(Prelim_comm_date[3:5])
									Prelim_comm_month = int(Prelim_comm_date[0:2])
									Prelim_comm_year = int(Prelim_comm_date[6:8])
									ind.Prelim_Comm_Date_rec = Prelim_comm_date

									Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year)
									Completion_month = int(Completion_date[0:2])
									ind.Completion_Date_rec = Completion_date
									if(Prelim_comm_rule==""):
										Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year)
									else:
										Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year)

									ind.Ind_Cmte_Comm_Date_rec = Committee_comm_date
									Committee_comm_month = int(Committee_comm_date[0:2])

									selection_date1 = selection_date_1_Calculate(sel_day,sel_month,sel_year)
									sel_1_month = int(selection_date1[0:2])
									ind.Selection_Date_Cycle_1_rec = selection_date1
									data = {}
								
				
								elif from_month <= eff_month_rebal:
									if eff_rule_rebal:
										date1_rebal = EffectiveDate_Calculate(month1_rebal,year,eff_rule_rebal)
										eff_day_rebal = int(date1_rebal[3:5])
										eff_month_rebal = int(date1_rebal[0:2])
										eff_year_rebal = int(date1_rebal[6:8])
										ind.Effective_Date_re = date1_rebal

									if sel_rule2_rebal:
										sel_date_rebal = SelectionDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,sel_rule2_rebal)
										sel_day_rebal = int(sel_date_rebal[3:5])
										sel_month_rebal = int(sel_date_rebal[0:2])
										sel_year_rebal = int(sel_date_rebal[6:8])
										ind.Selection_Date_Cycle_2_re = sel_date_rebal

									if fre_rule_rebal:
										freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,fre_rule_rebal)
										freeze_day_rebal = int(freeze_date_rebal[3:5])
										freeze_month_rebal = int(freeze_date_rebal[0:2])
										freeze_year_rebal = int(freeze_date_rebal[6:8])
										ind.Weights_Share_Freeze_Date_re = freeze_date_rebal

									if(ann_rule_rebal==1):
										announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,ann_rule_rebal)
									else:
										announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,ann_rule_rebal)
									announce_day_rebal = int(announce_date_rebal[3:5])
									announce_month_rebal = int(announce_date_rebal[0:2])
									announce_year_rebal = int(announce_date_rebal[6:8])
									ind.Public_Announcement_Date_re = announce_date_rebal

									if(client_comm_rule_rebal==1):
										client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,client_comm_rule_rebal)
									else:
										client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,client_comm_rule_rebal)
									Client_comm_day_rebal = int(client_comm_date_rebal[3:5])
									Client_comm_month_rebal = int(client_comm_date_rebal[0:2])
									Client_comm_year_rebal = int(client_comm_date_rebal[6:8])
									ind.Client_Comm_Date_re = client_comm_date_rebal

									Prelim_comm_date_rebal = PreliminaryCommDate_Calculate(announce_day_rebal,announce_month_rebal,announce_year_rebal)
									Prelim_comm_day_rebal = int(Prelim_comm_date_rebal[3:5])
									Prelim_comm_month_rebal = int(Prelim_comm_date_rebal[0:2])
									Prelim_comm_year_rebal = int(Prelim_comm_date_rebal[6:8])
									ind.Prelim_Comm_Date_re = Prelim_comm_date_rebal

									Completion_date_rebal = CompletionDate_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
									Completion_month_rebal = int(Completion_date_rebal[0:2])
									ind.Completion_Date_re = Completion_date_rebal
									if(Prelim_comm_rule_rebal==""):
										Committee_comm_date_rebal = Committe_comm_date_Calculate(Client_comm_day_rebal,Client_comm_month_rebal,Client_comm_year_rebal)
									else:
										Committee_comm_date_rebal = Committe_comm_date_Calculate(Prelim_comm_day_rebal,Prelim_comm_month_rebal,Prelim_comm_year_rebal)

									ind.Ind_Cmte_Comm_Date_re = Committee_comm_date_rebal
									Committee_comm_month_rebal = int(Committee_comm_date_rebal[0:2])

									selection_date1_rebal= selection_date_1_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
									sel_1_month_rebal = int(selection_date1_rebal[0:2])
									ind.Selection_Date_Cycle_1_re = selection_date1_rebal
									data = {}
								
								
								informa.append(ind)
					

							elif(sel_type =="selection"):
								if eff_rule:
									date1 = EffectiveDate_Calculate(month1,year,eff_rule)
									eff_day = int(date1[3:5])
									eff_month = int(date1[0:2])
									eff_year = int(date1[6:8])
									#ind.Effective_Date = date1

								if sel_rule2:
									sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,sel_rule2)
									sel_day = int(sel_date[3:5])
									sel_month = int(sel_date[0:2])
									sel_year = int(sel_date[6:8])
									
								if eff_rule_rebal:
									date1_rebal = EffectiveDate_Calculate(month1_rebal,year,eff_rule_rebal)
									eff_day_rebal = int(date1_rebal[3:5])
									eff_month_rebal = int(date1_rebal[0:2])
									eff_year_rebal = int(date1_rebal[6:8])

								if sel_rule2_rebal:
									sel_date_rebal = SelectionDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,sel_rule2_rebal)
									sel_day_rebal = int(sel_date_rebal[3:5])
									sel_month_rebal = int(sel_date_rebal[0:2])


								if(from_month <= sel_month):
									if eff_rule:
										date1 = EffectiveDate_Calculate(month1,year,eff_rule)
										eff_day = int(date1[3:5])
										eff_month = int(date1[0:2])
										eff_year = int(date1[6:8])
										ind.Effective_Date_rec = date1

									if sel_rule2:
										sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,sel_rule2)
										sel_day = int(sel_date[3:5])
										sel_month = int(sel_date[0:2])
										sel_year = int(sel_date[6:8])
										ind.Selection_Date_Cycle_2_rec = sel_date

									if fre_rule:
										freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,fre_rule)
										freeze_day = int(freeze_date[3:5])
										freeze_month = int(freeze_date[0:2])
										freeze_year = int(freeze_date[6:8])
										ind.Weights_Share_Freeze_Date_rec = freeze_date

									if(ann_rule==1):
										announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,ann_rule)
									else:
										announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,ann_rule)
									announce_day = int(announce_date[3:5])
									announce_month = int(announce_date[0:2])
									announce_year = int(announce_date[6:8])
									ind.Public_Announcement_Date_rec = announce_date

									if(client_comm_rule==1):
										client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,client_comm_rule)
									else:
										client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,client_comm_rule)
									Client_comm_day = int(client_comm_date[3:5])
									Client_comm_month = int(client_comm_date[0:2])
									Client_comm_year = int(client_comm_date[6:8])
									ind.Client_Comm_Date_rec = client_comm_date

									Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year)
									Prelim_comm_day = int(Prelim_comm_date[3:5])
									Prelim_comm_month = int(Prelim_comm_date[0:2])
									Prelim_comm_year = int(Prelim_comm_date[6:8])
									ind.Prelim_Comm_Date_rec = Prelim_comm_date

									Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year)
									Completion_month = int(Completion_date[0:2])
									ind.Completion_Date_rec = Completion_date
									if(Prelim_comm_rule==""):
										Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year)
									else:
										Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year)

									ind.Ind_Cmte_Comm_Date_rec = Committee_comm_date
									Committee_comm_month = int(Committee_comm_date[0:2])

									selection_date1 = selection_date_1_Calculate(sel_day,sel_month,sel_year)
									sel_1_month = int(selection_date1[0:2])
									ind.Selection_Date_Cycle_1_rec = selection_date1
									data = {}
									
									
								if from_month <= sel_month_rebal:
									if eff_rule_rebal:
										date1_rebal = EffectiveDate_Calculate(month1_rebal,year,eff_rule_rebal)
										eff_day_rebal = int(date1_rebal[3:5])
										eff_month_rebal = int(date1_rebal[0:2])
										eff_year_rebal = int(date1_rebal[6:8])
										ind.Effective_Date_re = date1_rebal

									if sel_rule2_rebal:
										sel_date_rebal = SelectionDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,sel_rule2_rebal)
										sel_day_rebal = int(sel_date_rebal[3:5])
										sel_month_rebal = int(sel_date_rebal[0:2])
										sel_year_rebal = int(sel_date_rebal[6:8])
										ind.Selection_Date_Cycle_2_re = sel_date_rebal

									if fre_rule_rebal:
										freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,fre_rule_rebal)
										freeze_day_rebal = int(freeze_date_rebal[3:5])
										freeze_month_rebal = int(freeze_date_rebal[0:2])
										freeze_year_rebal = int(freeze_date_rebal[6:8])
										ind.Weights_Share_Freeze_Date_re = freeze_date_rebal

									if(ann_rule_rebal==1):
										announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,ann_rule_rebal)
									else:
										announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,ann_rule_rebal)
									announce_day_rebal = int(announce_date_rebal[3:5])
									announce_month_rebal = int(announce_date_rebal[0:2])
									announce_year_rebal = int(announce_date_rebal[6:8])
									ind.Public_Announcement_Date_re = announce_date_rebal

									if(client_comm_rule_rebal==1):
										client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,client_comm_rule_rebal)
									else:
										client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,client_comm_rule_rebal)
									Client_comm_day_rebal = int(client_comm_date_rebal[3:5])
									Client_comm_month_rebal = int(client_comm_date_rebal[0:2])
									Client_comm_year_rebal = int(client_comm_date_rebal[6:8])
									ind.Client_Comm_Date_re = client_comm_date_rebal

									Prelim_comm_date_rebal = PreliminaryCommDate_Calculate(announce_day_rebal,announce_month_rebal,announce_year_rebal)
									Prelim_comm_day_rebal = int(Prelim_comm_date_rebal[3:5])
									Prelim_comm_month_rebal = int(Prelim_comm_date_rebal[0:2])
									Prelim_comm_year_rebal = int(Prelim_comm_date_rebal[6:8])
									ind.Prelim_Comm_Date_re = Prelim_comm_date_rebal

									Completion_date_rebal = CompletionDate_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
									Completion_month_rebal = int(Completion_date_rebal[0:2])
									ind.Completion_Date_re = Completion_date_rebal
									if(Prelim_comm_rule_rebal==""):
										Committee_comm_date_rebal = Committe_comm_date_Calculate(Client_comm_day_rebal,Client_comm_month_rebal,Client_comm_year_rebal)
									else:
										Committee_comm_date_rebal = Committe_comm_date_Calculate(Prelim_comm_day_rebal,Prelim_comm_month_rebal,Prelim_comm_year_rebal)

									ind.Ind_Cmte_Comm_Date_re = Committee_comm_date_rebal
									Committee_comm_month_rebal = int(Committee_comm_date_rebal[0:2])

									selection_date1_rebal= selection_date_1_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
									sel_1_month_rebal = int(selection_date1_rebal[0:2])
									ind.Selection_Date_Cycle_1_re = selection_date1_rebal
									data = {}
								informa.append(ind)
								

							elif(sel_type =="icom"):
								if eff_rule:
									date1 = EffectiveDate_Calculate(month1,year,eff_rule)
									eff_day = int(date1[3:5])
									eff_month = int(date1[0:2])
									eff_year = int(date1[6:8])

								if sel_rule2:
									sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,sel_rule2)
									sel_day = int(sel_date[3:5])
									sel_month = int(sel_date[0:2])
									sel_year = int(sel_date[6:8])

								if fre_rule:
									freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,fre_rule)
									freeze_day = int(freeze_date[3:5])
									freeze_month = int(freeze_date[0:2])
									freeze_year = int(freeze_date[6:8])

								if(ann_rule==1):
									announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,ann_rule)
								else:
									announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,ann_rule)
								announce_day = int(announce_date[3:5])
								announce_month = int(announce_date[0:2])
								announce_year = int(announce_date[6:8])

								if(client_comm_rule==1):
									client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,client_comm_rule)
								else:
									client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,client_comm_rule)
								Client_comm_day = int(client_comm_date[3:5])
								Client_comm_month = int(client_comm_date[0:2])
								Client_comm_year = int(client_comm_date[6:8])

								Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year)
								Prelim_comm_day = int(Prelim_comm_date[3:5])
								Prelim_comm_month = int(Prelim_comm_date[0:2])
								Prelim_comm_year = int(Prelim_comm_date[6:8])

								Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year)
								Completion_month = int(Completion_date[0:2])
								if(Prelim_comm_rule==""):
									Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year)
								else:
									Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year)

								
								Committee_comm_month = int(Committee_comm_date[0:2])
								
								if eff_rule_rebal:
									date1_rebal = EffectiveDate_Calculate(month1_rebal,year,eff_rule_rebal)
									eff_day_rebal = int(date1_rebal[3:5])
									eff_month_rebal = int(date1_rebal[0:2])
									eff_year_rebal = int(date1_rebal[6:8])
									

								if sel_rule2_rebal:
									sel_date_rebal = SelectionDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,sel_rule2_rebal)
									sel_day_rebal = int(sel_date_rebal[3:5])
									sel_month_rebal = int(sel_date_rebal[0:2])
									sel_year_rebal = int(sel_date_rebal[6:8])
									

								if fre_rule_rebal:
									freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,fre_rule_rebal)
									freeze_day_rebal = int(freeze_date_rebal[3:5])
									freeze_month_rebal = int(freeze_date_rebal[0:2])
									freeze_year_rebal = int(freeze_date_rebal[6:8])
									

								if(ann_rule_rebal==1):
									announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,ann_rule_rebal)
								else:
									announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,ann_rule_rebal)
								announce_day_rebal = int(announce_date_rebal[3:5])
								announce_month_rebal = int(announce_date_rebal[0:2])
								announce_year_rebal = int(announce_date_rebal[6:8])
								

								if(client_comm_rule_rebal==1):
									client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,client_comm_rule_rebal)
								else:
									client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,client_comm_rule_rebal)
								Client_comm_day_rebal = int(client_comm_date_rebal[3:5])
								Client_comm_month_rebal = int(client_comm_date_rebal[0:2])
								Client_comm_year_rebal = int(client_comm_date_rebal[6:8])
								

								Prelim_comm_date_rebal = PreliminaryCommDate_Calculate(announce_day_rebal,announce_month_rebal,announce_year_rebal)
								Prelim_comm_day_rebal = int(Prelim_comm_date_rebal[3:5])
								Prelim_comm_month_rebal = int(Prelim_comm_date_rebal[0:2])
								Prelim_comm_year_rebal = int(Prelim_comm_date_rebal[6:8])
								Completion_date_rebal = CompletionDate_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
								Completion_month_rebal = int(Completion_date_rebal[0:2])
							
								if(Prelim_comm_rule_rebal==""):
									Committee_comm_date_rebal = Committe_comm_date_Calculate(Client_comm_day_rebal,Client_comm_month_rebal,Client_comm_year_rebal)
								else:
									Committee_comm_date_rebal = Committe_comm_date_Calculate(Prelim_comm_day_rebal,Prelim_comm_month_rebal,Prelim_comm_year_rebal)

								Committee_comm_month_rebal = int(Committee_comm_date_rebal[0:2])

								if from_month <= Committee_comm_month:
									if eff_rule:
										date1 = EffectiveDate_Calculate(month1,year,eff_rule)
										eff_day = int(date1[3:5])
										eff_month = int(date1[0:2])
										eff_year = int(date1[6:8])
										ind.Effective_Date_rec = date1

									if sel_rule2:
										sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,sel_rule2)
										sel_day = int(sel_date[3:5])
										sel_month = int(sel_date[0:2])
										sel_year = int(sel_date[6:8])
										ind.Selection_Date_Cycle_2_rec = sel_date

									if fre_rule:
										freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,fre_rule)
										freeze_day = int(freeze_date[3:5])
										freeze_month = int(freeze_date[0:2])
										freeze_year = int(freeze_date[6:8])
										ind.Weights_Share_Freeze_Date_rec = freeze_date

									if(ann_rule==1):
										announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,ann_rule)
									else:
										announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,ann_rule)
									announce_day = int(announce_date[3:5])
									announce_month = int(announce_date[0:2])
									announce_year = int(announce_date[6:8])
									ind.Public_Announcement_Date_rec = announce_date

									if(client_comm_rule==1):
										client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,client_comm_rule)
									else:
										client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,client_comm_rule)
									Client_comm_day = int(client_comm_date[3:5])
									Client_comm_month = int(client_comm_date[0:2])
									Client_comm_year = int(client_comm_date[6:8])
									ind.Client_Comm_Date_rec = client_comm_date

									Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year)
									Prelim_comm_day = int(Prelim_comm_date[3:5])
									Prelim_comm_month = int(Prelim_comm_date[0:2])
									Prelim_comm_year = int(Prelim_comm_date[6:8])
									ind.Prelim_Comm_Date_rec = Prelim_comm_date

									Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year)
									Completion_month = int(Completion_date[0:2])
									ind.Completion_Date_rec = Completion_date
									if(Prelim_comm_rule==""):
										Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year)
									else:
										Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year)

									ind.Ind_Cmte_Comm_Date_rec = Committee_comm_date
									Committee_comm_month = int(Committee_comm_date_rebal[0:2])

									selection_date1 = selection_date_1_Calculate(sel_day,sel_month,sel_year)
									sel_1_month = int(selection_date1[0:2])
									ind.Selection_Date_Cycle_1_rec = selection_date1
									data = {}
								
								if from_month <= Committee_comm_month_rebal:
									if eff_rule_rebal:
										date1_rebal = EffectiveDate_Calculate(month1_rebal,year,eff_rule_rebal)
										eff_day_rebal = int(date1_rebal[3:5])
										eff_month_rebal = int(date1_rebal[0:2])
										eff_year_rebal = int(date1_rebal[6:8])
										ind.Effective_Date_re = date1_rebal

									if sel_rule2_rebal:
										sel_date_rebal = SelectionDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,sel_rule2_rebal)
										sel_day_rebal = int(sel_date_rebal[3:5])
										sel_month_rebal = int(sel_date_rebal[0:2])
										sel_year_rebal = int(sel_date_rebal[6:8])
										ind.Selection_Date_Cycle_2_re = sel_date_rebal

									if fre_rule_rebal:
										freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,fre_rule_rebal)
										freeze_day_rebal = int(freeze_date_rebal[3:5])
										freeze_month_rebal = int(freeze_date_rebal[0:2])
										freeze_year_rebal = int(freeze_date_rebal[6:8])
										ind.Weights_Share_Freeze_Date_re = freeze_date_rebal

									if(ann_rule_rebal==1):
										announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,ann_rule_rebal)
									else:
										announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,ann_rule_rebal)
									announce_day_rebal = int(announce_date_rebal[3:5])
									announce_month_rebal = int(announce_date_rebal[0:2])
									announce_year_rebal = int(announce_date_rebal[6:8])
									ind.Public_Announcement_Date_re = announce_date_rebal

									if(client_comm_rule_rebal==1):
										client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,client_comm_rule_rebal)
									else:
										client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,client_comm_rule_rebal)
									Client_comm_day_rebal = int(client_comm_date_rebal[3:5])
									Client_comm_month_rebal = int(client_comm_date_rebal[0:2])
									Client_comm_year_rebal = int(client_comm_date_rebal[6:8])
									ind.Client_Comm_Date_re = client_comm_date_rebal

									Prelim_comm_date_rebal = PreliminaryCommDate_Calculate(announce_day_rebal,announce_month_rebal,announce_year_rebal)
									Prelim_comm_day_rebal = int(Prelim_comm_date_rebal[3:5])
									Prelim_comm_month_rebal = int(Prelim_comm_date_rebal[0:2])
									Prelim_comm_year_rebal = int(Prelim_comm_date_rebal[6:8])
									ind.Prelim_Comm_Date_re = Prelim_comm_date_rebal

									Completion_date_rebal = CompletionDate_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
									Completion_month_rebal = int(Completion_date_rebal[0:2])
									ind.Completion_Date_re = Completion_date_rebal
									if(Prelim_comm_rule_rebal==""):
										Committee_comm_date_rebal = Committe_comm_date_Calculate(Client_comm_day_rebal,Client_comm_month_rebal,Client_comm_year_rebal)
									else:
										Committee_comm_date_rebal = Committe_comm_date_Calculate(Prelim_comm_day_rebal,Prelim_comm_month_rebal,Prelim_comm_year_rebal)

									ind.Ind_Cmte_Comm_Date_re = Committee_comm_date_rebal
									Committee_comm_month_rebal = int(Committee_comm_date[0:2])

									selection_date1_rebal= selection_date_1_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
									sel_1_month_rebal = int(selection_date1_rebal[0:2])
									ind.Selection_Date_Cycle_1_re = selection_date1_rebal
									data = {}
								informa.append(ind)
						else:
							if eff_rule:
								date1 = EffectiveDate_Calculate(month1,year,eff_rule)
								eff_day = int(date1[3:5])
								eff_month = int(date1[0:2])
								eff_year = int(date1[6:8])
								ind.Effective_Date_rec = date1


							if sel_rule2:
								sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,sel_rule2)
								sel_day = int(sel_date[3:5])
								sel_month = int(sel_date[0:2])
								sel_year = int(sel_date[6:8])
								ind.Selection_Date_Cycle_1_rec = sel_date


							if fre_rule:
								freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,fre_rule)
								freeze_day = int(freeze_date[3:5])
								freeze_month = int(freeze_date[0:2])
								freeze_year = int(freeze_date[6:8])
								ind.Weights_Share_Freeze_Date_rec = freeze_date

							if(ann_rule==1):
								announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,ann_rule)
							else:
								announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,ann_rule)
							announce_day = int(announce_date[3:5])
							announce_month = int(announce_date[0:2])
							announce_year = int(announce_date[6:8])
							ind.Public_Announcement_Date_rec = announce_date

							if(client_comm_rule==1):
								client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,client_comm_rule)
							else:
								client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,client_comm_rule)
							Client_comm_day = int(client_comm_date[3:5])
							Client_comm_month = int(client_comm_date[0:2])
							Client_comm_year = int(client_comm_date[6:8])
							ind.Client_Comm_Date_rec = client_comm_date

							Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year)
							Prelim_comm_day = int(Prelim_comm_date[3:5])
							Prelim_comm_month = int(Prelim_comm_date[0:2])
							Prelim_comm_year = int(Prelim_comm_date[6:8])
							ind.Prelim_Comm_Date_rec = Prelim_comm_date

							Completion_date = CompletionDate_Calculate(sel_day,month1,year)
							Completion_month = int(Completion_date[0:2])
							ind.Completion_Date_rec = Completion_date
							if(Prelim_comm_rule==""):
								Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year)
							else:
								Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year)

							ind.Ind_Cmte_Comm_Date_rec = Committee_comm_date
							Committee_comm_month = int(Committee_comm_date[0:2])

							selection_date1 = selection_date_1_Calculate(sel_day,sel_month,sel_year)
							sel_1_month = int(selection_date1[0:2])
							ind.Selection_Date_Cycle_1_rec = selection_date1
							data = {}
							informa.append(ind)
							
							if eff_rule_rebal:
								date1_rebal = EffectiveDate_Calculate(month1_rebal,year,eff_rule_rebal)
								eff_day_rebal = int(date1_rebal[3:5])
								eff_month_rebal = int(date1_rebal[0:2])
								eff_year_rebal = int(date1_rebal[6:8])
								ind.Effective_Date_re = date1_rebal

							if sel_rule2_rebal:
								sel_date_rebal = SelectionDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,sel_rule2_rebal)
								sel_day_rebal = int(sel_date_rebal[3:5])
								sel_month_rebal = int(sel_date_rebal[0:2])
								sel_year_rebal = int(sel_date_rebal[6:8])
								ind.Selection_Date_Cycle_2_re = sel_date_rebal

							if fre_rule_rebal:
								freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,fre_rule_rebal)
								freeze_day_rebal = int(freeze_date_rebal[3:5])
								freeze_month_rebal = int(freeze_date_rebal[0:2])
								freeze_year_rebal = int(freeze_date_rebal[6:8])
								ind.Weights_Share_Freeze_Date_re = freeze_date_rebal

							if(ann_rule_rebal==1):
								announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,ann_rule_rebal)
							else:
								announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,ann_rule_rebal)
							announce_day_rebal = int(announce_date_rebal[3:5])
							announce_month_rebal = int(announce_date_rebal[0:2])
							#announce_year_rebal = int(announce_year_rebal[6:8])
							announce_year_rebal = int(announce_date_rebal[6:8])
							ind.Public_Announcement_Date_re = announce_date_rebal

							if(client_comm_rule_rebal==1):
								client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,client_comm_date_rebal)
							else:
								client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,client_comm_date_rebal)
							Client_comm_day_rebal = int(client_comm_date_rebal[3:5])
							Client_comm_month_rebal = int(client_comm_date_rebal[0:2])
							Client_comm_year_rebal = int(client_comm_date_rebal[6:8])
							ind.Client_Comm_Date_re = client_comm_date_rebal

							Prelim_comm_date_rebal = PreliminaryCommDate_Calculate(announce_day_rebal,announce_month_rebal,announce_year_rebal)
							Prelim_comm_day_rebal = int(Prelim_comm_date_rebal[3:5])
							Prelim_comm_month_rebal = int(Prelim_comm_date_rebal[0:2])
							Prelim_comm_year_rebal = int(Prelim_comm_date_rebal[6:8])
							ind.Prelim_Comm_Date_re = Prelim_comm_date_rebal

							Completion_date_rebal = CompletionDate_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
							Completion_month_rebal = int(Completion_date_rebal[0:2])
							ind.Completion_Date_re = Completion_date_rebal
							if(Prelim_comm_rule_rebal==""):
								Committee_comm_date_rebal = Committe_comm_date_Calculate(Client_comm_day_rebal,Client_comm_month_rebal,Client_comm_year_rebal)
							else:
								Committee_comm_date_rebal = Committe_comm_date_Calculate(Prelim_comm_day_rebal,Prelim_comm_month_rebal,Prelim_comm_year_rebal)

							ind.Ind_Cmte_Comm_Date_re = Committee_comm_date_rebal
							Committee_comm_month_rebal = int(Committee_comm_date_rebal[0:2])

							selection_date1_rebal= selection_date_1_Calculate(sel_day_rebal,sel_month_rebal,sel_year_rebal)
							sel_1_month_rebal = int(selection_date1_rebal[0:2])
							ind.Selection_Date_Cycle_1_re = selection_date1_rebal
							data = {}
							informa.append(ind)

					else:
						informa.append(ind)
			
			page = request.GET.get('page',1)
			paginator = Paginator(informa, 10)
			try:
				informa = paginator.page(page)
			except PageNotAnInteger:
				informa = paginator.page(1)
			except EmptyPage:
				informa = paginator.page(paginator.num_pages)
				
			context = {
				'info' : informa,
				'client_name' : client_name,
				'client_code' : cname1,
				'from_MY' : month,
				'to_MY' : month2,
				'sel_type' : sel_type,


			}



	else:
		page = request.GET.get('page',1)
		paginator = Paginator(info, 10)
		try:
			info = paginator.page(page)
		except PageNotAnInteger:
			info = paginator.page(1)
		except EmptyPage:
			info = paginator.page(paginator.num_pages)
			
		context = {
			'info' : info,
			'client_name' : client_name,

		}
	templateName = "reportGenerate.html"
	return render(request, templateName, context)


def logout(request):
	try:
		del request.session['username']
	except:
		pass
	templateName = "login.html"
	context = {"message":"You are logged out."}
	return render(request, templateName, context)

	
def error_404_view(request, exception):
	templateName = "404.html"
	context = {}
	return render(request, templateName, context)

def error_500_view(request, exception):
	templateName = "500.html"
	context = {}
	return render(request, templateName, context)

def export_index_to_xlsx(request):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    Regindex_queryset = Registerindex.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the column titles and widths
    columns = [
        ('ID', 8),
        ('Index_Name', 40),
        ('Ident_ISIN', 80),
        ('Ident_Bloomberg', 15),
        ('Ind_Ver', 15),
        ('Ind_Ver_ID', 15),
    ]
    
    # Iterate through movie categories
    for category_index, category in enumerate(Regindex_queryset):
        # Create a worksheet/tab with the title of the category
        worksheet = workbook.create_sheet(
            title=category.Index_Name,
            index=category_index,
        )
        # Define the background color of the header cells
		
        fill = PatternFill(
            #start_color=category.html_color,
            #end_color=category.html_color,
            fill_type="solid",
        )
        row_num = 1
        
        # Assign values, styles, and formatting for each cell in the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            #cell.fill = fill
            # set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

        # Iterate through all movies of a category
        for movie in Registerindex.objects.all():
            row_num += 1
            
            # Define data and formats for each cell in the row
            row = [
                (movie.pk, 'Normal'),
                (movie.Index_Name, 'Normal'),
                (movie.Ident_ISIN, 'Normal'),
                #(timedelta(minutes=movie.length_in_minutes), 'Normal'),
				(movie.Ident_Bloomberg, 'Normal'),
                (movie.Ind_Ver, 'Normal'),
                (movie.Ind_Ver_ID, 'Normal'),
            ]
            
            # Assign values, styles, and formatting for each cell in the row
            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                '''if cell_format == 'Currency':
                    cell.number_format = '#,##0.00 €'
                if col_num == 4:
                    cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment
				'''

        # freeze the first row
        worksheet.freeze_panes = worksheet['A2']
        
        # set tab color
        #worksheet.sheet_properties.tabColor = category.html_color

    workbook.save(response)
    return response