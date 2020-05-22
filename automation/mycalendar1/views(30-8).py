# calendar/views.py
"""
 * Version : 1.0
 * Project: calendar Automation
 * Copyright : Indxx Capital Management
 * Author: Pavan Rajput
 * Created Date: 08-04-2019
 * Modified Date: dd-mm-yyyy
 * Licensed under : Self
"""
# Create your views here.
from django.contrib.auth.models import User
from django.core.paginator import Paginator,EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from django.shortcuts import render, render_to_response
from django.template import loader
from django.template.loader import render_to_string
import calendar
from api.models import Calendarlist, Registerindex, Ruleslist,Priordays
from mycalendar.Priorday import priorday_calculate
from django.utils.translation import ugettext as _
from mycalendar.effectivedate_rule1 import EffectiveDate_Calculate
from mycalendar.AnnouncementDate_rule import AnnouncementDate_Calculate
from mycalendar.SelectionDate2_rule import SelectionDate_Calculate
from mycalendar.FreezeDate_rule import FreezeDate_Calculate
from mycalendar.QualityDate_rule import QualityDate_Calculate
from datetime import date, datetime, timedelta
from mycalendar.Preliminarycommdate_rule import PreliminaryCommDate_Calculate
from mycalendar.CompletionDate_cycle1_rule import CompletionDate_Calculate
from mycalendar.Committee_commdate_rule import Committe_comm_date_Calculate
from mycalendar.SelectionDate1_rule import selection_date_1_Calculate
from django.db.models import Q
from mycalendar.numdays import num_of_days
from openpyxl import Workbook
from mycalendar.mail import send_mail
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pdb;

import datetime as dt
from pandas.tseries.holiday import AbstractHolidayCalendar, Holiday, nearest_workday, \
    USMartinLutherKingJr, USPresidentsDay, GoodFriday, USMemorialDay, \
    USLaborDay, USThanksgivingDay


class USTradingCalendar(AbstractHolidayCalendar):
    rules = [
        Holiday('NewYearsDay', month=1, day=1, observance=nearest_workday),
        USMartinLutherKingJr,
        USPresidentsDay,
        GoodFriday,
        USMemorialDay,
        Holiday('USIndependenceDay', month=7, day=4, observance=nearest_workday),
        USLaborDay,
        USThanksgivingDay,
        Holiday('Christmas', month=12, day=25, observance=nearest_workday)
    ]


def index(request):
	templateName = "index.html"
	clientList = Calendarlist.objects.filter(category = 'Client_Name')

	indexlist = Registerindex.objects.all().order_by('id').reverse()
	page = request.GET.get('page', 1)
	paginator = Paginator(indexlist, 20)
	try:
		indexdata = paginator.page(page)
	except PageNotAnInteger:
		indexdata = paginator.page(1)
	except EmptyPage:
		indexdata = paginator.page(paginator.num_pages)

	context = {
		'indexlist': indexdata,
		'clientList': clientList,
	}

	return render(request, templateName, context)
"""
if request.method == 'POST':
search_result = request.POST['search2']
if(search_result!=""):
#templatename = "SearchIndex.html"
lookups = Q(Index_Name__icontains = search_result) | Q(Contract_Type__icontains = search_result) | Q(Reconstitution__icontains = search_result) | Q(Calc_Agent__icontains = search_result) | Q(Ind_Sty__icontains = search_result) | Q(Contract_Type__icontains = search_result) | Q(Type_of_Ind__icontains = search_result)
results= Registerindex.objects.filter(lookups).distinct().reverse()

if(len(results) == 0):
obj = Calendarlist.objects.filter(description = search_result)
code = ""
for ind in obj:
	code = int(ind.code)
	if(code!=""):
		results = Registerindex.objects.filter(Client_Name_id = code)
	else:
		templateName = "Nosearch.html"


context={
	'search_text': search_result,
	'results': results,
	'clientList': clientList,
}
return render(request, templateName, context)

else:
templateName = "Nosearch.html"
context={
	'search_text': search_result,
	'results': '',
	'clientList': clientList,
}

else:
"""


def mycalendar(request):
    templates = loader.get_template("index.html")
    return HttpResponse(templates.render());

def addindex(request):
	template = "addindex.html"
	client_name = Calendarlist.objects.filter(category = 'Client_Name')
	ind_Sty = Calendarlist.objects.filter(category = 'Ind_Sty')
	ind_Version = Calendarlist.objects.filter(category = 'Ind_Ver')
	cal = Calendarlist.objects.filter(category = 'Cal')
	cal_Agent = Calendarlist.objects.filter(category = 'Cal_Agent')
	contract_Type = Calendarlist.objects.filter(category = 'Contract_Type')
	type_of_Ind = Calendarlist.objects.filter(category = 'Type_of_Ind')
	prod_Stat = Calendarlist.objects.filter(category = 'Prod_Stat')
	etf_Launched = Calendarlist.objects.filter(category = 'ETF_Launched')
	reconstitution = Calendarlist.objects.filter(category = 'Reconstitution')
	rebalance = Calendarlist.objects.filter(category = 'Rebalance')
	review = Calendarlist.objects.filter(category = 'Review')
	theme_Review = Calendarlist.objects.filter(category = 'Theme_Review')
	comp_Date = Calendarlist.objects.filter(category = 'Comp_Date')
	selec_Date_Cyc_1 = Calendarlist.objects.filter(category = 'Selec_Date_Cyc_1')
	selec_Date_Cyc_2 = Calendarlist.objects.filter(category = 'Selec_Date_Cyc_2')
	weights_Share_Freeze = Calendarlist.objects.filter(category = 'Weights_Share_Freeze_Date')
	public_Announcement = Calendarlist.objects.filter(category = 'Public_Announcement_Date')
	client_Comm = Calendarlist.objects.filter(category = 'Client_Comm_Date')
	ind_Cmte_Comm_Date = Calendarlist.objects.filter(category = 'Ind_Cmte_Comm_Date')
	effec_Date = Calendarlist.objects.filter(category = 'Effec_Date')
	prelim_Comm_Date = Calendarlist.objects.filter(category = 'Prelim_Comm_Date')
	comm_to_Calc_Agent = Calendarlist.objects.filter(category = 'Comm_to_Calc_Agent')
	qc_Date = Calendarlist.objects.filter(category = 'QC_Date')
	#send_mail('subject','message','mjain@indxx.com','pavank@indxx.com',    fail_silently=False)

	context = {
		'client_name':client_name,
		'ind_Sty':ind_Sty,
		'ind_Version':ind_Version,
		'cal':cal,
		'cal_Agent':cal_Agent,
		'contract_Type':contract_Type,
		'type_of_Ind':type_of_Ind,
		'prod_Stat':prod_Stat,
		'etf_Launched':etf_Launched,
		'reconstitution':reconstitution,
		'rebalance':rebalance,
		'review':review,
		'theme_Review':theme_Review,
		'comp_Date':comp_Date,
		'selec_Date_Cyc_1':selec_Date_Cyc_1,
		'selec_Date_Cyc_2':selec_Date_Cyc_2,
		'weights_Share_Freeze':weights_Share_Freeze,
		'public_Announcement':public_Announcement,
		'client_Comm':client_Comm,
		'ind_Cmte_Comm_Date':ind_Cmte_Comm_Date,
		'effec_Date':effec_Date,
		'prelim_Comm_Date':prelim_Comm_Date,
		'comm_to_Calc_Agent':comm_to_Calc_Agent,
		'qc_Date':qc_Date,
		}

	return render(request, template, context)



def post_index(request):
	inst = USTradingCalendar()
	y1 = date.today().year
	holidays = inst.holidays(dt.datetime(y1-1, 9, 30), dt.datetime(y1, 12, 31))
	if request.method == 'POST':

			check1 = request.POST.get('check1', False)
			check2 = request.POST.get('check2', False)
			check3 = request.POST.get('check3', False)
			choice1 = request.POST.get('choice', False)

			"""
			if(check3=="rev" and choice1== "rule"):

				if(request.POST['review']== "" or request.POST['review_month']== "" or request.POST['effective_date_review']== ""
			  or request.POST['ind_qc_date_review'] == "" or request.POST['selec_Date_Cyc_2_review']== "" or request.POST['ind_Comm_Date_review']== "" or request.POST['weights_Share_Freeze_review']== ""
			or request.POST['public_Announcement_review']== "" or request.POST['client_Comm_review']== "" or request.POST['comm_to_Calc_Agent_review']== ""):
					templatename = "600.html"
					return render(request, templatename)
			elif(check3=="rev" and choice1=="date"):
				if(request.POST['review']== "" or request.POST['review_month']== "" or request.POST['mnl_eff_date_review']== ""or request.POST['mnl_qc_date_review']==""
			  or request.POST['mnl_sel2_date_review']== "" or request.POST['mnl_ind_Comm_date_review']== "" or request.POST['mnl_freeze_date_review']== ""
			or request.POST['mnl_pb_announce_date_review']== "" or request.POST['mnl_client_comm_date_review']== "" or request.POST['mnl_comm_cal_date_review']== ""):
					templatename = "600.html"
					return render(request, templatename)
			"""
			reconst_Effective_Date = ""
			reconst_Selection_Date_Cycle_2 = ""
			reconst_Weights_Share_Freeze_Date = ""
			reconst_Public_Announcement_Date = ""
			reconst_Client_Comm_Date = ""
			reconst_Prelim_Comm_Date = ""
			reconst_Completion_Date = ""
			reconst_qc_date = ""
			#reconst_Ind_Cmte_Comm_Date = ""
			reconst_comm_cal_date = ""
			re_comm_cal_date=""
			rev_comm_cal_date= ""
			reconst_Selection_Date_Cycle_1 = ""
			rev_effective = ""
			rev_selection2 = ""
			rev_freeze = ""
			rev_announce = ""
			rev_clientcomm = ""
			rev_completion = ""
			rev_qc_date = ""
			rev_commiteecomm = ""
			rev_selection1 = ""
			re_effective = ""
			re_selection2 = ""
			re_freeze = ""
			re_announce = ""
			re_clientcomm = ""
			re_prelim = ""
			re_completion = ""
			re_qc = ""
			re_selection1 = ""
			color_1 = 1
			if(request.POST['mnl_eff_date']=="" and request.POST['mnl_eff_date_review']=="" and request.POST['mnl_eff_date_rebal']==""):
				if(request.POST['reconst_month'] != ""):
					if((request.POST['reconstitution']=="Annual") or (request.POST['reconstitution']=="Monthly")):
						num_month = 1
						gap=1
					elif(request.POST['reconstitution']== "Quarterly"):
						num_month = 4
						gap = 3
					else:
						num_month = 2
						gap = 6
					months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
							  October=10, November=11, December=12)
					month1 = months[request.POST['reconst_month']]
					month1 = int(month1)
					cal_month = month1
					year = date.today().year
					for i in range(0,num_month):
						if(month1>12):
							month1 = month1%12;
						if(month1<cal_month):
							year = year+1
						eff_rule = int(request.POST['effective_date'])
						sel_rule = int(request.POST['selec_Date_Cyc_2'])
						fre_rule = int(request.POST['weights_Share_Freeze'])
						ann_rule = int(request.POST['public_Announcement'])
						client_comm_rule = int(request.POST['client_Comm'])
						#icom_rule = int(request.POST['ind_Comm_Date'])
						prelim_comm_rule = int(request.POST['pre_comm_date']) if(request.POST['pre_comm_date']!="") else request.POST['pre_comm_date']
						date1 = EffectiveDate_Calculate(month1, year, eff_rule,holidays)
						eff_day = int(date1[3:5])
						eff_mon = int(date1[0:2])
						eff_year = int(date1[6:8])
						sel2_prior_day = int(request.POST['sel2_prior_days']) if(request.POST['sel2_prior_days']!="")else request.POST['sel2_prior_days']
						if(sel_rule==7):
							sel_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,sel2_prior_day)
						else:
						    sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule,holidays)
						sel_day = int(sel_date[3:5])
						sel_month = int(sel_date[0:2])
						sel_year = int(sel_date[6:8])
						fre_prior_day = int(request.POST['fre_prior_days']) if(request.POST['fre_prior_days']!="") else request.POST['fre_prior_days']
						if(fre_rule==6):

							freeze_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,fre_prior_day)
						else:
						    freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule,holidays)
						freeze_day = int(freeze_date[3:5])
						freeze_mon = int(freeze_date[0:2])
						freeze_year = int(freeze_date[6:8])
						ann_prior_day = int(request.POST['ann_prior_days']) if(request.POST['ann_prior_days']!="") else request.POST['ann_prior_days']
						if(ann_rule==3):
							announce_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,ann_prior_day)
						else:
							if (ann_rule == 1):
								announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule,holidays)
							else:
								announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule,holidays)
						announce_day = int(announce_date[3:5])
						announce_mon = int(announce_date[0:2])
						announce_year = int(announce_date[6:8])
						cl_prior_day = int(request.POST['cl_prior_days']) if(request.POST['cl_prior_days']!="") else request.POST['cl_prior_days']
						if(client_comm_rule==3):
							client_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,cl_prior_day)
						else:
							if (client_comm_rule == 1):
								client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, client_comm_rule,holidays)
							else:
								client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year,client_comm_rule,holidays)
						Client_comm_day = int(client_comm_date[3:5])
						Client_comm_mon = int(client_comm_date[0:2])
						Client_comm_year = int(client_comm_date[6:8])

						comm_cal_prior_day = int(request.POST['cal_ag_prior_days']) if(request.POST['cal_ag_prior_days']!="") else request.POST['cal_ag_prior_days']
						comm_cal_rule = int(request.POST['comm_to_Calc_Agent'])
						if(comm_cal_rule==2):
							comm_cal_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,comm_cal_prior_day)
						else:
							comm_cal_date = client_comm_date
						pre_prior_day = int(request.POST['pre_prior_days']) if(request.POST['pre_prior_days']!="") else request.POST['pre_prior_days']
						if(prelim_comm_rule != ""):
							if(prelim_comm_rule==2):

								Prelim_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,pre_prior_day)
							else:
								Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day, announce_mon, announce_year,holidays)
							Prelim_comm_day = int(Prelim_comm_date[3:5])
							Prelim_comm_mon = int(Prelim_comm_date[0:2])
							Prelim_comm_year = int(Prelim_comm_date[6:8])
						else:
							Prelim_comm_date=""
						"""
						icom_prior_day = int(request.POST['icom_prior_days']) if(request.POST['icom_prior_days']!="") else request.POST['icom_prior_days']
						if (icom_rule == 2):
							Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon,
																		   Client_comm_year,holidays)
						else:
							if(icom_rule==3):
								Committee_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,icom_prior_day)
							else:
							    Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day, Prelim_comm_mon,
																		   Prelim_comm_year,holidays)
						Committee_comm_day = int(Committee_comm_date[3:5])
						Committee_comm_mon = int(Committee_comm_date[0:2])
						Committee_comm_year = int(Committee_comm_date[6:8])
						"""
						qc_rule = int(request.POST['qc_date'])
						qc_prior_day = int(request.POST['qc_prior_days']) if(request.POST['qc_prior_days']!="") else request.POST['qc_prior_days']
						if(qc_rule==3):
							Quality_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,qc_prior_day)
						else:
							Quality_date = QualityDate_Calculate(freeze_day,freeze_mon,freeze_year,holidays)
						cmp_prior_day = int(request.POST['cmp_prior_days']) if(request.POST['cmp_prior_days']!="") else request.POST['cmp_prior_days']

						comp_rule = int(request.POST['cmp_Date'])
						if(comp_rule==2):
							Completion_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,cmp_prior_day)
						else:
							Completion_date = CompletionDate_Calculate(sel_day, sel_month, sel_year,holidays)


						sel1_rule = int(request.POST['selec_Date_Cyc_1']) if(request.POST['selec_Date_Cyc_1']!="") else request.POST['selec_Date_Cyc_1']
						if(sel1_rule==2):
							sel1_prior_day = int(request.POST['sel1_prior_days']) if(request.POST['sel1_prior_days']!="") else request.POST['sel1_prior_days']
							selection_date1 = priorday_calculate(eff_day,eff_mon,eff_year,holidays,sel1_prior_day)
						elif(sel1_rule==1):
							selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year,holidays)
						else:
							sel1_prior_day=""
							selection_date1 =""

						reconst_Selection_Date_Cycle_2 = reconst_Selection_Date_Cycle_2 + sel_date + "\n"
						reconst_Completion_Date = reconst_Completion_Date + Completion_date + "\n"
						reconst_Selection_Date_Cycle_1 = reconst_Selection_Date_Cycle_1+selection_date1 + "\n"

						cal_month = month1
						month1= month1+gap
						reconst_Effective_Date = reconst_Effective_Date + date1 + "\n"
						reconst_Weights_Share_Freeze_Date = reconst_Weights_Share_Freeze_Date + freeze_date + "\n"
						reconst_Public_Announcement_Date = reconst_Public_Announcement_Date + announce_date + "\n"
						reconst_Client_Comm_Date = reconst_Client_Comm_Date + client_comm_date + "\n"
						reconst_comm_cal_date =  reconst_comm_cal_date + comm_cal_date + "\n"
						reconst_Prelim_Comm_Date = reconst_Prelim_Comm_Date + Prelim_comm_date + "\n"
						#reconst_Ind_Cmte_Comm_Date = reconst_Ind_Cmte_Comm_Date + Committee_comm_date + "\n"
						reconst_qc_date = reconst_qc_date + Quality_date + "\n"
				if(request.POST['review_month'] != ""):
					eff_rule_rev = int(request.POST['effective_date_review'])
					sel_rule_rev = int(request.POST['selec_Date_Cyc_2_review'])
					fre_rule_rev = int(request.POST['weights_Share_Freeze_review'])
					ann_rule_rev = int(request.POST['public_Announcement_review'])
					client_comm_rule_rev = int(request.POST['client_Comm_review'])
				#prelim_comm_rule_rev = int(request.POST['pre_comm_date_review'])
				#prelim_comm_rule_rev = int(request.POST['pre_comm_date_review'])
					if(request.POST['review']=="Annual" or request.POST['review']=="Monthly"):
						num_month = 1
						gap=1
					elif(request.POST['review']== "Quarterly"):
						num_month = 4
						gap = 3
					else:
						num_month = 2
						gap = 6
					months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
						  October=10, November=11, December=12)
					month1 = months[request.POST['review_month']]
					month1 = int(month1)
					cal_month = month1
					year = date.today().year
					for i in range(0,num_month):
						if(month1>12):
							month1 = month1%12;
						if(month1<cal_month):
							year = year+1
						date1 = EffectiveDate_Calculate(month1, year, eff_rule_rev,holidays)
						eff_day = int(date1[3:5])
						eff_mon = int(date1[0:2])
						eff_year = int(date1[6:8])
						sel2_prior_day_review = int(request.POST['sel2_prior_days_review']) if(request.POST['sel2_prior_days_review']!="") else request.POST['sel2_prior_days_review']
						if(sel_rule_rev==7):
							sel_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,sel2_prior_day_review)
						else:
							sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule_rev,holidays)
						sel_day = int(sel_date[3:5])
						sel_month = int(sel_date[0:2])
						sel_year = int(sel_date[6:8])
						fre_prior_day_review = int(request.POST['fre_prior_days_review']) if(request.POST['fre_prior_days_review']!="") else request.POST['fre_prior_days_review']
						if(fre_rule_rev==6):
							freeze_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,fre_prior_day_review)
						else:
							freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule_rev,holidays)

						freeze_day = int(freeze_date[3:5])
						freeze_mon = int(freeze_date[0:2])
						freeze_year = int(freeze_date[6:8])
						ann_prior_day_review = int(request.POST['ann_prior_days_review']) if(request.POST['ann_prior_days_review']!="") else request.POST['ann_prior_days_review']
						if(ann_rule_rev==3):
						    announce_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,ann_prior_day_review)
						else:
							if (ann_rule_rev == 1):
								announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rev,holidays)
							else:
								announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rev,holidays)
						announce_day = int(announce_date[3:5])
						announce_mon = int(announce_date[0:2])
						announce_year = int(announce_date[6:8])
						cl_prior_day_review = int(request.POST['cl_prior_days_review']) if(request.POST['cl_prior_days_review']!="") else request.POST['cl_prior_days_review']
						if(client_comm_rule_rev==3):
						    client_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,cl_prior_day_review)
						else:
							if (client_comm_rule_rev == 1):
								client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, client_comm_rule_rev,holidays)
							else:
								client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, client_comm_rule_rev,holidays)
						Client_comm_day = int(client_comm_date[3:5])
						Client_comm_mon = int(client_comm_date[0:2])
						Client_comm_year = int(client_comm_date[6:8])
						comm_cal_rule_review = int(request.POST['comm_to_Calc_Agent_review'])
						comm_cal_prior_day_review = int(request.POST['cal_ag_prior_days_review']) if(request.POST['cal_ag_prior_days_review']!="") else request.POST['cal_ag_prior_days_review']
						if(comm_cal_rule_review==2):
							comm_cal_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,comm_cal_prior_day_review)
						else:
							comm_cal_date = client_comm_date
						"""	
						icom_rule_review = int(request.POST['ind_Comm_Date_review'])
						icom_prior_day_review = int(request.POST['icom_prior_days_review']) if(request.POST['icom_prior_days_review']!="") else request.POST['icom_prior_days_review']
						if(icom_rule_review==3):

						     Committee_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,icom_prior_day_review)
						else:
							 Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon, Client_comm_year,holidays)
						Committee_comm_day = int(Committee_comm_date[3:5])
						Committee_comm_mon = int(Committee_comm_date[0:2])
						Committee_comm_year = int(Committee_comm_date[6:8])
						"""
						qc_rule_rev = int(request.POST['ind_qc_date_review'])
						qc_prior_day_review = int(request.POST['qc_prior_days_review']) if(request.POST['qc_prior_days_review']!="") else request.POST['qc_prior_days_review']
						if(qc_rule_rev==3):
							Quality_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,qc_prior_day_review)
						else:
							Quality_date = QualityDate_Calculate(freeze_day,freeze_mon,freeze_year,holidays)

						cal_month = month1
						month1= month1+gap
						rev_effective = rev_effective + date1 + "\n"
						if(request.POST['selec_Date_Cyc_2_review'] != ""):

							rev_selection2 = rev_selection2 + sel_date + "\n"
						#rev_selection1 = rev_selection1 + selection_date1 + ", "
						else:
							rev_selection2 = rev_selection2
							#rev_selection1 = rev_selection1
						rev_freeze = rev_freeze + freeze_date + "\n"
						rev_announce = rev_announce + announce_date + "\n"
						rev_clientcomm = rev_clientcomm + client_comm_date + "\n"
						rev_qc_date = rev_qc_date + Quality_date + "\n"
						#rev_commiteecomm = rev_commiteecomm + Committee_comm_date + "\n"
						rev_comm_cal_date = rev_comm_cal_date + comm_cal_date + "\n"
				if(request.POST['rebalance_month'] != ""):
					eff_rule_rebal = int(request.POST['effective_date_rebal'])
					fre_prior_day_rebal = int(request.POST['fre_prior_days_rebal']) if(request.POST['fre_prior_days_rebal']!="") else request.POST['fre_prior_days_rebal']
					fre_rule_rebal = int(request.POST['weights_Share_Freeze_rebal'])
					ann_rule_rebal = int(request.POST['public_Announcement_rebal'])
					client_comm_rule_rebal = int(request.POST['client_Comm_rebal'])
					if(request.POST['rebalance']=="Annual" or request.POST['rebalance']=="Monthly"):
						num_month = 1
						gap=1
					elif(request.POST['rebalance']== "Quarterly"):
						num_month = 4
						gap = 3
					else:
						num_month = 2
						gap = 6
					months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
						  October=10, November=11, December=12)
					month1 = months[request.POST['rebalance_month']]
					month1 = int(month1)
					cal_month = month1
					year = date.today().year
					for i in range(0,num_month):
						if(month1>12):
							month1 = month1%12;
						if(month1<cal_month):
							year = year+1
						date1 = EffectiveDate_Calculate(month1, year, eff_rule_rebal,holidays)
						eff_day = int(date1[3:5])
						eff_mon = int(date1[0:2])
						eff_year = int(date1[6:8])
						fre_prior_day_rebal = int(request.POST['fre_prior_days_rebal']) if(request.POST['fre_prior_days_rebal']!="") else ""
						if(fre_rule_rebal==6):
							freeze_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,fre_prior_day_rebal)
						else:
							freeze_date = FreezeDate_Calculate(eff_day, eff_mon,eff_year,fre_rule_rebal,holidays)
						freeze_day = int(freeze_date[3:5])
						freeze_mon = int(freeze_date[0:2])
						freeze_year = int(freeze_date[6:8])
						ann_prior_day_rebal = int(request.POST['ann_prior_days_rebal']) if(request.POST['ann_prior_days_rebal']!="") else request.POST['ann_prior_days_rebal']
						if(ann_rule_rebal==3):
							announce_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,ann_prior_day_rebal)
						else:
							if (ann_rule_rebal == 1):
								announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rebal,holidays)
							else:
								announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rebal,holidays)
						announce_day = int(announce_date[3:5])
						announce_mon = int(announce_date[0:2])
						announce_year = int(announce_date[6:8])
						cl_prior_day_rebal = int(request.POST['cl_prior_days_rebal']) if(request.POST['cl_prior_days_rebal']!="") else request.POST['cl_prior_days_rebal']
						if(client_comm_rule_rebal ==3):
							client_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,cl_prior_day_rebal)
						else:
							if (client_comm_rule_rebal == 1):
								client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, client_comm_rule_rebal,holidays)
							else:
								client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, client_comm_rule_rebal,holidays)
							Client_comm_day = int(client_comm_date[3:5])
							Client_comm_mon = int(client_comm_date[0:2])
							Client_comm_year = int(client_comm_date[6:8])
						comm_cal_prior_day_rebal = int(request.POST['cal_ag_prior_days_rebal']) if(request.POST['cal_ag_prior_days_rebal']!="") else request.POST['cal_ag_prior_days_rebal']
						comm_cal_rule_rebal = int(request.POST['comm_to_Calc_Agent_rebal'])
						if(comm_cal_rule_rebal==2):
							comm_cal_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,comm_cal_prior_day_rebal)
						else:
							comm_cal_date = client_comm_date

						qc_prior_day_rebal= int(request.POST['qc_prior_days_rebal']) if(request.POST['qc_prior_days_rebal']!="") else request.POST['qc_prior_days_rebal']
						qc_rule_rebal = int(request.POST['qc_Date_rebal'])
						if(qc_rule_rebal==3):
							qc_date_rebal = priorday_calculate(eff_day,eff_mon,eff_year,holidays,qc_prior_day_rebal)
						else:
							qc_date_rebal = freeze_date
						cal_month = month1
						month1= month1+gap
						re_effective = re_effective + date1 + "\n"
						re_freeze = re_freeze + freeze_date + "\n"
						re_announce = re_announce + announce_date + "\n"
						re_clientcomm = re_clientcomm + client_comm_date + "\n"
						re_qc = re_qc + qc_date_rebal + "\n"
						re_comm_cal_date = re_comm_cal_date + comm_cal_date + "\n"


			else:

				color_1 = 0

				reconst_Effective_Date = request.POST['mnl_eff_date'] if(request.POST['mnl_eff_date']=="") else datetime.strptime(request.POST['mnl_eff_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Selection_Date_Cycle_2 = request.POST['mnl_sel2_date'] if(request.POST['mnl_sel2_date']=="") else datetime.strptime(request.POST['mnl_sel2_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Weights_Share_Freeze_Date = request.POST['mnl_freeze_date'] if(request.POST['mnl_freeze_date']=="") else datetime.strptime(request.POST['mnl_freeze_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Public_Announcement_Date = request.POST['mnl_pb_announce_date'] if(request.POST['mnl_pb_announce_date']=="") else datetime.strptime(request.POST['mnl_pb_announce_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Client_Comm_Date = request.POST['mnl_client_comm_date'] if(request.POST['mnl_client_comm_date']=="") else datetime.strptime(request.POST['mnl_client_comm_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Prelim_Comm_Date = request.POST['mnl_prelim_date'] if(request.POST['mnl_prelim_date']=="") else datetime.strptime(request.POST['mnl_prelim_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Completion_Date = request.POST['mnl_cmp_date'] if(request.POST['mnl_cmp_date']=="") else datetime.strptime(request.POST['mnl_cmp_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_comm_cal_date = request.POST['mnl_comm_cal_date'] if(request.POST['mnl_comm_cal_date']=="") else datetime.strptime(request.POST['mnl_comm_cal_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				#reconst_Ind_Cmte_Comm_Date = request.POST['mnl_ind_Comm_date'] if(request.POST['mnl_ind_Comm_date']=="") else datetime.strptime(request.POST['mnl_ind_Comm_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Selection_Date_Cycle_1 =request.POST['mnl_sel1_date'] if(request.POST['mnl_sel1_date']=="") else datetime.strptime(request.POST['mnl_sel1_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_qc_date = request.POST['mnl_qc_date'] if(request.POST['mnl_qc_date']=="") else datetime.strptime(request.POST['mnl_qc_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_effective = request.POST['mnl_eff_date_review'] if(request.POST['mnl_eff_date_review']=="") else datetime.strptime(request.POST['mnl_eff_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_announce = request.POST['mnl_pb_announce_date_review'] if(request.POST['mnl_pb_announce_date_review']=="") else datetime.strptime(request.POST['mnl_pb_announce_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_clientcomm = request.POST['mnl_client_comm_date_review'] if(request.POST['mnl_client_comm_date_review']=="") else datetime.strptime(request.POST['mnl_client_comm_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				#rev_commiteecomm = request.POST['mnl_ind_Comm_date_review'] if(request.POST['mnl_ind_Comm_date_review']=="") else datetime.strptime(request.POST['mnl_ind_Comm_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_qc_date = request.POST['mnl_qc_date_review'] if(request.POST['mnl_qc_date_review']=="") else datetime.strptime(request.POST['mnl_qc_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				#rev_completion = request.POST['mnl_cmp_date_review'] if(request.POST['mnl_cmp_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_cmp_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
				#rev_prelim = request.POST['mnl_prelim_date_review'] if(request.POST['mnl_prelim_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_prelim_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
				rev_freeze = request.POST['mnl_freeze_date_review'] if(request.POST['mnl_freeze_date_review']=="") else datetime.strptime(request.POST['mnl_freeze_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_selection2 = request.POST['mnl_sel2_date_review'] if(request.POST['mnl_sel2_date_review']=="") else datetime.strptime(request.POST['mnl_sel2_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_announce = request.POST['mnl_pb_announce_date_rebal'] if(request.POST['mnl_pb_announce_date_rebal']=="") else datetime.strptime(request.POST['mnl_pb_announce_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_clientcomm = request.POST['mnl_client_comm_date_rebal'] if(request.POST['mnl_client_comm_date_rebal']=="") else datetime.strptime(request.POST['mnl_client_comm_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_effective = request.POST['mnl_eff_date_rebal'] if(request.POST['mnl_eff_date_rebal']=="") else datetime.strptime(request.POST['mnl_eff_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_freeze = request.POST['mnl_freeze_date_rebal'] if(request.POST['mnl_freeze_date_rebal']=="") else datetime.strptime(request.POST['mnl_freeze_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_qc = request.POST['mnl_qc_date_rebal'] if(request.POST['mnl_qc_date_rebal']=="") else datetime.strptime(request.POST['mnl_qc_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_comm_cal_date = request.POST['mnl_comm_cal_date_rebal'] if(request.POST['mnl_comm_cal_date_rebal']=="") else datetime.strptime(request.POST['mnl_comm_cal_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_comm_cal_date = request.POST['mnl_comm_cal_date_review'] if(request.POST['mnl_comm_cal_date_review']=="") else datetime.strptime(request.POST['mnl_comm_cal_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")


			index_data = Registerindex(
				Ident_ISIN=request.POST['isin_id'],
				Ident_Bloomberg=request.POST['bloomberg_id'],
				Ident_Reuters=request.POST['thomson_id'],
				Index_Name=request.POST['index_name'],
				Client_Name_id=request.POST['client_name'],
				Ind_Sty=request.POST['index_Style'],
				Ind_Ver=request.POST['ind_version'],
				Ind_Ver_ID="",
				Calc=request.POST['calculation'],
				Calc_Agent=request.POST['cal_agent']if(request.POST['cal_agent_des']=="") else request.POST['cal_agent_des'],
				Data_Platform=request.POST['data_platform'],
				Data_Vendors=request.POST['data_vendors'],
				Contract_Type=request.POST['contract_Type'],
				Type_of_Ind=request.POST['type_index'],
				Product_Status=request.POST['prod_Status'],
				ETF_Launched=request.POST['etf_Launched'],
				Reconstitution=request.POST['reconstitution'],
				Rebalance=request.POST['rebalance'],
				Review=request.POST['review'],
				Theme_Review=request.POST['theme_Review'],
				Selection_Date_Cycle_1=reconst_Selection_Date_Cycle_1,
				Selection_Date_Cycle_1_rebal=re_selection1,
				Selection_Date_Cycle_1_review=rev_selection1,
				Completion_Date=reconst_Completion_Date,
				Completion_Date_rebal=re_completion,
				Completion_Date_review="",
				Selection_Date_Cycle_2=reconst_Selection_Date_Cycle_2,
				Selection_Date_Cycle_2_rebal=re_selection2,
				Selection_Date_Cycle_2_review=rev_selection2,
				#Ind_Cmte_Comm_Date=reconst_Ind_Cmte_Comm_Date,
				Ind_Cmte_Comm_Date_rebal="",
				#Ind_Cmte_Comm_Date_review=rev_commiteecomm,
				Prelim_Comm_Date=reconst_Prelim_Comm_Date,
				Prelim_Comm_Date_rebal="",
				Prelim_Comm_Date_review="",
				Weights_Share_Freeze_Date=reconst_Weights_Share_Freeze_Date,
				Weights_Share_Freeze_Date_rebal=re_freeze,
				QC_Date_rebal=re_qc,
				QC_Date=reconst_qc_date,
				QC_Date_review=rev_qc_date,
				Weights_Share_Freeze_Date_review=rev_freeze,
				Public_Announcement_Date=reconst_Public_Announcement_Date,
				Public_Announcement_Date_rebal=re_announce,
				Public_Announcement_Date_review=rev_announce,
				Client_Comm_Date=reconst_Client_Comm_Date,
				Client_Comm_Date_rebal=re_clientcomm,
				Client_Comm_Date_review=rev_clientcomm,
				Effective_Date=reconst_Effective_Date,
				Effective_Date_rebal=re_effective,
				Effective_Date_review=rev_effective,
				reconst_month=request.POST['reconst_month'],
				rebalance_month=request.POST['rebalance_month'],
				review_month=request.POST['review_month'],
				Comm_to_Calc_Agent=reconst_comm_cal_date,
				Comm_to_Calc_Agent_rebal=re_comm_cal_date,
				Comm_to_Calc_Agent_review=rev_comm_cal_date,
				color_code=color_1,
				live_date=request.POST['live_date'],
				etf_launch_date=request.POST['etf_date'],
				backtest_comp_date=request.POST['backtest_date'],
			)

			index_data.save()
			rule_data = Ruleslist(
				eff_rule = request.POST['effective_date'],
				eff_rule_rebal = request.POST['effective_date_rebal'],
				eff_rule_review = request.POST['effective_date_review'],
				sel_rule1 = request.POST['selec_Date_Cyc_1'],
				#sel_rule1_rebal = request.POST['selec_Date_Cyc_1_rebal'],
				#sel_rule1_review = request.POST['selec_Date_Cyc_1_review'],
				sel_rule2 = request.POST['selec_Date_Cyc_2'],
				#sel_rule2_rebal = request.POST['selec_Date_Cyc_2_rebal'],
				sel_rule2_review = request.POST['selec_Date_Cyc_2_review'],
				announce_rule = request.POST['public_Announcement'],
				announce_rule_rebal = request.POST['public_Announcement_rebal'],
				announce_rule_review = request.POST['public_Announcement_review'],
				prelim_rule = request.POST['pre_comm_date'],
				#prelim_rule_rebal = request.POST['pre_comm_date_rebal'],
				#prelim_rule_review = request.POST['pre_comm_date_review'],
				clientcomm_rule = request.POST['client_Comm'],
				clientcomm_rule_rebal = request.POST['client_Comm_rebal'],
				clientcomm_rule_review = request.POST['client_Comm_review'],
				#indcommittee_rule = request.POST['ind_Comm_Date'],
				#indcommittee_rule_rebal = request.POST['ind_Comm_Date_rebal'],
				#indcommittee_rule_review = request.POST['ind_Comm_Date_review'],
				freeze_rule = request.POST['weights_Share_Freeze'],
				freeze_rule_rebal = request.POST['weights_Share_Freeze_rebal'],
				qc_rule_rebal = request.POST['qc_Date_rebal'],
				qc_rule = request.POST['qc_date'],
				qc_rule_review = request.POST['ind_qc_date_review'],
				freeze_rule_review = request.POST['weights_Share_Freeze_review'],
				comp_rule = request.POST['cmp_Date'],
				#comp_rule_rebal = request.POST['cmp_Date_rebal'],
				#comp_rule_review = request.POST['cmp_Date_review'],
				comm_cal_rule = request.POST['comm_to_Calc_Agent'],
				comm_cal_rule_rebal = request.POST['comm_to_Calc_Agent_rebal'],
				comm_cal_rule_review = request.POST['comm_to_Calc_Agent_review'],
				index_id = index_data.id,

			)
			rule_data.save()

			prior_days = Priordays(

	            sel_cycle_1_day_recon = int(request.POST['sel1_prior_days']) if(request.POST['sel1_prior_days']!="") else request.POST['sel1_prior_days'],
				comp_cycle_1_day_recon = int(request.POST['cmp_prior_days']) if(request.POST['cmp_prior_days']!="") else request.POST['cmp_prior_days'],
				sel_cycle_2_day_recon = int(request.POST['sel2_prior_days']) if(request.POST['sel2_prior_days']!="")else request.POST['sel2_prior_days'],
			    prelim_day_recon = int(request.POST['pre_prior_days']) if(request.POST['pre_prior_days']!="")else request.POST['pre_prior_days'],
			    freeze_day_recon = int(request.POST['fre_prior_days']) if(request.POST['fre_prior_days']!="")else request.POST['fre_prior_days'],
	            #indcommittee_day_recon = int(request.POST['icom_prior_days']) if(request.POST['icom_prior_days']!="") else request.POST['icom_prior_days'],
	            qc_day_recon = int(request.POST['qc_prior_days']) if(request.POST['qc_prior_days']!="") else request.POST['qc_prior_days'],
			    announce_day_recon = int(request.POST['ann_prior_days']) if(request.POST['ann_prior_days']!="")else request.POST['ann_prior_days'],
	            clientcomm_day_recon = int(request.POST['cl_prior_days']) if(request.POST['cl_prior_days']!="")else request.POST['cl_prior_days'],
	            comm_cal_agent_day_recon = int(request.POST['cal_ag_prior_days']) if(request.POST['cal_ag_prior_days']!="") else request.POST['cal_ag_prior_days'],
	            freeze_day_rebal = int(request.POST['fre_prior_days_rebal']) if(request.POST['fre_prior_days_rebal']!="")else request.POST['fre_prior_days_rebal'],
	            qc_day_rebal = int(request.POST['qc_prior_days_rebal']) if(request.POST['qc_prior_days_rebal']!="")else request.POST['qc_prior_days_rebal'],
	            announce_day_rebal = int(request.POST['ann_prior_days_rebal']) if(request.POST['ann_prior_days_rebal']!="")else request.POST['ann_prior_days_rebal'],
	            clientcomm_day_rebal = int(request.POST['cl_prior_days_rebal']) if(request.POST['cl_prior_days_rebal']!="")else request.POST['cl_prior_days_rebal'],
	            comm_cal_agent_day_rebal = int(request.POST['cal_ag_prior_days_rebal']) if(request.POST['cal_ag_prior_days_rebal']!="") else request.POST['cal_ag_prior_days_rebal'],
	            sel_cycle_2_day_review = int(request.POST['sel2_prior_days_review']) if(request.POST['sel2_prior_days_review']!="") else request.POST['sel2_prior_days_review'],
	            #indcommittee_day_review = int(request.POST['icom_prior_days_review']) if(request.POST['icom_prior_days_review']!="") else request.POST['icom_prior_days_review'],
	            qc_day_review = int(request.POST['qc_prior_days_review']) if(request.POST['qc_prior_days_review']!="") else request.POST['qc_prior_days_review'],
	            freeze_day_review = int(request.POST['fre_prior_days_review']) if(request.POST['fre_prior_days_review']!="") else request.POST['fre_prior_days_review'],
	            announce_day_review = int(request.POST['ann_prior_days_review']) if(request.POST['ann_prior_days_review']!="") else request.POST['ann_prior_days_review'] ,
	            clientcomm_day_review = int(request.POST['cl_prior_days_review']) if(request.POST['cl_prior_days_review']!="") else request.POST['cl_prior_days_review'] ,
	            comm_cal_agent_day_review = int(request.POST['cal_ag_prior_days_review']) if(request.POST['cal_ag_prior_days_review']!="") else request.POST['cal_ag_prior_days_review'],
				index_id = index_data.id,
			)
			prior_days.save()
			template = "postindex.html"
			context = {"message": "success",}
			return render(request, template, context)
	else:
		context = {"message": "index not saved!"}
		template = "index.html"
		return render(request, template, context)


def view_index(request, id):
	templateName = "ViewIndex.html"
	in_id = Registerindex.objects.filter(id = id)
	clientList = Calendarlist.objects.filter(category = 'Client_Name')
	context = {
		'in_id' : in_id,'clientList' : clientList,
	}
	return render(request, templateName, context)


def search_index(request):
	if request.method == 'POST':
		search_result = request.POST['search2']
		results = []
		if(search_result!=""):
			templatename = "SearchIndex.html"
			lookups = Q(Index_Name__icontains = search_result) | Q(Contract_Type__icontains = search_result) | Q(Reconstitution__icontains = search_result) | Q(Calc_Agent__icontains = search_result) | Q(Ind_Sty__icontains = search_result) | Q(Contract_Type__icontains = search_result) | Q(Type_of_Ind__icontains = search_result)
			results= Registerindex.objects.filter(lookups).distinct().order_by('id').reverse()
			if(len(results) == 0):
				obj = Calendarlist.objects.filter(description = search_result)
				code = ""
				for ind in obj:
					code = int(ind.code)
					if(code!=""):
						results = Registerindex.objects.filter(Client_Name_id = code).order_by('id').reverse()
					else:
						templatename = "SearchIndex.html"
		else:
			templatename = "Nosearch.html"


		clientList = Calendarlist.objects.filter(category = 'Client_Name')
		page = request.GET.get('page',1)
		paginator = Paginator(results, 20)
		try:
			indexdata = paginator.page(page)
		except PageNotAnInteger:
			indexdata = paginator.page(1)
		except EmptyPage:
			indexdata = paginator.page(paginator.num_pages)
		context={
			'search_text': search_result,
			'indexdata': indexdata,
			'clientList': clientList,
		}
		return render(request, templatename, context)

	else:
		templatename = "SearchIndex.html"
		return render(request, templatename)


def report(request):
	templateName = "report.html"
	client_name = Calendarlist.objects.filter(category = 'Client_Name')
	info = Registerindex.objects.all().order_by('id').reverse()
	page = request.GET.get('page',1)
	paginator = Paginator(info, 10)
	try:
		infodata = paginator.page(page)
	except PageNotAnInteger:
		infodata = paginator.page(1)
	except EmptyPage:
		infodata = paginator.page(paginator.num_pages)
	now = dt.datetime.now()
	context = {
		'client_name': client_name,
		'info': infodata,
		'curr_year': now.year
	}
	return render(request, templateName, context)

def back(request):
    templates = loader.get_template("home.html");
    return HttpResponse(templates.render());


def edit_index(request, id):
	templateName = "EditIndex.html"
	in_id = Registerindex.objects.filter(id=id)

	rule = Ruleslist.objects.filter(index = id)
	try:
		priordays = Priordays.objects.get(index_id=id)
	except Priordays.DoesNotExist:
		priordays = None
	client_name = Calendarlist.objects.filter(category='Client_Name')
	ind_Sty = Calendarlist.objects.filter(category='Ind_Sty')
	ind_Version = Calendarlist.objects.filter(category='Ind_Ver')
	cal = Calendarlist.objects.filter(category='Cal')
	cal_Agent = Calendarlist.objects.filter(category='Cal_Agent')
	contract_Type = Calendarlist.objects.filter(category='Contract_Type')
	type_of_Ind = Calendarlist.objects.filter(category='Type_of_Ind')
	prod_Stat = Calendarlist.objects.filter(category='Prod_Stat')
	etf_Launched = Calendarlist.objects.filter(category='ETF_Launched')
	reconstitution = Calendarlist.objects.filter(category='Reconstitution')
	rebalance = Calendarlist.objects.filter(category='Rebalance')
	review = Calendarlist.objects.filter(category='Review')
	theme_Review = Calendarlist.objects.filter(category='Theme_Review')
	comp_Date = Calendarlist.objects.filter(category='Comp_Date')
	selec_Date_Cyc_1 = Calendarlist.objects.filter(category='Selec_Date_Cyc_1')
	selec_Date_Cyc_2 = Calendarlist.objects.filter(category='Selec_Date_Cyc_2')
	weights_Share_Freeze = Calendarlist.objects.filter(category='Weights_Share_Freeze_Date')
	public_Announcement = Calendarlist.objects.filter(category='Public_Announcement_Date')
	client_Comm = Calendarlist.objects.filter(category='Client_Comm_Date')
	ind_Cmte_Comm_Date = Calendarlist.objects.filter(category='Ind_Cmte_Comm_Date')
	effec_Date = Calendarlist.objects.filter(category='Effec_Date')
	QC_Date = Calendarlist.objects.filter(category='QC_Date')
	prelim_Comm_Date = Calendarlist.objects.filter(category='Prelim_Comm_Date')
	comm_to_Calc_Agent = Calendarlist.objects.filter(category='Comm_to_Calc_Agent')
	for ind1 in in_id:
		if(ind1.color_code) == 0:
			ind1.Completion_Date= datetime.strptime(ind1.Completion_Date,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Completion_Date!="") else ""
			ind1.Selection_Date_Cycle_2= datetime.strptime(ind1.Selection_Date_Cycle_2,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Selection_Date_Cycle_2!="") else ""
			ind1.Selection_Date_Cycle_1= datetime.strptime(ind1.Selection_Date_Cycle_1,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Selection_Date_Cycle_1!="") else ""
			#ind1.Ind_Cmte_Comm_Date= datetime.strptime(ind1.Ind_Cmte_Comm_Date,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Ind_Cmte_Comm_Date!="") else ""
			ind1.Prelim_Comm_Date= datetime.strptime(ind1.Prelim_Comm_Date,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Prelim_Comm_Date!="") else ""
			ind1.Weights_Share_Freeze_Date=datetime.strptime(ind1.Weights_Share_Freeze_Date,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Weights_Share_Freeze_Date!="") else ""
			ind1.QC_Date=datetime.strptime(ind1.QC_Date,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.QC_Date!="") else ""
			ind1.Public_Announcement_Date=datetime.strptime(ind1.Public_Announcement_Date,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Public_Announcement_Date!="") else ""
			ind1.Client_Comm_Date=datetime.strptime(ind1.Client_Comm_Date,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Client_Comm_Date!="") else ""
			ind1.Effective_Date=datetime.strptime(ind1.Effective_Date,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Effective_Date!="") else ""
			ind1.Comm_to_Calc_Agent=datetime.strptime(ind1.Comm_to_Calc_Agent,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Comm_to_Calc_Agent!="") else ""
			ind1.Comm_to_Calc_Agent_rebal=datetime.strptime(ind1.Comm_to_Calc_Agent_rebal,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Comm_to_Calc_Agent_rebal!="") else ""
			ind1.Comm_to_Calc_Agent_review=datetime.strptime(ind1.Comm_to_Calc_Agent_review,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Comm_to_Calc_Agent_review!="") else ""
			ind1.Weights_Share_Freeze_Date_rebal=datetime.strptime(ind1.Weights_Share_Freeze_Date_rebal,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Weights_Share_Freeze_Date_rebal!="") else ""
			ind1.QC_Date_rebal=datetime.strptime(ind1.QC_Date_rebal,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.QC_Date_rebal!="") else ""
			ind1.Public_Announcement_Date_rebal=datetime.strptime(ind1.Public_Announcement_Date_rebal,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Public_Announcement_Date_rebal!="") else ""
			ind1.Client_Comm_Date_rebal=datetime.strptime(ind1.Client_Comm_Date_rebal,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Client_Comm_Date_rebal!="") else ""
			ind1.Effective_Date_rebal=datetime.strptime(ind1.Effective_Date_rebal,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Effective_Date_rebal!="") else ""
			ind1.Effective_Date_review=datetime.strptime(ind1.Effective_Date_review,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Effective_Date_review!="") else ""
			ind1.Selection_Date_Cycle_2_review= datetime.strptime(ind1.Selection_Date_Cycle_2_review,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Selection_Date_Cycle_2_review!="") else ""
			#ind1.Ind_Cmte_Comm_Date_review= datetime.strptime(ind1.Ind_Cmte_Comm_Date_review,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Ind_Cmte_Comm_Date_review!="") else ""
			ind1.Weights_Share_Freeze_Date_review=datetime.strptime(ind1.Weights_Share_Freeze_Date_review,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Weights_Share_Freeze_Date_review!="") else ""
			ind1.QC_Date_review=datetime.strptime(ind1.QC_Date_review,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.QC_Date_review!="") else ""
			ind1.Public_Announcement_Date_review=datetime.strptime(ind1.Public_Announcement_Date_review,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Public_Announcement_Date_review!="") else ""
			ind1.Client_Comm_Date_review=datetime.strptime(ind1.Client_Comm_Date_review,"%m/%d/%y").strftime("%Y-%m-%d") if(ind1.Client_Comm_Date_review!="") else ""





	context = {
		'in_id': in_id,
		'priordays': priordays,
		'client_name': client_name,
		'ind_Sty': ind_Sty,
		'ind_Version': ind_Version,
		'cal': cal,
		'cal_Agent': cal_Agent,
		'contract_Type': contract_Type,
		'type_of_Ind': type_of_Ind,
		'prod_Stat': prod_Stat,
		'etf_Launched': etf_Launched,
		'reconstitution': reconstitution,
		'rebalance': rebalance,
		'review': review,
		'theme_Review': theme_Review,
		'comp_Date': comp_Date,
		'selec_Date_Cyc_1': selec_Date_Cyc_1,
		'selec_Date_Cyc_2': selec_Date_Cyc_2,
		'weights_Share_Freeze': weights_Share_Freeze,
		'public_Announcement': public_Announcement,
		'client_Comm': client_Comm,
		#'ind_Cmte_Comm_Date': ind_Cmte_Comm_Date,
		'effec_Date': effec_Date,
		'prelim_Comm_Date': prelim_Comm_Date,
		'comm_to_Calc_Agent': comm_to_Calc_Agent,
		'rule':rule,
		'QC_Date':QC_Date,
	}

	return render(request, templateName, context)

def update_index(request, id):
	inst = USTradingCalendar()
	y1 = date.today().year
	holidays = inst.holidays(dt.datetime(y1-1, 9, 30), dt.datetime(y1, 12, 31))
	in_id = Registerindex.objects.filter(id=id)
	index1 = Ruleslist.objects.filter(index = id)
	priorday = Priordays.objects.filter(index=id)
	clientList = Calendarlist.objects.filter(category='Client_Name')
	for ind in in_id:
		if request.method == 'POST':

			check1 = request.POST.get('check1', False)
			check2 = request.POST.get('check2', False)
			check3 = request.POST.get('check3', False)
			choice1 = request.POST.get('choice', False)

			"""	
			if(check3=="rev" and choice1== "rule"):
				if(request.POST['review']== "" or request.POST['review_month']== "" or request.POST['effective_date_review']== ""
			  or request.POST['selec_Date_Cyc_2_review']== "" or request.POST['ind_Comm_Date_review']== "" or request.POST['weights_Share_Freeze_review']== ""
			or request.POST['public_Announcement_review']== "" or request.POST['client_Comm_review']== "" or request.POST['comm_to_Calc_Agent_review']== ""):
					templatename = "600.html"
					return render(request, templatename)
			elif(check3=="rev" and choice1=="date"):
				if(request.POST['review']== "" or request.POST['review_month']== "" or request.POST['mnl_eff_date_review']== ""
			  or request.POST['mnl_sel2_date_review']== "" or request.POST['mnl_ind_Comm_date_review']== "" or request.POST['mnl_freeze_date_review']== ""
			or request.POST['mnl_pb_announce_date_review']== "" or request.POST['mnl_client_comm_date_review']== "" or request.POST['mnl_comm_cal_date_review']== ""):
					templatename = "600.html"
					return render(request, templatename)
			"""
			reconst_Effective_Date = ""
			reconst_Selection_Date_Cycle_2 = ""
			reconst_Weights_Share_Freeze_Date = ""
			reconst_Public_Announcement_Date = ""
			reconst_Client_Comm_Date = ""
			reconst_Prelim_Comm_Date = ""
			reconst_Completion_Date = ""
			reconst_qc_date = ""
			#reconst_Ind_Cmte_Comm_Date = ""
			reconst_comm_cal_date = ""
			re_comm_cal_date=""
			rev_comm_cal_date= ""
			reconst_Selection_Date_Cycle_1 = ""
			rev_effective = ""
			rev_selection2 = ""
			rev_freeze = ""
			rev_announce = ""
			rev_clientcomm = ""
			rev_completion = ""
			rev_qc_date = ""
			rev_commiteecomm = ""
			rev_selection1 = ""
			re_effective = ""
			re_selection2 = ""
			re_freeze = ""
			re_announce = ""
			re_clientcomm = ""
			re_prelim = ""
			re_completion = ""
			re_qc = ""
			re_selection1 = ""
			color_1 = 1

			if(choice1=="rule"):
				if(request.POST['effective_date'] != ""):
					if((request.POST['reconstitution']=="Annual") or (request.POST['reconstitution']=="Monthly")):
						num_month = 1
						gap=1
					elif(request.POST['reconstitution']== "Quarterly"):
						num_month = 4
						gap = 3
					else:
						num_month = 2
						gap = 6
					months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
							  October=10, November=11, December=12)
					month1 = months[request.POST['reconst_month']]
					month1 = int(month1)
					cal_month = month1
					year = date.today().year
					for i in range(0,num_month):
						if(month1>12):
							month1 = month1%12;
						if(month1<cal_month):
							year = year+1
						eff_rule = int(request.POST['effective_date'])
						sel_rule = int(request.POST['selec_Date_Cyc_2'])
						fre_rule = int(request.POST['weights_Share_Freeze'])
						ann_rule = int(request.POST['public_Announcement'])
						client_comm_rule = int(request.POST['client_Comm'])
						#icom_rule = int(request.POST['ind_Comm_Date'])
						prelim_comm_rule = int(request.POST['pre_comm_date']) if(request.POST['pre_comm_date']!="") else request.POST['pre_comm_date']
						date1 = EffectiveDate_Calculate(month1, year, eff_rule,holidays)
						eff_day = int(date1[3:5])
						eff_mon = int(date1[0:2])
						eff_year = int(date1[6:8])
						sel2_prior_day = int(request.POST['sel2_prior_days']) if(request.POST['sel2_prior_days']!="")else request.POST['sel2_prior_days']
						if(sel_rule==7):
							sel_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,sel2_prior_day)
						else:
						    sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule,holidays)
						sel_day = int(sel_date[3:5])
						sel_month = int(sel_date[0:2])
						sel_year = int(sel_date[6:8])
						fre_prior_day = int(request.POST['fre_prior_days']) if(request.POST['fre_prior_days']!="") else request.POST['fre_prior_days']
						if(fre_rule==6):

							freeze_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,fre_prior_day)
						else:
						    freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule,holidays)
						freeze_day = int(freeze_date[3:5])
						freeze_mon = int(freeze_date[0:2])
						freeze_year = int(freeze_date[6:8])
						ann_prior_day = int(request.POST['ann_prior_days']) if(request.POST['ann_prior_days']!="") else request.POST['ann_prior_days']
						if(ann_rule==3):
							announce_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,ann_prior_day)
						else:
							if (ann_rule == 1):
								announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule,holidays)
							else:
								announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule,holidays)
						announce_day = int(announce_date[3:5])
						announce_mon = int(announce_date[0:2])
						announce_year = int(announce_date[6:8])
						cl_prior_day = int(request.POST['cl_prior_days']) if(request.POST['cl_prior_days']!="") else request.POST['cl_prior_days']
						if(client_comm_rule==3):
							client_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,cl_prior_day)
						else:
							if (client_comm_rule == 1):
								client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, client_comm_rule,holidays)
							else:
								client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year,client_comm_rule,holidays)
						Client_comm_day = int(client_comm_date[3:5])
						Client_comm_mon = int(client_comm_date[0:2])
						Client_comm_year = int(client_comm_date[6:8])

						comm_cal_prior_day = int(request.POST['cal_ag_prior_days']) if(request.POST['cal_ag_prior_days']!="") else request.POST['cal_ag_prior_days']
						comm_cal_rule = int(request.POST['comm_to_Calc_Agent'])
						if(comm_cal_rule==2):
							comm_cal_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,comm_cal_prior_day)
						else:
							comm_cal_date = client_comm_date
						pre_prior_day = int(request.POST['pre_prior_days']) if(request.POST['pre_prior_days']!="") else request.POST['pre_prior_days']
						if(prelim_comm_rule != ""):
							if(prelim_comm_rule==2):

								Prelim_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,pre_prior_day)
							else:
								Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day, announce_mon, announce_year,holidays)
							Prelim_comm_day = int(Prelim_comm_date[3:5])
							Prelim_comm_mon = int(Prelim_comm_date[0:2])
							Prelim_comm_year = int(Prelim_comm_date[6:8])
						else:
							Prelim_comm_date=""
						"""
						icom_prior_day = int(request.POST['icom_prior_days']) if(request.POST['icom_prior_days']!="") else request.POST['icom_prior_days']
						if (icom_rule==2):
							Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon,
																		   Client_comm_year,holidays)
						else:
							if(icom_rule==3):
								Committee_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,icom_prior_day)
							else:
							    Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day, Prelim_comm_mon,
																		   Prelim_comm_year,holidays)
						Committee_comm_day = int(Committee_comm_date[3:5])
						Committee_comm_mon = int(Committee_comm_date[0:2])
						Committee_comm_year = int(Committee_comm_date[6:8])
						"""
						qc_rule = int(request.POST['qc_date'])
						qc_prior_day = int(request.POST['qc_prior_days']) if(request.POST['qc_prior_days']!="") else request.POST['qc_prior_days']
						if(qc_rule==3):
							Quality_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,qc_prior_day)
						else:
							Quality_date = QualityDate_Calculate(freeze_day,freeze_mon,freeze_year,holidays)
						cmp_prior_day = int(request.POST['cmp_prior_days']) if(request.POST['cmp_prior_days']!="") else request.POST['cmp_prior_days']

						comp_rule = int(request.POST['cmp_Date'])
						if(comp_rule==2):
							Completion_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,cmp_prior_day)
						else:
							Completion_date = CompletionDate_Calculate(sel_day, sel_month, sel_year,holidays)


						sel1_rule = int(request.POST['selec_Date_Cyc_1']) if(request.POST['selec_Date_Cyc_1']!="") else request.POST['selec_Date_Cyc_1']
						if(sel1_rule==2):
							sel1_prior_day = int(request.POST['sel1_prior_days']) if(request.POST['sel1_prior_days']!="") else request.POST['sel1_prior_days']
							selection_date1 = priorday_calculate(eff_day,eff_mon,eff_year,holidays,sel1_prior_day)
						elif(sel1_rule==1):
							selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year,holidays)
						else:
							sel1_prior_day=""
							selection_date1 =""

						reconst_Selection_Date_Cycle_2 = reconst_Selection_Date_Cycle_2 + sel_date + "\n"
						reconst_Completion_Date = reconst_Completion_Date + Completion_date + "\n"
						reconst_Selection_Date_Cycle_1 = reconst_Selection_Date_Cycle_1+selection_date1 + "\n"

						cal_month = month1
						month1= month1+gap
						reconst_Effective_Date = reconst_Effective_Date + date1 + "\n"
						reconst_Weights_Share_Freeze_Date = reconst_Weights_Share_Freeze_Date + freeze_date + "\n"
						reconst_Public_Announcement_Date = reconst_Public_Announcement_Date + announce_date + "\n"
						reconst_Client_Comm_Date = reconst_Client_Comm_Date + client_comm_date + "\n"
						reconst_comm_cal_date =  reconst_comm_cal_date + comm_cal_date + "\n"
						reconst_Prelim_Comm_Date = reconst_Prelim_Comm_Date + Prelim_comm_date + "\n"
						#reconst_Ind_Cmte_Comm_Date = reconst_Ind_Cmte_Comm_Date + Committee_comm_date + "\n"
						reconst_qc_date = reconst_qc_date + Quality_date + "\n"
				else:
					reconst_Effective_Date=""
					reconst_Weights_Share_Freeze_Date=""
					reconst_Public_Announcement_Date=""
					reconst_Client_Comm_Date=""
					reconst_comm_cal_date=""
					reconst_Prelim_Comm_Date=""
					#reconst_Ind_Cmte_Comm_Date=""
					reconst_qc_date=""
					reconst_Selection_Date_Cycle_2=""
					reconst_Completion_Date=""
					reconst_Selection_Date_Cycle_1=""
				if(request.POST['effective_date_review'] != ""):
					eff_rule_rev = int(request.POST['effective_date_review'])
					#print(eff_rule_rev)
					sel_rule_rev = int(request.POST['selec_Date_Cyc_2_review'])
					fre_rule_rev = int(request.POST['weights_Share_Freeze_review'])
					ann_rule_rev = int(request.POST['public_Announcement_review'])
					client_comm_rule_rev = int(request.POST['client_Comm_review'])
				#prelim_comm_rule_rev = int(request.POST['pre_comm_date_review'])
				#prelim_comm_rule_rev = int(request.POST['pre_comm_date_review'])
					if(request.POST['review']=="Annual" or request.POST['review']=="Monthly"):
						num_month = 1
						gap=1
					elif(request.POST['review']== "Quarterly"):
						num_month = 4
						gap = 3
					else:
						num_month = 2
						gap = 6
					months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
						  October=10, November=11, December=12)
					month1 = months[request.POST['review_month']]
					month1 = int(month1)
					cal_month = month1
					year = date.today().year
					for i in range(0,num_month):
						if(month1>12):
							month1 = month1%12;
						if(month1<cal_month):
							year = year+1
						date1 = EffectiveDate_Calculate(month1, year, eff_rule_rev,holidays)
						eff_day = int(date1[3:5])
						eff_mon = int(date1[0:2])
						eff_year = int(date1[6:8])
						sel2_prior_day_review = int(request.POST['sel2_prior_days_review']) if(request.POST['sel2_prior_days_review']!="") else request.POST['sel2_prior_days_review']
						if(sel_rule_rev==7):
							sel_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,sel2_prior_day_review)
						else:
							sel_date = SelectionDate_Calculate(eff_day, eff_mon, eff_year, sel_rule_rev,holidays)
						sel_day = int(sel_date[3:5])
						sel_month = int(sel_date[0:2])
						sel_year = int(sel_date[6:8])
						fre_prior_day_review = int(request.POST['fre_prior_days_review']) if(request.POST['fre_prior_days_review']!="") else request.POST['fre_prior_days_review']
						if(fre_rule_rev==6):
							freeze_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,fre_prior_day_review)
						else:
							freeze_date = FreezeDate_Calculate(eff_day, eff_mon, eff_year, fre_rule_rev,holidays)

						freeze_day = int(freeze_date[3:5])
						freeze_mon = int(freeze_date[0:2])
						freeze_year = int(freeze_date[6:8])
						ann_prior_day_review = int(request.POST['ann_prior_days_review']) if(request.POST['ann_prior_days_review']!="") else request.POST['ann_prior_days_review']
						if(ann_rule_rev==3):
						    announce_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,ann_prior_day_review)
						else:
							if (ann_rule_rev == 1):
								announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rev,holidays)
							else:
								announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rev,holidays)
						announce_day = int(announce_date[3:5])
						announce_mon = int(announce_date[0:2])
						announce_year = int(announce_date[6:8])
						cl_prior_day_review = int(request.POST['cl_prior_days_review']) if(request.POST['cl_prior_days_review']!="") else request.POST['cl_prior_days_review']
						if(client_comm_rule_rev==3):
						    client_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,cl_prior_day_review)
						else:
							if (client_comm_rule_rev == 1):
								client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, client_comm_rule_rev,holidays)
							else:
								client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, client_comm_rule_rev,holidays)
						Client_comm_day = int(client_comm_date[3:5])
						Client_comm_mon = int(client_comm_date[0:2])
						Client_comm_year = int(client_comm_date[6:8])
						comm_cal_rule_review = int(request.POST['comm_to_Calc_Agent_review'])
						comm_cal_prior_day_review = int(request.POST['cal_ag_prior_days_review']) if(request.POST['cal_ag_prior_days_review']!="") else request.POST['cal_ag_prior_days_review']
						if(comm_cal_rule_review==2):
							comm_cal_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,comm_cal_prior_day_review)
						else:
							comm_cal_date = client_comm_date
						"""
						icom_rule_review = int(request.POST['ind_Comm_Date_review'])
						icom_prior_day_review = int(request.POST['icom_prior_days_review']) if(request.POST['icom_prior_days_review']!="") else request.POST['icom_prior_days_review']
						if(icom_rule_review==3):

						     Committee_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,icom_prior_day_review)
						else:
							 Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day, Client_comm_mon, Client_comm_year,holidays)
						Committee_comm_day = int(Committee_comm_date[3:5])
						Committee_comm_mon = int(Committee_comm_date[0:2])
						Committee_comm_year = int(Committee_comm_date[6:8])
						"""
						qc_rule_rev = int(request.POST['ind_qc_date_review'])
						qc_prior_day_review = int(request.POST['qc_prior_days_review']) if(request.POST['qc_prior_days_review']!="") else request.POST['qc_prior_days_review']
						if(qc_rule_rev==3):
							Quality_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,qc_prior_day_review)
						else:
							Quality_date = QualityDate_Calculate(freeze_day,freeze_mon,freeze_year,holidays)

						cal_month = month1
						month1= month1+gap
						rev_effective = rev_effective + date1 + "\n"
						if(request.POST['selec_Date_Cyc_2_review'] != ""):

							rev_selection2 = rev_selection2 + sel_date + "\n"
						#rev_selection1 = rev_selection1 + selection_date1 + ", "
						else:
							rev_selection2 = rev_selection2
							#rev_selection1 = rev_selection1
						rev_freeze = rev_freeze + freeze_date + "\n"
						rev_announce = rev_announce + announce_date + "\n"
						rev_clientcomm = rev_clientcomm + client_comm_date + "\n"
						rev_qc_date = rev_qc_date + Quality_date + "\n"
						#rev_commiteecomm = rev_commiteecomm + Committee_comm_date + "\n"
						rev_comm_cal_date = rev_comm_cal_date + comm_cal_date + "\n"
				else:
					rev_effective=""
					rev_selection2=""
					rev_freeze=""
					rev_announce=""
					rev_clientcomm=""
					rev_qc_date=""
					rev_commiteecomm=""
					rev_comm_cal_date=""
				if(request.POST['effective_date_rebal'] != ""):
					eff_rule_rebal = int(request.POST['effective_date_rebal'])
					fre_prior_day_rebal = int(request.POST['fre_prior_days_rebal']) if(request.POST['fre_prior_days_rebal']!="") else request.POST['fre_prior_days_rebal']
					fre_rule_rebal = int(request.POST['weights_Share_Freeze_rebal'])
					ann_rule_rebal = int(request.POST['public_Announcement_rebal'])
					client_comm_rule_rebal = int(request.POST['client_Comm_rebal'])
					if(request.POST['rebalance']=="Annual" or request.POST['rebalance']=="Monthly"):
						num_month = 1
						gap=1
					elif(request.POST['rebalance']== "Quarterly"):
						num_month = 4
						gap = 3
					else:
						num_month = 2
						gap = 6
					months = dict(January=1, February=2, March=3, April=4, May=5, June=6, July=7, August=8, September=9,
						  October=10, November=11, December=12)
					month1 = months[request.POST['rebalance_month']]
					month1 = int(month1)
					cal_month = month1
					year = date.today().year
					for i in range(0,num_month):
						if(month1>12):
							month1 = month1%12;
						if(month1<cal_month):
							year = year+1
						date1 = EffectiveDate_Calculate(month1, year, eff_rule_rebal,holidays)
						eff_day = int(date1[3:5])
						eff_mon = int(date1[0:2])
						eff_year = int(date1[6:8])
						fre_prior_day_rebal = int(request.POST['fre_prior_days_rebal']) if(request.POST['fre_prior_days_rebal']!="") else ""
						if(fre_rule_rebal==6):
							freeze_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,fre_prior_day_rebal)
						else:
							freeze_date = FreezeDate_Calculate(eff_day, eff_mon,eff_year,fre_rule_rebal,holidays)
						freeze_day = int(freeze_date[3:5])
						freeze_mon = int(freeze_date[0:2])
						freeze_year = int(freeze_date[6:8])
						ann_prior_day_rebal = int(request.POST['ann_prior_days_rebal']) if(request.POST['ann_prior_days_rebal']!="") else request.POST['ann_prior_days_rebal']
						if(ann_rule_rebal==3):
							announce_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,ann_prior_day_rebal)
						else:
							if (ann_rule_rebal == 1):
								announce_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, ann_rule_rebal,holidays)
							else:
								announce_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, ann_rule_rebal,holidays)
						announce_day = int(announce_date[3:5])
						announce_mon = int(announce_date[0:2])
						announce_year = int(announce_date[6:8])
						cl_prior_day_rebal = int(request.POST['cl_prior_days_rebal']) if(request.POST['cl_prior_days_rebal']!="") else request.POST['cl_prior_days_rebal']
						if(client_comm_rule_rebal ==3):
							client_comm_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,cl_prior_day_rebal)
						else:
							if (client_comm_rule_rebal == 1):
								client_comm_date = AnnouncementDate_Calculate(eff_day, eff_mon, eff_year, client_comm_rule_rebal,holidays)
							else:
								client_comm_date = AnnouncementDate_Calculate(freeze_day, freeze_mon, freeze_year, client_comm_rule_rebal,holidays)
							Client_comm_day = int(client_comm_date[3:5])
							Client_comm_mon = int(client_comm_date[0:2])
							Client_comm_year = int(client_comm_date[6:8])
						comm_cal_prior_day_rebal = int(request.POST['cal_ag_prior_days_rebal']) if(request.POST['cal_ag_prior_days_rebal']!="") else request.POST['cal_ag_prior_days_rebal']
						comm_cal_rule_rebal = int(request.POST['comm_to_Calc_Agent_rebal'])
						if(comm_cal_rule_rebal==2):
							comm_cal_date = priorday_calculate(eff_day,eff_mon,eff_year,holidays,comm_cal_prior_day_rebal)
						else:
							comm_cal_date = client_comm_date

						qc_prior_day_rebal= int(request.POST['qc_prior_days_rebal']) if(request.POST['qc_prior_days_rebal']!="") else request.POST['qc_prior_days_rebal']
						qc_rule_rebal = int(request.POST['qc_Date_rebal'])
						if(qc_rule_rebal==3):
							qc_date_rebal = priorday_calculate(eff_day,eff_mon,eff_year,holidays,qc_prior_day_rebal)
						else:
							qc_date_rebal = freeze_date
						cal_month = month1
						month1= month1+gap
						re_effective = re_effective + date1 + "\n"
						re_freeze = re_freeze + freeze_date + "\n"
						re_announce = re_announce + announce_date + "\n"
						re_clientcomm = re_clientcomm + client_comm_date + "\n"
						re_qc = re_qc + qc_date_rebal + "\n"
						re_comm_cal_date = re_comm_cal_date + comm_cal_date + "\n"
				else:
					#print("hi")
					re_effective=""
					re_freeze=""
					re_announce=""
					re_clientcomm=""
					re_qc=""
					re_comm_cal_date=""
			else:

				color_1 = 0


				reconst_Effective_Date = request.POST['mnl_eff_date'] if(request.POST['mnl_eff_date']=="") else datetime.strptime(request.POST['mnl_eff_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Selection_Date_Cycle_2 = request.POST['mnl_sel2_date'] if(request.POST['mnl_sel2_date']=="") else datetime.strptime(request.POST['mnl_sel2_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Weights_Share_Freeze_Date = request.POST['mnl_freeze_date'] if(request.POST['mnl_freeze_date']=="") else datetime.strptime(request.POST['mnl_freeze_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Public_Announcement_Date = request.POST['mnl_pb_announce_date'] if(request.POST['mnl_pb_announce_date']=="") else datetime.strptime(request.POST['mnl_pb_announce_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Client_Comm_Date = request.POST['mnl_client_comm_date'] if(request.POST['mnl_client_comm_date']=="") else datetime.strptime(request.POST['mnl_client_comm_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Prelim_Comm_Date = request.POST['mnl_prelim_date'] if(request.POST['mnl_prelim_date']=="") else datetime.strptime(request.POST['mnl_prelim_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Completion_Date = request.POST['mnl_cmp_date'] if(request.POST['mnl_cmp_date']=="") else datetime.strptime(request.POST['mnl_cmp_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				#print("man_date", request.POST['mnl_cmp_date'])
				#reconst_Ind_Cmte_Comm_Date = request.POST['mnl_ind_Comm_date'] if(request.POST['mnl_ind_Comm_date']=="") else datetime.strptime(request.POST['mnl_ind_Comm_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_Selection_Date_Cycle_1 =request.POST['mnl_sel1_date'] if(request.POST['mnl_sel1_date']=="") else datetime.strptime(request.POST['mnl_sel1_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_qc_date = request.POST['mnl_qc_date'] if(request.POST['mnl_qc_date']=="") else datetime.strptime(request.POST['mnl_qc_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_effective = request.POST['mnl_eff_date_review'] if(request.POST['mnl_eff_date_review']=="") else datetime.strptime(request.POST['mnl_eff_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_announce = request.POST['mnl_pb_announce_date_review'] if(request.POST['mnl_pb_announce_date_review']=="") else datetime.strptime(request.POST['mnl_pb_announce_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_clientcomm = request.POST['mnl_client_comm_date_review'] if(request.POST['mnl_client_comm_date_review']=="") else datetime.strptime(request.POST['mnl_client_comm_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				#rev_commiteecomm = request.POST['mnl_ind_Comm_date_review'] if(request.POST['mnl_ind_Comm_date_review']=="") else datetime.strptime(request.POST['mnl_ind_Comm_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_qc_date = request.POST['mnl_qc_date_review'] if(request.POST['mnl_qc_date_review']=="") else datetime.strptime(request.POST['mnl_qc_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				#rev_completion = request.POST['mnl_cmp_date_review'] if(request.POST['mnl_cmp_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_cmp_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
				#rev_prelim = request.POST['mnl_prelim_date_review'] if(request.POST['mnl_prelim_date_review']=="") else datetime.datetime.strptime(request.POST['mnl_prelim_date_review'],"%Y-%m-%d").strftime("%m/%d/%Y")
				rev_freeze = request.POST['mnl_freeze_date_review'] if(request.POST['mnl_freeze_date_review']=="") else datetime.strptime(request.POST['mnl_freeze_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_selection2 = request.POST['mnl_sel2_date_review'] if(request.POST['mnl_sel2_date_review']=="") else datetime.strptime(request.POST['mnl_sel2_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_announce = request.POST['mnl_pb_announce_date_rebal'] if(request.POST['mnl_pb_announce_date_rebal']=="") else datetime.strptime(request.POST['mnl_pb_announce_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_clientcomm = request.POST['mnl_client_comm_date_rebal'] if(request.POST['mnl_client_comm_date_rebal']=="") else datetime.strptime(request.POST['mnl_client_comm_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_effective = request.POST['mnl_eff_date_rebal'] if(request.POST['mnl_eff_date_rebal']=="") else datetime.strptime(request.POST['mnl_eff_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_freeze = request.POST['mnl_freeze_date_rebal'] if(request.POST['mnl_freeze_date_rebal']=="") else datetime.strptime(request.POST['mnl_freeze_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_qc = request.POST['mnl_qc_date_rebal'] if(request.POST['mnl_qc_date_rebal']=="") else datetime.strptime(request.POST['mnl_qc_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				reconst_comm_cal_date = request.POST['mnl_comm_cal_date'] if(request.POST['mnl_comm_cal_date']=="") else datetime.strptime(request.POST['mnl_comm_cal_date'],"%Y-%m-%d").strftime("%m/%d/%y")
				re_comm_cal_date = request.POST['mnl_comm_cal_date_rebal'] if(request.POST['mnl_comm_cal_date_rebal']=="") else datetime.strptime(request.POST['mnl_comm_cal_date_rebal'],"%Y-%m-%d").strftime("%m/%d/%y")
				rev_comm_cal_date = request.POST['mnl_comm_cal_date_review'] if(request.POST['mnl_comm_cal_date_review']=="") else datetime.strptime(request.POST['mnl_comm_cal_date_review'],"%Y-%m-%d").strftime("%m/%d/%y")
			ind.Ident_ISIN=request.POST['isin_id']
			ind.Ident_Bloomberg=request.POST['bloomberg_id']
			ind.Ident_Reuters=request.POST['thomson_id']
			ind.Index_Name=request.POST['index_name']
			ind.Client_Name_id=request.POST['client_name']
			ind.Ind_Sty=request.POST['index_Style']
			ind.Ind_Ver=request.POST['ind_version']
			ind.Ind_Ver_ID=request.POST['index_version_id']
			ind.Calc=request.POST['calculation']
			ind.Calc_Agent=request.POST['cal_agent'] # if(request.POST['cal_agent_des']=="") else request.POST['cal_agent_des']
			ind.Data_Platform=request.POST['data_platform']
			ind.Data_Vendors=request.POST['data_vendors']
			ind.Contract_Type=request.POST['contract_Type']
			ind.Type_of_Ind=request.POST['type_index']
			ind.Product_Status=request.POST['prod_Status']
			ind.ETF_Launched=request.POST['etf_Launched']
			if(request.POST['etf_Launched']=="Yes"):
				ind.etf_launch_date=request.POST['etf_date']
			else:
				ind.etf_launch_date=""

			ind.Reconstitution=request.POST['reconstitution']
			ind.Rebalance=request.POST['rebalance']
			ind.Review=request.POST['review']
			ind.Theme_Review=request.POST['theme_Review']
			ind.Selection_Date_Cycle_1=reconst_Selection_Date_Cycle_1
			ind.Completion_Date=reconst_Completion_Date
			ind.Selection_Date_Cycle_2=reconst_Selection_Date_Cycle_2
			ind.Selection_Date_Cycle_2_review=rev_selection2
			#ind.Ind_Cmte_Comm_Date=reconst_Ind_Cmte_Comm_Date
			#ind.Ind_Cmte_Comm_Date_review=rev_commiteecomm
			ind.Prelim_Comm_Date=reconst_Prelim_Comm_Date
			ind.Weights_Share_Freeze_Date=reconst_Weights_Share_Freeze_Date
			ind.Weights_Share_Freeze_Date_rebal=re_freeze
			ind.Weights_Share_Freeze_Date_review=rev_freeze
			ind.Public_Announcement_Date=reconst_Public_Announcement_Date
			ind.Public_Announcement_Date_rebal= re_announce
			ind.Public_Announcement_Date_review=rev_announce
			ind.Client_Comm_Date=reconst_Client_Comm_Date
			ind.Client_Comm_Date_rebal=re_clientcomm
			ind.Client_Comm_Date_review=rev_clientcomm
			ind.Effective_Date=reconst_Effective_Date
			ind.Effective_Date_rebal=re_effective
			ind.Effective_Date_review=rev_effective
			ind.QC_Date_rebal=re_qc
			ind.QC_Date=reconst_qc_date
			ind.QC_Date_review=rev_qc_date
			ind.reconst_month=request.POST['reconst_month']
			ind.rebalance_month=request.POST['rebalance_month']
			ind.review_month=request.POST['review_month']
			ind.Comm_to_Calc_Agent=reconst_Client_Comm_Date
			ind.Comm_to_Calc_Agent_rebal=re_clientcomm
			ind.Comm_to_Calc_Agent_review=rev_clientcomm
			ind.color_code=color_1
			ind.live_date=request.POST['live_date']

			ind.etf_launch_date=request.POST['etf_date']
			ind.backtest_comp_date=request.POST['backtest_date']
			ind.save()
	for indrule in index1:
		if request.method == 'POST':
			
			#print(request.POST['qc_Date_rebal'])
			#pdb.set_trace()
			indrule.eff_rule = request.POST['effective_date']
			indrule.eff_rule_rebal = request.POST['effective_date_rebal']
			indrule.eff_rule_review = request.POST['effective_date_review']
			indrule.sel_rule1 = request.POST['selec_Date_Cyc_1']
			indrule.sel_rule2 = request.POST['selec_Date_Cyc_2']
			indrule.sel_rule2_review = request.POST['selec_Date_Cyc_2_review']
			indrule.announce_rule = request.POST['public_Announcement']
			indrule.announce_rule_rebal = request.POST['public_Announcement_rebal']
			indrule.announce_rule_review = request.POST['public_Announcement_review']
			indrule.prelim_rule = request.POST['pre_comm_date']
			indrule.clientcomm_rule = request.POST['client_Comm']
			indrule.clientcomm_rule_rebal = request.POST['client_Comm_rebal']
			indrule.clientcomm_rule_review = request.POST['client_Comm_review']
			#indrule.indcommittee_rule = request.POST['ind_Comm_Date']
			#indrule.indcommittee_rule_review = request.POST['ind_Comm_Date_review']
			indrule.freeze_rule = request.POST['weights_Share_Freeze']
			indrule.freeze_rule_rebal = request.POST['weights_Share_Freeze_rebal']
			indrule.freeze_rule_review = request.POST['weights_Share_Freeze_review']
			indrule.comp_rule = request.POST['cmp_Date']
			indrule.comm_cal_rule = request.POST['comm_to_Calc_Agent']
			indrule.comm_cal_rule_rebal = request.POST['comm_to_Calc_Agent_rebal']
			indrule.comm_cal_rule_review = request.POST['comm_to_Calc_Agent_review']
			indrule.qc_rule_rebal = request.POST['qc_Date_rebal']
			indrule.qc_rule = request.POST['qc_date']
			indrule.qc_rule_review = request.POST['ind_qc_date_review']
			indrule.save()
	for ind1 in priorday:
		if request.method == 'POST':
			ind1.sel_cycle_1_day_recon = int(request.POST['sel1_prior_days']) if(request.POST['sel1_prior_days']!="") else request.POST['sel1_prior_days']
			ind1.comp_cycle_1_day_recon = int(request.POST['cmp_prior_days']) if(request.POST['cmp_prior_days']!="") else request.POST['cmp_prior_days']
			ind1.sel_cycle_2_day_recon = int(request.POST['sel2_prior_days']) if(request.POST['sel2_prior_days']!="")else request.POST['sel2_prior_days']
			ind1.prelim_day_recon = int(request.POST['pre_prior_days']) if(request.POST['pre_prior_days']!="")else request.POST['pre_prior_days']
			ind1.freeze_day_recon = int(request.POST['fre_prior_days']) if(request.POST['fre_prior_days']!="")else request.POST['fre_prior_days']
			#ind1.indcommittee_day_recon = int(request.POST['icom_prior_days']) if(request.POST['icom_prior_days']!="") else request.POST['icom_prior_days']
			ind1.qc_day_recon = int(request.POST['qc_prior_days']) if(request.POST['qc_prior_days']!="") else request.POST['qc_prior_days']
			ind1.announce_day_recon = int(request.POST['ann_prior_days']) if(request.POST['ann_prior_days']!="")else request.POST['ann_prior_days']
			ind1.clientcomm_day_recon = int(request.POST['cl_prior_days']) if(request.POST['cl_prior_days']!="")else request.POST['cl_prior_days']
			ind1.comm_cal_agent_day_recon = int(request.POST['cal_ag_prior_days']) if(request.POST['cal_ag_prior_days']!="") else request.POST['cal_ag_prior_days']
			ind1.freeze_day_rebal = int(request.POST['fre_prior_days_rebal']) if(request.POST['fre_prior_days_rebal']!="")else request.POST['fre_prior_days_rebal']
			ind1.qc_day_rebal = int(request.POST['qc_prior_days_rebal']) if(request.POST['qc_prior_days_rebal']!="")else request.POST['qc_prior_days_rebal']
			ind1.announce_day_rebal = int(request.POST['ann_prior_days_rebal']) if(request.POST['ann_prior_days_rebal']!="")else request.POST['ann_prior_days_rebal']
			ind1.clientcomm_day_rebal = int(request.POST['cl_prior_days_rebal']) if(request.POST['cl_prior_days_rebal']!="")else request.POST['cl_prior_days_rebal']
			ind1.comm_cal_agent_day_rebal = int(request.POST['cal_ag_prior_days_rebal']) if(request.POST['cal_ag_prior_days_rebal']!="") else request.POST['cal_ag_prior_days_rebal']
			ind1.sel_cycle_2_day_review = int(request.POST['sel2_prior_days_review']) if(request.POST['sel2_prior_days_review']!="") else request.POST['sel2_prior_days_review']
			#ind1.indcommittee_day_review = int(request.POST['icom_prior_days_review']) if(request.POST['icom_prior_days_review']!="") else request.POST['icom_prior_days_review']
			ind1.qc_day_review = int(request.POST['qc_prior_days_review']) if(request.POST['qc_prior_days_review']!="") else request.POST['qc_prior_days_review']
			ind1.freeze_day_review = int(request.POST['fre_prior_days_review']) if(request.POST['fre_prior_days_review']!="") else request.POST['fre_prior_days_review']
			ind1.announce_day_review = int(request.POST['ann_prior_days_review']) if(request.POST['ann_prior_days_review']!="") else request.POST['ann_prior_days_review']
			ind1.clientcomm_day_review = int(request.POST['cl_prior_days_review']) if(request.POST['cl_prior_days_review']!="") else request.POST['cl_prior_days_review']
			ind1.comm_cal_agent_day_review = int(request.POST['cal_ag_prior_days_review']) if(request.POST['cal_ag_prior_days_review']!="") else request.POST['cal_ag_prior_days_review']
			ind1.save()


	templates = loader.get_template("DeleteIndex.html")
	return HttpResponse(templates.render())

def delete_index(request, id):
	Registerindex.objects.filter(id = id).delete()
	templates = loader.get_template("DeleteIndex.html")
	return HttpResponse(templates.render());



def report_generate(request):
	client_name = Calendarlist.objects.filter(category = 'Client_Name')
	info = Registerindex.objects.all().order_by('id').reverse()
	inst = USTradingCalendar()
	
	
	if request.method == 'POST':
		indexName = request.POST['Index_Name']
		cname1 = request.POST['Client_Code']
		month = request.POST['month1']
		month2 = request.POST['month2']
		sel_type1 = request.POST.get('radiobutton', False)



		if(indexName == "" and cname1 == "" and month =="" and month2 =="" and (sel_type1!="")):
			templateName = "no-selection.html";
			context = {
				'info' : info,
				'client_name' : client_name,}
			return render(request,templateName,context)
		elif(month!="" and month2==""):
			templateName = "DateError.html";
			
			return render(request,templateName)
		else:

			#print(year)
			#pdb.set_trace()
			calendardata = Calendarlist.objects.all()
			#client_name = Calendarlist.objects.filter(category = 'Client_Name')

			if(indexName != "" and cname1 != ""):
				lookups = Q(Index_Name__icontains = indexName)
				info = Registerindex.objects.filter(Client_Name = cname1 and lookups).order_by('id').reverse()
			elif(indexName != ""):
				lookups = Q(Index_Name__icontains = indexName)
				info = Registerindex.objects.filter(lookups).order_by('id').reverse()
			elif(cname1 != ""):
				info = Registerindex.objects.filter(Client_Name = cname1).order_by('id').reverse()
			else:
				info = Registerindex.objects.all().order_by('id').reverse()

			informa=[]
			months = dict(January=1, February=2, March=3, April = 4, May=5, June=6, July=7, August=8,September=9, October=10, November=11, December=12)
			for ind in info:
				
				today_date = date.today()
				ind.today_date = today_date.strftime("%x")
				#print(type(ind.reconst_month))
				#pdb.set_trace()
				if(ind.color_code==0):
					if ind.Effective_Date:
						month1 = months[ind.reconst_month]
					else:
						month1 = ""
					if ind.Effective_Date_review:
						month1_review = months[ind.review_month]
					else:
						month1_review = ""
					if ind.Effective_Date_rebal:
						month1_rebal = months[ind.rebalance_month]
					else:
						month1_rebal=""

					flag1=0

					if(month!=""):
						year = int(month[:4])
						mon = int(month[5:7])
						d =1
						from_month = date(year,mon,d)
					if(month2!=""):
						year1 = int(month2[:4])
						mon1 = int(month2[5:7])
						d1= num_of_days(mon1,year1)
						to_month = date(year1,mon1,d1)
					if(sel_type1=="effective"):
						if ind.Effective_Date:
							temp=0
							eff_date = datetime.strptime(ind.Effective_Date,"%m/%d/%y").strftime("%Y-%m-%d")
							eff_year = int(eff_date[:4])
							eff_mon = int(eff_date[5:7])
							eff_day = int(eff_date[8:10])
							e_date = date(eff_year,eff_mon,eff_day)
							if(month!=""):
								if(from_month > e_date):
									temp=1
							if(month2!=""):
								if(e_date>to_month):
									temp=1
							if(temp==0):
							    ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1;
							    ind.Completion_Date_rec = ind.Completion_Date;
							    ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2;
							    #ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date;
							    ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date;
							    ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date;
							    ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date;
							    ind.Client_Comm_Date_rec = ind.Client_Comm_Date;
							    ind.Effective_Date_rec = ind.Effective_Date;
							    flag1=1;
						if ind.Effective_Date_rebal:
							temp=0
							eff_date_rebal = datetime.strptime(ind.Effective_Date_rebal,"%m/%d/%y").strftime("%Y-%m-%d")
							eff_year_rebal = int(eff_date_rebal[:4])
							eff_mon_rebal = int(eff_date_rebal[5:7])
							eff_day_rebal = int(eff_date_rebal[8:10])
							e_date_rebal = date(eff_year_rebal,eff_mon_rebal,eff_day_rebal)
							if(month!=""):
								if(from_month > e_date_rebal):
									temp=1
							if(month2!=""):
								if(e_date_rebal>to_month):
									temp=1
							if(temp==0):
								ind.Ind_Cmte_Comm_Date_re = ind.Ind_Cmte_Comm_Date_rebal;
								ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_rebal;
								ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_rebal;
								ind.Client_Comm_Date_re = ind.Client_Comm_Date_rebal;
								ind.Effective_Date_re = ind.Effective_Date_rebal;
								flag1=1;
						if ind.Effective_Date_review:
							temp=0
							eff_date_review = datetime.strptime(ind.Effective_Date_review,"%m/%d/%y").strftime("%Y-%m-%d")
							eff_year_review = int(eff_date_review[:4])
							eff_mon_review = int(eff_date_review[5:7])
							eff_day_review = int(eff_date_review[8:10])
							e_date_review = date(eff_year_review,eff_mon_review,eff_day_review)
							if(month!=""):
								if(from_month > e_date_review):
									temp=1
							if(month2!=""):
								if(e_date_review>to_month):
									temp=1
							if(temp==0):
								ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_review;
								ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_review;
								ind.Prelim_Comm_Date_rev = ind.Prelim_Comm_Date_review;
								ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_review;
								ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_review;
								ind.Client_Comm_Date_rev = ind.Client_Comm_Date_review;
								ind.Effective_Date_rev = ind.Effective_Date_review;
								flag1=1;
						if(flag1==1):
						    informa.append(ind)

					if(sel_type1=="reco" and month1!=""):
						if ind.Effective_Date:
							temp=0
							eff_date = datetime.strptime(ind.Effective_Date,"%m/%d/%y").strftime("%Y-%m-%d")
							eff_year = int(eff_date[:4])
							eff_mon = int(eff_date[5:7])
							eff_day = int(eff_date[8:10])
							e_date = date(eff_year,eff_mon,eff_day)
							sel_date = datetime.strptime(ind.Selection_Date_Cycle_2,"%m/%d/%y").strftime("%Y-%m-%d")
							sel_year = int(sel_date[:4])
							sel_mon = int(sel_date[5:7])
							sel_day = int(sel_date[8:10])
							s_date = date(sel_year,sel_mon,sel_day)
							freeze_date = datetime.strptime(ind.Weights_Share_Freeze_Date,"%m/%d/%y").strftime("%Y-%m-%d")
							freeze_year = int(freeze_date[:4])
							freeze_mon = int(freeze_date[5:7])
							freeze_day = int(freeze_date[8:10])
							f_date = date(freeze_year,freeze_mon,freeze_day)
							if(ind.Selection_Date_Cycle_1!=""):
								sel1_date = datetime.strptime(ind.Selection_Date_Cycle_1,"%m/%d/%y").strftime("%Y-%m-%d")
								sel1_year = int(sel1_date[:4])
								sel1_mon = int(sel1_date[5:7])
								sel1_day = int(sel1_date[8:10])
								sel1_date = date(sel1_year,sel1_mon,sel1_day)
							if(month!="" and month2==""):
								if(from_month < e_date or from_month < sel1_date or from_month < s_date or from_month < f_date) :
									temp=1
							if(month2!="" and month!=""):
								if((e_date<to_month and from_month < e_date)  or (from_month < s_date and to_month < s_date) or (from_month < f_date and to_month < f_date)):
									temp=1
									#or (from_month < sel1_date and to_month < sel1_date)
							if(month2=="" and month==""):
								temp=1
							if(temp==1):
							    ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1;
							    ind.Completion_Date_rec = ind.Completion_Date;
							    ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2;
							    #ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date;
							    ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date;
							    ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date;
							    ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date;
							    ind.Client_Comm_Date_rec = ind.Client_Comm_Date;
							    ind.Effective_Date_rec = ind.Effective_Date;
							    flag1=1;
						if(flag1==1):
						    informa.append(ind)
					if(sel_type1=="rev" and month1_review!=""):
						if ind.Effective_Date_review:
							temp=0
							eff_date = datetime.strptime(ind.Effective_Date_review,"%m/%d/%y").strftime("%Y-%m-%d")
							eff_year = int(eff_date[:4])
							eff_mon = int(eff_date[5:7])
							eff_day = int(eff_date[8:10])
							e_date = date(eff_year,eff_mon,eff_day)
							sel_date = datetime.strptime(ind.Selection_Date_Cycle_2_review,"%m/%d/%y").strftime("%Y-%m-%d")
							sel_year = int(sel_date[:4])
							sel_mon = int(sel_date[5:7])
							sel_day = int(sel_date[8:10])
							s_date = date(sel_year,sel_mon,sel_day)
							freeze_date = datetime.strptime(ind.Weights_Share_Freeze_Date_review,"%m/%d/%y").strftime("%Y-%m-%d")
							freeze_year = int(freeze_date[:4])
							freeze_mon = int(freeze_date[5:7])
							freeze_day = int(freeze_date[8:10])
							f_date = date(freeze_year,freeze_mon,freeze_day)

							if(month!="" and month2==""):
								if(from_month < e_date or from_month < s_date or from_month < f_date) :
									temp=1
							if(month2!="" and month!=""):
								if((e_date<to_month and from_month < e_date) or (from_month < s_date and to_month < s_date) or (from_month < f_date and to_month < f_date)):
									temp=1
							if(month2=="" and month==""):
								temp=1
							if(temp==1):
								ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_review
								ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_review
								ind.Prelim_Comm_Date_rev = ind.Prelim_Comm_Date_review
								ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_review
								ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_review
								ind.Client_Comm_Date_rev = ind.Client_Comm_Date_review
								ind.Effective_Date_rev = ind.Effective_Date_review
								flag1=1
						if(flag1==1):
						    informa.append(ind)
					if(sel_type1=="reb" and month1_rebal!=""):
						if ind.Effective_Date_rebal:
							temp=0
							eff_date = datetime.strptime(ind.Effective_Date_rebal,"%m/%d/%y").strftime("%Y-%m-%d")
							eff_year = int(eff_date[:4])
							eff_mon = int(eff_date[5:7])
							eff_day = int(eff_date[8:10])
							e_date = date(eff_year,eff_mon,eff_day)

							freeze_date = datetime.strptime(ind.Weights_Share_Freeze_Date_rebal,"%m/%d/%y").strftime("%Y-%m-%d")
							freeze_year = int(freeze_date[:4])
							freeze_mon = int(freeze_date[5:7])
							freeze_day = int(freeze_date[8:10])
							f_date = date(freeze_year,freeze_mon,freeze_day)

							if(month!="" and month2==""):
								if(from_month < e_date or from_month < f_date) :
									temp=1
							if(month2!="" and month!=""):
								if((e_date<to_month and from_month < e_date) or (from_month < f_date and to_month < f_date)):
									temp=1
							if(month2=="" and month==""):
								temp=1
							if(temp==1):
								ind.Ind_Cmte_Comm_Date_re = ind.Ind_Cmte_Comm_Date_rebal;
								ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_rebal;
								ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_rebal;
								ind.Client_Comm_Date_re = ind.Client_Comm_Date_rebal;
								ind.Effective_Date_re = ind.Effective_Date_rebal;
								flag1=1
						if(flag1==1):
						    informa.append(ind)


				else:

					eff_rule=""
					eff_rule_review="";
					eff_rule_rebal="";
					month1="";
					eff_mon=0;
					month1_rebal="";
					eff_mon_rebal=0;
					month1_review="";
					eff_mon_review=0;
					num_month=0;
					num_month_rebal=0;
					num_month_review=0;
					#print(ind.id)
					#pdb.set_trace()
					rulelist = Ruleslist.objects.get(index_id=ind.id)
					try:
						priordays = Priordays.objects.get(index_id=ind.id)
					except Priordays.DoesNotExist:
						priordays = None
					if(rulelist.eff_rule!=""):
						eff_rule = rulelist.eff_rule
						sel_rule = int(rulelist.sel_rule1) if(rulelist.sel_rule1!="") else rulelist.sel_rule1
						sel_rule2 = rulelist.sel_rule2
						fre_rule = rulelist.freeze_rule
						ann_rule = rulelist.announce_rule
						comp_rule = rulelist.comp_rule
						qc_rule = rulelist.qc_rule
						comm_cal_rule = rulelist.comm_cal_rule
						#icom_rule = rulelist.indcommittee_rule
						client_comm_rule = rulelist.clientcomm_rule
						Prelim_comm_rule = rulelist.prelim_rule
						month1 = months[ind.reconst_month]
						time = ind.Reconstitution
						cal_month = month1
						if(time=="Annual"):
							num_month=1
							gap=12
						elif(time=="Monthly"):
							num_month =12
							gap=1
						elif(time=="Quarterly"):
							num_month=4
							gap=3
						else:
							num_month=2
							gap=6
						month1 = int(month1)



					if(rulelist.eff_rule_rebal!=""):
						eff_rule_rebal = rulelist.eff_rule_rebal
						sel_rule_rebal = rulelist.sel_rule1_rebal
						sel_rule2_rebal = rulelist.sel_rule2_rebal
						fre_rule_rebal = rulelist.freeze_rule_rebal
						ann_rule_rebal = rulelist.announce_rule_rebal
						qc_rule_rebal = rulelist.qc_rule_rebal
						comm_cal_rule_rebal = rulelist.comm_cal_rule_rebal
						client_comm_rule_rebal = rulelist.clientcomm_rule_rebal
						Prelim_comm_rule_rebal = rulelist.prelim_rule_rebal
						month1_rebal = months[ind.rebalance_month]

						month1_rebal = int(month1_rebal)

						time_rebal = ind.Rebalance
						cal_month_rebal = month1_rebal
						if(time_rebal=="Annual"):
							num_month_rebal=1
							gap_rebal=12
						elif(time_rebal=="Monthly"):
							num_month_rebal =12
							gap_rebal=1
						elif(time_rebal=="Quarterly"):
							num_month_rebal=4
							gap_rebal=3
						else:
							num_month_rebal=2
							gap_rebal=6

					if(rulelist.eff_rule_review!=""):
						eff_rule_review = rulelist.eff_rule_review

						sel_rule2_review = rulelist.sel_rule2_review
						fre_rule_review = rulelist.freeze_rule_review
						ann_rule_review = rulelist.announce_rule_review
						comm_cal_rule_review = rulelist.comm_cal_rule_review
						client_comm_rule_review = rulelist.clientcomm_rule_review
						Prelim_comm_rule_review = rulelist.prelim_rule_review
						#icom_rule_review = rulelist.indcommittee_rule_review
						qc_rule_review = rulelist.qc_rule_review
						month1_review = months[ind.review_month]
						month1_review = int(month1_review)

						time_review = ind.Review
						cal_month_review = month1_review
						if(time_review=="Annual"):
							num_month_review=1
							gap_review=12
						elif(time_review=="Monthly"):
							num_month_review =12
							gap_review=1
						elif(time_review=="Quarterly"):
							num_month_review=4
							gap_review=3
						else:
							num_month_review=2
							gap_review=6																										

					
					if(month!="" and month2!=""):
						print(ind.Index_Name)
						year = int(month[:4])
						mon = int(month[5:7])
						d = 1
						holidays = inst.holidays(dt.datetime(year-1, 9, 30), dt.datetime(year, 12, 31))
						from_month = date(year,mon,d)

						year1 = int(month2[:4])
						mon1 = int(month2[5:7])
						d1 = num_of_days(mon1,year1)
						to_month = date(year1,mon1,d1)
						flag=0;
						ind.Selection_Date_Cycle_1_rec="";
						ind.Completion_Date_rec ="";
						ind.Selection_Date_Cycle_2_rec="";
						ind.Ind_Cmte_Comm_Date_rec=""
						ind.Prelim_Comm_Date_rec=""
						ind.Weights_Share_Freeze_Date_rec=""
						ind.Public_Announcement_Date_rec=""
						ind.Client_Comm_Date_rec=""
						ind.Effective_Date_rec=""
						ind.Ind_Cmte_Comm_Date_re=""
						ind.Prelim_Comm_Date_re=""
						ind.Weights_Share_Freeze_Date_re=""
						ind.Public_Announcement_Date_re=""
						ind.Client_Comm_Date_re=""
						ind.Effective_Date_re=""
						ind.Selection_Date_Cycle_2_rev="";
						ind.Ind_Cmte_Comm_Date_rev=""
						ind.Prelim_Comm_Date_rev=""
						ind.Weights_Share_Freeze_Date_rev=""
						ind.Public_Announcement_Date_rev=""
						ind.Client_Comm_Date_rev=""
						ind.Effective_Date_rev=""
						ind.QC_Date_rec = ""
						ind.QC_Date_re = ""
						ind.QC_Date_rev = ""
						ind.Comm_cal_Date_rec=""
						ind.Comm_cal_Date_re=""
						ind.Comm_cal_Date_rev=""
						if(sel_type1 =="effective"):


							if eff_rule:
								date1 = EffectiveDate_Calculate(month1, year, int(eff_rule),holidays)
								eff_day = int(date1[3:5])
								eff_month = int(date1[0:2])




									#ind.Effective_Date = date1
							if eff_rule_rebal:
								date1_rebal = EffectiveDate_Calculate(month1_rebal, year, int(eff_rule_rebal),holidays)
								eff_day_rebal = int(date1_rebal[3:5])
								eff_mon_rebal = int(date1_rebal[0:2])

							if eff_rule_review:
								date1_review = EffectiveDate_Calculate(month1_review, year, int(eff_rule_review),holidays)
								eff_day_review = int(date1_review[3:5])
								eff_mon_review = int(date1_review[0:2])
								dt_obj = datetime.strptime(date1_review,'%m/%d/%y')
								e_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								e_y_review = int(e_date_review[:4])
								e_date_review = date(e_y_review,eff_mon_review,eff_day_review)
					
							year = date.today().year
							from_year = year
							
							while(from_year<=year1):
								
								holidays = inst.holidays(dt.datetime(from_year-1, 9, 30), dt.datetime(from_year, 12, 31))
								for i in range(0,num_month):
									if(month1!=""):
										if(month1>12):
											month1 = month1%12
											
										if(month1<cal_month):
											year = year+1
											if(month1==0):
												month1=12

									if eff_rule:
										
										date1 = EffectiveDate_Calculate(month1,year,int(eff_rule),holidays)
										eff_day = int(date1[3:5])
										eff_month = int(date1[0:2])
										eff_year = int(date1[6:8])
										dt_obj = datetime.strptime(date1,'%m/%d/%y')
										e_date = datetime.strftime(dt_obj, "%Y-%m-%d")
										e_y = int(e_date[:4])
										e_date = date(e_y,eff_month,eff_day)
									if sel_rule2:
										if(int(sel_rule2)==7):
											sel_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.sel_cycle_2_day_recon))
										else:
											sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,int(sel_rule2),holidays)
										sel_day = int(sel_date[3:5])
										sel_month = int(sel_date[0:2])
										sel_year = int(sel_date[6:8])

									if fre_rule:
										if(int(fre_rule)==6):
											freeze_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.freeze_day_recon))
										else:
											freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,int(fre_rule),holidays)
										freeze_day = int(freeze_date[3:5])
										freeze_month = int(freeze_date[0:2])
										freeze_year = int(freeze_date[6:8])
									if(int(ann_rule)==3):
											announce_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.announce_day_recon))
									else:
										 if(int(ann_rule)==1):
											  announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(ann_rule),holidays)
										 else:
											  announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(ann_rule),holidays)
									announce_day = int(announce_date[3:5])
									announce_month = int(announce_date[0:2])
									announce_year = int(announce_date[6:8])
									if(int(client_comm_rule)==3):
											client_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.clientcomm_day_recon))
									else:
										if(int(client_comm_rule)==1):
											client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(client_comm_rule),holidays)
										else:
											client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(client_comm_rule),holidays)
									Client_comm_day = int(client_comm_date[3:5])
									Client_comm_month = int(client_comm_date[0:2])
									Client_comm_year = int(client_comm_date[6:8])
									if(Prelim_comm_rule!=""):
										if(int(Prelim_comm_rule)==2):
											#print(ind.Index_Name)
											Prelim_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.prelim_day_recon))
										else:
											Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year,holidays)
										Prelim_comm_day = int(Prelim_comm_date[3:5])
										Prelim_comm_month = int(Prelim_comm_date[0:2])
										Prelim_comm_year = int(Prelim_comm_date[6:8])
									else:
										Prelim_comm_date = ""

									if(int(comp_rule)==2):
										Completion_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.comp_cycle_1_day_recon))
									else:
										Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year,holidays)
									Completion_month = int(Completion_date[0:2])
									"""
									if(icom_rule=="3"):
										Committee_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.indcommittee_day_recon))
									elif(icom_rule=="2"):
										Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year,holidays)
									else:
										Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year,holidays)
									Committee_comm_month = int(Committee_comm_date[0:2])
									Committee_comm_day = int(Committee_comm_date[3:5])
									Committee_comm_year = int(Committee_comm_date[6:8])
									"""
									if(sel_rule==2):
										selection_date1 = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.sel_cycle_1_day_recon))
									elif(sel_rule==1):
										selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year,holidays)
									else:
										selection_date1 =""


									if(int(qc_rule)==3):
										Quality_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.qc_day_recon))
									else:
										Quality_date = QualityDate_Calculate(freeze_day,freeze_month,freeze_year,holidays)

									if(comm_cal_rule=="2"):
										comm_cal_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.comm_cal_agent_day_recon))
									else:
										comm_cal_date = client_comm_date
									if(month1!=""):
										cal_month = month1										
										month1= month1+gap
										
									if month!="":
										if from_month <= e_date and e_date <= to_month and month1!="":
											flag=1
											ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
											ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
											ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
											#ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
											ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
											ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
											ind.QC_Date_rec = ind.QC_Date_rec + Quality_date + "\n"
											ind.Comm_cal_Date_rec = ind.Comm_cal_Date_rec + comm_cal_date + "\n"
											ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
											ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
											ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
											data = {}
									else:
										if e_date <= to_month and month1!="":
											flag=1
											ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
											ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
											ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
											#ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
											ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
											ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
											ind.Comm_cal_Date_rec= ind.Comm_cal_Date_rec= + comm_cal_date + "\n"
											ind.QC_Date_rec = ind.QC_Date_rec + Quality_date + "\n"
											ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
											ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
											ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
											data = {}
								from_year = from_year+1
								

							year = date.today().year
							from_year = year
							while(from_year<= year1):
								holidays = inst.holidays(dt.datetime(from_year-1, 9, 30), dt.datetime(from_year, 12, 31))
								for i in range(0,num_month_review):
									if(month1_review!=""):
										if(month1_review>12):
											month1_review = month1_review%12;
										if(month1_review<cal_month_review):
											year = year+1
											if(month1_review==0):
												month1_review=12
									if eff_rule_review:
										date1_review = EffectiveDate_Calculate(month1_review,year,int(eff_rule_review),holidays)
										eff_day_review = int(date1_review[3:5])
										eff_month_review = int(date1_review[0:2])
										eff_year_review = int(date1_review[6:8])
										dt_obj = datetime.strptime(date1_review,'%m/%d/%y')
										e_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
										e_y_review = int(e_date_review[:4])
										e_date_review = date(e_y_review,eff_month_review,eff_day_review)

									if (int(sel_rule2_review)==7):
										sel_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.sel_cycle_2_day_review))
									else:
										sel_date_review = SelectionDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(sel_rule2_review),holidays)

									if (int(fre_rule_review)==6):
										freeze_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.freeze_day_review))
									else:
										freeze_date_review = FreezeDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(fre_rule_review),holidays)
									freeze_day_review = int(freeze_date_review[3:5])
									freeze_month_review = int(freeze_date_review[0:2])
									freeze_year_review = int(freeze_date_review[6:8])

									if (int(ann_rule_review)==3):
										announce_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.announce_day_review))
									else:
										if(int(ann_rule_review)==1):
											announce_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(ann_rule_review),holidays)
										else:
											announce_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(ann_rule_review),holidays)
									if (int(client_comm_rule_review)==3):
										client_comm_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.clientcomm_day_review))
									else:
										if(int(client_comm_rule_review)==1):
											client_comm_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(client_comm_rule_review),holidays)
										else:
											client_comm_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(client_comm_rule_review),holidays)
									Client_comm_day_review = int(client_comm_date_review[3:5])
									Client_comm_month_review = int(client_comm_date_review[0:2])
									Client_comm_year_review = int(client_comm_date_review[6:8])

									"""
									if(Prelim_comm_rule_review==""):
										if (icom_rule_review=="3"):
											Committee_comm_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.indcommittee_day_review))
										else:
											Committee_comm_date_review = Committe_comm_date_Calculate(Client_comm_day_review,Client_comm_month_review,Client_comm_year_review,holidays)
									else:
										Committee_comm_date_review = Committe_comm_date_Calculate(Prelim_comm_day_review,Prelim_comm_month_review,Prelim_comm_year_review,holidays)


									Committee_comm_month_review = int(Committee_comm_date_review[0:2])
									Committee_comm_day_review = int(Committee_comm_date_review[3:5])
									Committee_comm_year_review = int(Committee_comm_date_review[6:8])
									"""
									if(int(qc_rule_review)==3):
										Quality_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.qc_day_review))
									else:
										Quality_date_review = QualityDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,holidays)
									if(comm_cal_rule_review=="2"):
										comm_cal_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.comm_cal_agent_day_review))
									else:
										comm_cal_date_review = client_comm_date_review

									if(month1_review!=""):
										cal_month_review = month1_review
										month1_review= month1_review+gap_review
									if month!="":
										if from_month <= e_date_review and e_date_review <= to_month and month1_review!="":
											flag=1

											#ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
											ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
											ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
											ind.QC_Date_rev= ind.QC_Date_rev + Quality_date_review + "\n"
											ind.Comm_cal_Date_rev= ind.Comm_cal_Date_rev + comm_cal_date_review + "\n"
											ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
											ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
											ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
											data = {}
									else:
										if e_date_review <= to_month and month1_review!="":
											flag=1

											#ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
											ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
											ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
											ind.QC_Date_rev= ind.QC_Date_rev + Quality_date_review + "\n"
											ind.Comm_cal_Date_rev= ind.Comm_cal_Date_rev + comm_cal_date_review + "\n"
											ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
											ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
											ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
											data = {}
								from_year = from_year+1
								
							year = date.today().year
							from_year = year
							while(from_year<= year1):
								holidays = inst.holidays(dt.datetime(from_year-1, 9, 30), dt.datetime(from_year, 12, 31))
								for i in range(0,num_month_rebal):
									if(month1_rebal!=""):
										if(month1_rebal>12):
											month1_rebal = month1_rebal%12
										if(month1_rebal<cal_month_rebal):
											year = year+1
											if(month1_rebal==0):
												month1_rebal=12
											
									if eff_rule_rebal:
										date1_rebal = EffectiveDate_Calculate(month1_rebal,year,int(eff_rule_rebal),holidays)
										eff_day_rebal = int(date1_rebal[3:5])
										eff_month_rebal = int(date1_rebal[0:2])
										eff_year_rebal = int(date1_rebal[6:8])
										dt_obj = datetime.strptime(date1_rebal,'%m/%d/%y')
										e_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
										e_y_rebal = int(e_date_rebal[:4])
										e_date_rebal = date(e_y_rebal,eff_month_rebal,eff_day_rebal)

									if (int(fre_rule_rebal)==6):
										freeze_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.freeze_day_rebal))
									else:
										freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(fre_rule_rebal),holidays)
										freeze_day_rebal = int(freeze_date_rebal[3:5])
										freeze_month_rebal = int(freeze_date_rebal[0:2])
										freeze_year_rebal = int(freeze_date_rebal[6:8])

									if (int(ann_rule_rebal)==3):
										announce_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.announce_day_rebal))
									else:
										if(int(ann_rule_rebal)==1):
											announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(ann_rule_rebal),holidays)
										else:
											announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(ann_rule_rebal),holidays)

									if (int(client_comm_rule_rebal)==3):
										client_comm_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.clientcomm_day_rebal))
									else:
										if(int(client_comm_rule_rebal)==1):
											client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(client_comm_rule_rebal),holidays)
										else:
											client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(client_comm_rule_rebal),holidays)
									if(int(qc_rule_rebal)==3):
										Quality_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.qc_day_rebal))
									else:
										Quality_date_rebal = freeze_date_rebal
									if(comm_cal_rule_rebal=="2"):
										comm_cal_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.comm_cal_agent_day_rebal))
									else:
										comm_cal_date_rebal = client_comm_date_rebal


									if(month1_rebal!=""):
										cal_month_rebal = month1_rebal
										month1_rebal= month1_rebal+gap_rebal
									if month!="":
										if from_month <= e_date_rebal and e_date_rebal <= to_month and month1_rebal!="":
											flag=1
											ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
											ind.QC_Date_re = ind.QC_Date_re  + Quality_date_rebal + "\n"
											ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
											ind.Comm_cal_Date_re= ind.Comm_cal_Date_re + comm_cal_date_rebal + "\n"
											ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
											ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
											data = {}
									else:
										if e_date_rebal <= to_month and month1_rebal!="":
											flag=1
											ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
											ind.QC_Date_re = ind.QC_Date_re  + Quality_date_rebal + "\n"
											ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
											ind.Comm_cal_Date_re= ind.Comm_cal_Date_re + comm_cal_date_rebal + "\n"
											ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
											ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
											data = {}
								from_year = from_year+1


						elif(sel_type1=="reco" and month1!=""):
							year = date.today().year
							from_year = year
							while(from_year <= year1):
								holidays = inst.holidays(dt.datetime(from_year-1, 9, 30), dt.datetime(from_year, 12, 31))
								for i in range(0,num_month):
									if(month1>12):
										month1 = month1%12
										
									if(month1<cal_month):
										year = year+1
										if(month1==0):
											month1=12
									if eff_rule:
										#print(month1)
										#print(year)
										#print(eff_rule)
										#pdb.set_trace()
										date1 = EffectiveDate_Calculate(month1,year,int(eff_rule),holidays)
										eff_day = int(date1[3:5])
										eff_month = int(date1[0:2])
										eff_year = int(date1[6:8])
										dt_obj = datetime.strptime(date1,'%m/%d/%y')
										e_date = datetime.strftime(dt_obj, "%Y-%m-%d")
										e_y = int(e_date[:4])
										e_date = date(e_y,eff_month,eff_day)
									if sel_rule2:
										if(int(sel_rule2)==7):
											sel_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.sel_cycle_2_day_recon))
										else:
											sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,int(sel_rule2),holidays)
										sel_day = int(sel_date[3:5])
										sel_month = int(sel_date[0:2])
										sel_year = int(sel_date[6:8])
										dt_obj = datetime.strptime(sel_date,'%m/%d/%y')
										s_date = datetime.strftime(dt_obj, "%Y-%m-%d")
										s_y = int(s_date[:4])
										s_date = date(s_y,sel_month,sel_day)

									if fre_rule:
										if(int(fre_rule)==6):
											freeze_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.freeze_day_recon))
										else:
											freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,int(fre_rule),holidays)
										freeze_day = int(freeze_date[3:5])
										freeze_month = int(freeze_date[0:2])
										freeze_year = int(freeze_date[6:8])
										dt_obj = datetime.strptime(freeze_date,'%m/%d/%y')
										f_date = datetime.strftime(dt_obj, "%Y-%m-%d")
										f_y = int(f_date[:4])
										f_date = date(f_y,freeze_month,freeze_day)
									if(int(ann_rule)==3):
											announce_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.announce_day_recon))
									else:
										 if(int(ann_rule)==1):
											  announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(ann_rule),holidays)
										 else:
											  announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(ann_rule),holidays)
									announce_day = int(announce_date[3:5])
									announce_month = int(announce_date[0:2])
									announce_year = int(announce_date[6:8])
									dt_obj = datetime.strptime(announce_date,'%m/%d/%y')
									a_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									a_y = int(a_date[:4])
									a_date = date(a_y,announce_month,announce_day)

									if(int(client_comm_rule)==3):
											client_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.clientcomm_day_recon))
									else:
										if(int(client_comm_rule)==1):
											client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(client_comm_rule),holidays)
										else:
											client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(client_comm_rule),holidays)
									Client_comm_day = int(client_comm_date[3:5])
									Client_comm_month = int(client_comm_date[0:2])
									Client_comm_year = int(client_comm_date[6:8])
									dt_obj = datetime.strptime(client_comm_date,'%m/%d/%y')
									cl_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									cl_y = int(cl_date[:4])
									cl_date = date(cl_y,Client_comm_month,Client_comm_day)
									if(Prelim_comm_rule!=""):
										if(int(Prelim_comm_rule)==2):
											Prelim_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.prelim_day_recon))
										else:
											Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year,holidays)
										Prelim_comm_day = int(Prelim_comm_date[3:5])
										Prelim_comm_month = int(Prelim_comm_date[0:2])
										Prelim_comm_year = int(Prelim_comm_date[6:8])
										dt_obj = datetime.strptime(Prelim_comm_date,'%m/%d/%y')
										p_date = datetime.strftime(dt_obj, "%Y-%m-%d")
										p_y = int(p_date[:4])
										p_date = date(p_y,Prelim_comm_month,Prelim_comm_day)
									else:
										Prelim_comm_date=""


									if(int(comp_rule)==2):
										Completion_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.comp_cycle_1_day_recon))
									else:
										Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year,holidays)
									Completion_month = int(Completion_date[0:2])
									Completion_day = int(Completion_date[3:5])
									dt_obj = datetime.strptime(Completion_date,'%m/%d/%y')
									c_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									c_y = int(c_date[:4])
									c_date = date(c_y,Completion_month,Completion_day)
									"""
									if(icom_rule=="3"):
										Committee_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.indcommittee_day_recon))
									elif(icom_rule=="2"):
										Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year,holidays)
									else:
										Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year,holidays)
									Committee_comm_month = int(Committee_comm_date[0:2])
									Committee_comm_day = int(Committee_comm_date[3:5])
									Committee_comm_year = int(Committee_comm_date[6:8])
									"""
									dt_obj = datetime.strptime(freeze_date,'%m/%d/%y')
									com_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									com_y = int(com_date[:4])
									com_date = date(com_y,freeze_month,freeze_day)
									if(sel_rule==2):
										selection_date1 = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.sel_cycle_1_day_recon))
										sel_1_month = int(selection_date1[0:2])
										sel_1_day = int(selection_date1[3:5])
										dt_obj = datetime.strptime(selection_date1,'%m/%d/%y')
										s1_date = datetime.strftime(dt_obj, "%Y-%m-%d")
										s1_y = int(s1_date[:4])
										sel_1_date = date(s1_y, sel_1_month, sel_1_day)
									elif(sel_rule==1):
										selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year,holidays)
										sel_1_month = int(selection_date1[0:2])
										sel_1_day = int(selection_date1[3:5])
										dt_obj = datetime.strptime(selection_date1,'%m/%d/%y')
										s1_date = datetime.strftime(dt_obj, "%Y-%m-%d")
										s1_y = int(s1_date[:4])
										sel_1_date = date(s1_y, sel_1_month, sel_1_day)
									else:
										selection_date1 =""


									if(int(qc_rule)==3):
										Quality_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.qc_day_recon))
									else:
										Quality_date = QualityDate_Calculate(freeze_day,freeze_month,freeze_year,holidays)
									qc_month = int(Quality_date[0:2])
									qc_day = int(Quality_date[3:5])
									dt_obj = datetime.strptime(Quality_date,'%m/%d/%y')
									qc_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									qc_y = int(qc_date[:4])
									qc_date = date(qc_y, qc_month, qc_day)
									if(int(comm_cal_rule)==2):
										comm_cal_ag_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.comm_cal_agent_day_recon))
									else:
										comm_cal_ag_date = client_comm_date
									comm_cal_month = int(comm_cal_ag_date[0:2])
									comm_cal_day = int(comm_cal_ag_date[3:5])
									dt_obj = datetime.strptime(comm_cal_ag_date,'%m/%d/%y')
									comm_cal_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									comm_cal_y = int(comm_cal_date[:4])
									comm_cal_date = date(comm_cal_y, comm_cal_month, comm_cal_day)

									cal_month = month1
									month1= month1+gap
									if month!="":
										if((from_month <= e_date and e_date <= to_month)  or (from_month<= s_date and s_date <= to_month )or (from_month<= f_date and f_date <= to_month)
											or (from_month <= comm_cal_date and comm_cal_date <= to_month)
											or (from_month<=a_date and a_date <= to_month) or (from_month<=cl_date and cl_date <= to_month) or (from_month<=qc_date and qc_date <= to_month)
										 or (from_month<=c_date and c_date <= to_month) or (from_month<=com_date and com_date <= to_month) ):
											flag=1
											#print(to_month)
											#or (from_month <= sel_1_date and sel_1_date <= to_month )
											#or (from_month <= p_date and p_date <= to_month)
											ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
											ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
											ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
											#ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
											ind.Comm_cal_Date_rec= ind.Comm_cal_Date_rec + comm_cal_ag_date + "\n"
											ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
											ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
											ind.QC_Date_rec = ind.QC_Date_rec + Quality_date + "\n"
											ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
											ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
											ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
											data = {}
									else:
										if(e_date <= to_month or s_date <= to_month or f_date <= to_month
											or a_date <= to_month or cl_date <= to_month or qc_date <= to_month or comm_cal_date <= to_month or c_date <= to_month or sel_1_date <= to_month or com_date <= to_month ):
											flag=1
											#or p_date <= to_month
											ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
											ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
											ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
											#ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
											ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
											ind.Comm_cal_Date_rec= ind.Comm_cal_Date_rec + "\n"
											ind.QC_Date_rec = ind.QC_Date_rec + Quality_date + "\n"
											ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
											ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
											ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
											ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
											data = {}								
								from_year = from_year +1
								
						elif(sel_type1=="reb" and month1_rebal!=""):
							year = date.today().year
							from_year = year
							while(from_year<= year1):
								holidays = inst.holidays(dt.datetime(from_year-1, 9, 30), dt.datetime(from_year, 12, 31))
								for i in range(0,num_month_rebal):
									if(month1_rebal>12):
										month1_rebal = month1_rebal%12
										
									if(month1_rebal<cal_month_rebal):
										year = year+1
										if(month1_rebal==0):
											month1_rebal = 12

									if eff_rule_rebal:
										date1_rebal = EffectiveDate_Calculate(month1_rebal,year,int(eff_rule_rebal),holidays)
										eff_day_rebal = int(date1_rebal[3:5])
										eff_month_rebal = int(date1_rebal[0:2])
										eff_year_rebal = int(date1_rebal[6:8])
										dt_obj = datetime.strptime(date1_rebal,'%m/%d/%y')
										e_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
										e_y_rebal = int(e_date_rebal[:4])
										e_date_rebal = date(e_y_rebal,eff_month_rebal,eff_day_rebal)
										if (int(fre_rule_rebal)==6):
											freeze_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.freeze_day_rebal))
										else:
											freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(fre_rule_rebal),holidays)

										#print(freeze_date_rebal)
										#pdb.set_trace()
										freeze_day_rebal = int(freeze_date_rebal[3:5])
										freeze_month_rebal = int(freeze_date_rebal[0:2])
										freeze_year_rebal = int(freeze_date_rebal[6:8])
										dt_obj = datetime.strptime(freeze_date_rebal,'%m/%d/%y')
										f_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
										f_y_rebal = int(f_date_rebal[:4])
										f_date_rebal = date(f_y_rebal,freeze_month_rebal,freeze_day_rebal)


									if (int(ann_rule_rebal)==3):
										announce_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.announce_day_rebal))
									elif(int(ann_rule_rebal)==1):
										announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(ann_rule_rebal),holidays)
									else:
										announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(ann_rule_rebal),holidays)
									announce_day_rebal = int(announce_date_rebal[3:5])
									announce_month_rebal = int(announce_date_rebal[0:2])
									announce_year_rebal = int(announce_date_rebal[6:8])
									dt_obj = datetime.strptime(announce_date_rebal,'%m/%d/%y')
									a_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
									a_y_rebal = int(a_date_rebal[:4])
									a_date_rebal = date(a_y_rebal,announce_month_rebal,announce_day_rebal)
									if (int(client_comm_rule_rebal)==3):
										client_comm_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.clientcomm_day_rebal))
									elif(int(client_comm_rule_rebal)==1):
										client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(client_comm_rule_rebal),holidays)
									else:
										client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(client_comm_rule_rebal),holidays)
									Client_comm_day_rebal = int(client_comm_date_rebal[3:5])
									Client_comm_month_rebal = int(client_comm_date_rebal[0:2])
									Client_comm_year_rebal = int(client_comm_date_rebal[6:8])
									dt_obj = datetime.strptime(client_comm_date_rebal,'%m/%d/%y')
									cl_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
									cl_y_rebal = int(cl_date_rebal[:4])
									cl_date_rebal = date(cl_y_rebal,Client_comm_month_rebal,Client_comm_day_rebal)

									if(int(qc_rule_rebal)==3):
										Quality_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.qc_day_rebal))
									else:
										Quality_date_rebal = freeze_date_rebal
									qc_month_rebal = int(Quality_date_rebal[0:2])
									qc_day_rebal = int(Quality_date_rebal[3:5])
									dt_obj = datetime.strptime(Quality_date_rebal,'%m/%d/%y')
									qc_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
									qc_y_rebal = int(qc_date_rebal[:4])
									qc_date_rebal = date(qc_y_rebal, qc_month_rebal, qc_day_rebal)

									if(int(comm_cal_rule_rebal)==2):
										comm_cal_ag_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.comm_cal_agent_day_rebal))
									else:
										comm_cal_ag_date_rebal = client_comm_date_rebal
									comm_cal_month_rebal = int(comm_cal_ag_date_rebal[0:2])
									comm_cal_day_rebal = int(comm_cal_ag_date_rebal[3:5])
									dt_obj = datetime.strptime(comm_cal_ag_date_rebal,'%m/%d/%y')
									comm_cal_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
									comm_cal_y_rebal = int(comm_cal_date_rebal[:4])
									comm_cal_date_rebal = date(comm_cal_y_rebal, comm_cal_month_rebal, comm_cal_day_rebal)


									cal_month_rebal = month1_rebal
									month1_rebal= month1_rebal+gap_rebal
									if month!="":
										if((from_month <= e_date_rebal and e_date_rebal <= to_month) or
											(from_month<= f_date_rebal and f_date_rebal <= to_month) or (from_month<= qc_date_rebal and qc_date_rebal <= to_month)
											or (from_month<=a_date_rebal and a_date_rebal <= to_month) or (from_month<=cl_date_rebal and cl_date_rebal <= to_month)
										  or (from_month<= comm_cal_date_rebal and comm_cal_date_rebal <= to_month)):
											ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
											ind.QC_Date_re = ind.QC_Date_re  + Quality_date_rebal + "\n"
											ind.Comm_cal_Date_re= ind.Comm_cal_Date_re + comm_cal_ag_date_rebal + "\n"
											ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
											ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
											ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
											data = {}
											flag=1
									else:
										if(e_date_rebal <= to_month or comm_cal_date_rebal <= to_month or f_date_rebal <= to_month or a_date_rebal <= to_month or cl_date_rebal <= to_month
										 or qc_date_rebal <= to_month):
											ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
											ind.QC_Date_re = ind.QC_Date_re  + Quality_date_rebal + "\n"
											ind.Comm_cal_Date_re= ind.Comm_cal_Date_re + comm_cal_ag_date_rebal + "\n"
											ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
											ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
											ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
											data = {}
											flag=1
								from_year = from_year +1


						elif(sel_type1=="rev" and month1_review!=""):
							year = date.today().year
							from_year = year
							while(from_year<= year1):
								holidays = inst.holidays(dt.datetime(from_year-1, 9, 30), dt.datetime(from_year, 12, 31))
								for i in range(0,num_month_review):
									if(month1_review>12):
										month1_review = month1_review%12
										
									if(month1_review<cal_month_review):
										year = year+1
										if(month1_review==0):
											month1_review = 12
									if eff_rule_review:
										date1_review = EffectiveDate_Calculate(month1_review,year,int(eff_rule_review),holidays)
										eff_day_review = int(date1_review[3:5])
										eff_month_review = int(date1_review[0:2])
										eff_year_review = int(date1_review[6:8])
										dt_obj = datetime.strptime(date1_review,'%m/%d/%y')
										e_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
										e_y_review = int(e_date_review[:4])
										e_date_review = date(e_y_review,eff_month_review,eff_day_review)

									if(int(sel_rule2_review)==7):
										sel_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.sel_cycle_2_day_review))
									else:
										sel_date_review = SelectionDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(sel_rule2_review),holidays)
									sel_day_review = int(sel_date_review[3:5])
									sel_month_review = int(sel_date_review[0:2])
									sel_year_review = int(sel_date_review[6:8])
									dt_obj = datetime.strptime(sel_date_review,'%m/%d/%y')
									s_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									s_y_review = int(s_date_review[:4])
									s_date_review = date(s_y_review,sel_month_review,sel_day_review)
									
									if (int(fre_rule_review)==6):
										freeze_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.freeze_day_review))
									else:
										freeze_date_review = FreezeDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(fre_rule_review),holidays)
									
									freeze_day_review = int(freeze_date_review[3:5])
									freeze_month_review = int(freeze_date_review[0:2])
									freeze_year_review = int(freeze_date_review[6:8])
									dt_obj = datetime.strptime(freeze_date_review,'%m/%d/%y')
									f_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									f_y_review = int(f_date_review[:4])
									f_date_review = date(f_y_review,freeze_month_review,freeze_day_review)

									if (int(ann_rule_review)==3):
										announce_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.announce_day_review))
									elif(int(ann_rule_review)==1):
										announce_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(ann_rule_review),holidays)
									else:
										announce_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(ann_rule_review),holidays)
									
									announce_day_review = int(announce_date_review[3:5])
									announce_month_review = int(announce_date_review[0:2])
									announce_year_review = int(announce_date_review[6:8])
									dt_obj = datetime.strptime(announce_date_review,'%m/%d/%y')
									a_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									a_y_review = int(a_date_review[:4])
									a_date_review = date(a_y_review,announce_month_review,announce_day_review)
									if (int(client_comm_rule_review)==3):
										client_comm_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.clientcomm_day_review))
									elif(int(client_comm_rule_review)==1):
										client_comm_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(client_comm_rule_review),holidays)
									else:
										client_comm_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(client_comm_rule_review),holidays)
									Client_comm_day_review = int(client_comm_date_review[3:5])
									Client_comm_month_review = int(client_comm_date_review[0:2])
									Client_comm_year_review = int(client_comm_date_review[6:8])
									dt_obj = datetime.strptime(client_comm_date_review,'%m/%d/%y')
									cl_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									cl_y_review = int(cl_date_review[:4])
									cl_date_review = date(cl_y_review,Client_comm_month_review,Client_comm_day_review)
									"""
									if(Prelim_comm_rule_review==""):
										if (icom_rule_review=="3"):
											Committee_comm_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.indcommittee_day_review))
										else:
											Committee_comm_date_review = Committe_comm_date_Calculate(Client_comm_day_review,Client_comm_month_review,Client_comm_year_review,holidays)
									else:
											Committee_comm_date_review = Committe_comm_date_Calculate(Prelim_comm_day_review,Prelim_comm_month_review,Prelim_comm_year_review,holidays)

									Committee_comm_month_review = int(Committee_comm_date_review[0:2])
									Committee_comm_year_review = int(Committee_comm_date_review[6:8])
									Committee_comm_day_review = int(Committee_comm_date_review[3:5])
									"""
									dt_obj = datetime.strptime(freeze_date_review,'%m/%d/%y')
									com_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									com_y_review = int(com_date_review[:4])
									com_date_review = date(com_y_review,freeze_month_review,freeze_day_review)
									if(int(qc_rule_review)==3):
										Quality_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.qc_day_review))
									else:
										Quality_date_review = QualityDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,holidays)
									qc_month_review = int(Quality_date_review[0:2])
									qc_day_review = int(Quality_date_review[3:5])
									dt_obj = datetime.strptime(Quality_date_review,'%m/%d/%y')
									qc_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									qc_y_review = int(qc_date_review[:4])
									qc_date_review = date(qc_y_review, qc_month_review, qc_day_review)

									if(int(comm_cal_rule_review)==2):
										comm_cal_ag_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.comm_cal_agent_day_review))
									else:
										comm_cal_ag_date_review = client_comm_date_review
									comm_cal_month_review = int(comm_cal_ag_date_review[0:2])
									comm_cal_day_review = int(comm_cal_ag_date_review[3:5])
									dt_obj = datetime.strptime(comm_cal_ag_date_review,'%m/%d/%y')
									comm_cal_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									comm_cal_y_review = int(comm_cal_date_review[:4])
									comm_cal_date_review = date(comm_cal_y_review, comm_cal_month_review, comm_cal_day_review)
									cal_month_review = month1_review
									month1_review= month1_review+gap_review
									if month!="":
										if((from_month <= e_date_review and e_date_review <= to_month) or (from_month<= qc_date_review and qc_date_review <= to_month) or
											(from_month<= f_date_review and f_date_review <= to_month) or (from_month<= s_date_review and s_date_review <= to_month)
											or (from_month<=a_date_review and a_date_review <= to_month) or (from_month<=cl_date_review and cl_date_review <= to_month)
										 or (from_month<=com_date_review and com_date_review <= to_month) or (from_month<=comm_cal_date_review and comm_cal_date_review <= to_month) ):

											#ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
											ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
											ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
											ind.QC_Date_rev= ind.QC_Date_rev + Quality_date_review + "\n"
											ind.Comm_cal_Date_rev= ind.Comm_cal_Date_rev + comm_cal_ag_date_review + "\n"
											ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
											ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
											ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
											data = {}
											flag=1
									else:
										if(e_date_review <= to_month or f_date_review <= to_month or s_date_review <= to_month or a_date_review <= to_month or (from_month<=cl_date_review and cl_date_review <= to_month)
										 or com_date_review <= to_month or qc_date_review <= to_month or comm_cal_date_review<= to_month):

											#ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
											ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
											ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
											ind.QC_Date_rev= ind.QC_Date_rev + Quality_date_review + "\n"
											ind.Comm_cal_Date_rev= ind.Comm_cal_Date_rev + comm_cal_ag_date_review + "\n"
											ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
											ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
											ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
											data = {}
											flag=1
								from_year = from_year +1


						else:
							pass
						if(flag==1):
							informa.append(ind)

					else:
						print(ind.Index_Name)
						month = str(dt.date.today())
						#print(month)
						#pdb.set_trace()
						year = int(month[:4])
						mon = int(month[5:7])
						d = 1
						holidays = inst.holidays(dt.datetime(year-1, 9, 30), dt.datetime(year, 12, 31))
						from_month = date(year,mon,d)

						year1 = year
						mon1 = 12
						d1 = num_of_days(mon1,year1)
						to_month = date(year1,mon1,d1)
						flag=0;
						ind.Selection_Date_Cycle_1_rec="";
						ind.Completion_Date_rec ="";
						ind.Selection_Date_Cycle_2_rec="";
						ind.Ind_Cmte_Comm_Date_rec=""
						ind.Prelim_Comm_Date_rec=""
						ind.Weights_Share_Freeze_Date_rec=""
						ind.Public_Announcement_Date_rec=""
						ind.Client_Comm_Date_rec=""
						ind.Effective_Date_rec=""
						ind.Ind_Cmte_Comm_Date_re=""
						ind.Prelim_Comm_Date_re=""
						ind.Weights_Share_Freeze_Date_re=""
						ind.Public_Announcement_Date_re=""
						ind.Client_Comm_Date_re=""
						ind.Effective_Date_re=""
						ind.Selection_Date_Cycle_2_rev="";
						ind.Ind_Cmte_Comm_Date_rev=""
						ind.Prelim_Comm_Date_rev=""
						ind.Weights_Share_Freeze_Date_rev=""
						ind.Public_Announcement_Date_rev=""
						ind.Client_Comm_Date_rev=""
						ind.Effective_Date_rev=""
						ind.QC_Date_rec = ""
						ind.QC_Date_re = ""
						ind.QC_Date_rev = ""
						ind.Comm_cal_Date_rec=""
						ind.Comm_cal_Date_re=""
						ind.Comm_cal_Date_rev=""
						if(sel_type1 =="effective"):


							if eff_rule:
								date1 = EffectiveDate_Calculate(month1, year, int(eff_rule),holidays)
								eff_day = int(date1[3:5])
								eff_month = int(date1[0:2])




									#ind.Effective_Date = date1
							if eff_rule_rebal:
								date1_rebal = EffectiveDate_Calculate(month1_rebal, year, int(eff_rule_rebal),holidays)
								eff_day_rebal = int(date1_rebal[3:5])
								eff_mon_rebal = int(date1_rebal[0:2])

							if eff_rule_review:
								date1_review = EffectiveDate_Calculate(month1_review, year, int(eff_rule_review),holidays)
								eff_day_review = int(date1_review[3:5])
								eff_mon_review = int(date1_review[0:2])
								dt_obj = datetime.strptime(date1_review,'%m/%d/%y')
								e_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								e_y_review = int(e_date_review[:4])
								e_date_review = date(e_y_review,eff_mon_review,eff_day_review)

							if eff_rule:
								date1 = EffectiveDate_Calculate(month1, year, int(eff_rule),holidays)
								eff_day = int(date1[3:5])
								eff_month = int(date1[0:2])




									#ind.Effective_Date = date1
							if eff_rule_rebal:
								date1_rebal = EffectiveDate_Calculate(month1_rebal, year, int(eff_rule_rebal),holidays)
								eff_day_rebal = int(date1_rebal[3:5])
								eff_mon_rebal = int(date1_rebal[0:2])

							if eff_rule_review:
								date1_review = EffectiveDate_Calculate(month1_review, year, int(eff_rule_review),holidays)
								eff_day_review = int(date1_review[3:5])
								eff_mon_review = int(date1_review[0:2])
								dt_obj = datetime.strptime(date1_review,'%m/%d/%y')
								e_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								e_y_review = int(e_date_review[:4])
								e_date_review = date(e_y_review,eff_mon_review,eff_day_review)

							for i in range(0,num_month):
								year = int(month[:4])
								if(month1!=""):
									if(month1>12):
										month1 = month1%12;
									if(month1<cal_month):
										year = year+1


								if eff_rule:

									date1 = EffectiveDate_Calculate(month1,year,int(eff_rule),holidays)
									eff_day = int(date1[3:5])
									eff_month = int(date1[0:2])
									eff_year = int(date1[6:8])
									dt_obj = datetime.strptime(date1,'%m/%d/%y')
									e_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									e_y = int(e_date[:4])
									e_date = date(e_y,eff_month,eff_day)
								if sel_rule2:
									if(int(sel_rule2)==7):
										sel_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.sel_cycle_2_day_recon))
									else:
										sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,int(sel_rule2),holidays)
									sel_day = int(sel_date[3:5])
									sel_month = int(sel_date[0:2])
									sel_year = int(sel_date[6:8])

								if fre_rule:
									if(int(fre_rule)==6):
										freeze_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.freeze_day_recon))
									else:
										freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,int(fre_rule),holidays)
									freeze_day = int(freeze_date[3:5])
									freeze_month = int(freeze_date[0:2])
									freeze_year = int(freeze_date[6:8])
								if(int(ann_rule)==3):
										announce_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.announce_day_recon))
								else:
								     if(int(ann_rule)==1):
									      announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(ann_rule),holidays)
								     else:
									      announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(ann_rule),holidays)
								announce_day = int(announce_date[3:5])
								announce_month = int(announce_date[0:2])
								announce_year = int(announce_date[6:8])
								if(int(client_comm_rule)==3):
										client_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.clientcomm_day_recon))
								else:
									if(int(client_comm_rule)==1):
										client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(client_comm_rule),holidays)
									else:
										client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(client_comm_rule),holidays)
								Client_comm_day = int(client_comm_date[3:5])
								Client_comm_month = int(client_comm_date[0:2])
								Client_comm_year = int(client_comm_date[6:8])
								if(Prelim_comm_rule!=""):
									if(int(Prelim_comm_rule)==2):
										#print(ind.Index_Name)
										Prelim_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.prelim_day_recon))
									else:
										Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year,holidays)
									Prelim_comm_day = int(Prelim_comm_date[3:5])
									Prelim_comm_month = int(Prelim_comm_date[0:2])
									Prelim_comm_year = int(Prelim_comm_date[6:8])
								else:
									Prelim_comm_date = ""

								if(int(comp_rule)==2):
									Completion_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.comp_cycle_1_day_recon))
								else:
									Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year,holidays)
								Completion_month = int(Completion_date[0:2])
								"""
								if(icom_rule=="3"):
									Committee_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.indcommittee_day_recon))
								elif(icom_rule=="2"):
									Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year,holidays)
								else:
									Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year,holidays)
								Committee_comm_month = int(Committee_comm_date[0:2])
								Committee_comm_day = int(Committee_comm_date[3:5])
								Committee_comm_year = int(Committee_comm_date[6:8])
								"""
								if(sel_rule==2):
									selection_date1 = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.sel_cycle_1_day_recon))
								elif(sel_rule==1):
									selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year,holidays)
								else:
									selection_date1 =""


								if(int(qc_rule)==3):
									Quality_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.qc_day_recon))
								else:
									Quality_date = QualityDate_Calculate(freeze_day,freeze_month,freeze_year,holidays)

								if(comm_cal_rule=="2"):
									comm_cal_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.comm_cal_agent_day_recon))
								else:
									comm_cal_date = client_comm_date
								if(month1!=""):
									cal_month = month1
									month1= month1+gap
								if month!="":
									if from_month <= e_date and e_date <= to_month and month1!="":
										flag=1
										ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
										ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
										ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
										#ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
										ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
										ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
										ind.QC_Date_rec = ind.QC_Date_rec + Quality_date + "\n"
										ind.Comm_cal_Date_rec = ind.Comm_cal_Date_rec + comm_cal_date + "\n"
										ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
										ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
										ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
										data = {}
								else:
									if e_date <= to_month and month1!="":
										flag=1
										ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
										ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
										ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
										#ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
										ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
										ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
										ind.Comm_cal_Date_rec= ind.Comm_cal_Date_rec= + comm_cal_date + "\n"
										ind.QC_Date_rec = ind.QC_Date_rec + Quality_date + "\n"
										ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
										ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
										ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
										data = {}



							for i in range(0,num_month_review):
								year = int(month[:4])
								if(month1_review!=""):
									if(month1_review>12):
										month1_review = month1_review%12;
									if(month1_review<cal_month_review):
									    year = year+1
								if eff_rule_review:
									date1_review = EffectiveDate_Calculate(month1_review,year,int(eff_rule_review),holidays)
									eff_day_review = int(date1_review[3:5])
									eff_month_review = int(date1_review[0:2])
									eff_year_review = int(date1_review[6:8])
									dt_obj = datetime.strptime(date1_review,'%m/%d/%y')
									e_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									e_y_review = int(e_date_review[:4])
									e_date_review = date(e_y_review,eff_month_review,eff_day_review)

								if (int(sel_rule2_review)==7):
									sel_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.sel_cycle_2_day_review))
								else:
									sel_date_review = SelectionDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(sel_rule2_review),holidays)

								if (int(fre_rule_review)==6):
									freeze_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.freeze_day_review))
								else:
									freeze_date_review = FreezeDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(fre_rule_review),holidays)
								freeze_day_review = int(freeze_date_review[3:5])
								freeze_month_review = int(freeze_date_review[0:2])
								freeze_year_review = int(freeze_date_review[6:8])

								if (int(ann_rule_review)==3):
									announce_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.announce_day_review))
								else:
									if(int(ann_rule_review)==1):
										announce_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(ann_rule_review),holidays)
									else:
										announce_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(ann_rule_review),holidays)
								if (int(client_comm_rule_review)==3):
									client_comm_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.clientcomm_day_review))
								else:
									if(int(client_comm_rule_review)==1):
										client_comm_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(client_comm_rule_review),holidays)
									else:
										client_comm_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(client_comm_rule_review),holidays)
								Client_comm_day_review = int(client_comm_date_review[3:5])
								Client_comm_month_review = int(client_comm_date_review[0:2])
								Client_comm_year_review = int(client_comm_date_review[6:8])

								"""
								if(Prelim_comm_rule_review==""):
									if (icom_rule_review=="3"):
										Committee_comm_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.indcommittee_day_review))
									else:
										Committee_comm_date_review = Committe_comm_date_Calculate(Client_comm_day_review,Client_comm_month_review,Client_comm_year_review,holidays)
								else:
									Committee_comm_date_review = Committe_comm_date_Calculate(Prelim_comm_day_review,Prelim_comm_month_review,Prelim_comm_year_review,holidays)


								Committee_comm_month_review = int(Committee_comm_date_review[0:2])
								Committee_comm_day_review = int(Committee_comm_date_review[3:5])
								Committee_comm_year_review = int(Committee_comm_date_review[6:8])
								"""
								if(int(qc_rule_review)==3):
									Quality_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.qc_day_review))
								else:
									Quality_date_review = QualityDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,holidays)
								if(comm_cal_rule_review=="2"):
									comm_cal_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.comm_cal_agent_day_review))
								else:
									comm_cal_date_review = client_comm_date_review

								if(month1_review!=""):
									cal_month_review = month1_review
									month1_review= month1_review+gap_review
								if month!="":
									if from_month <= e_date_review and e_date_review <= to_month and month1_review!="":
										flag=1

										#ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
										ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
										ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
										ind.QC_Date_rev= ind.QC_Date_rev + Quality_date_review + "\n"
										ind.Comm_cal_Date_rev= ind.Comm_cal_Date_rev + comm_cal_date_review + "\n"
										ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
										ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
										ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
										data = {}
								else:
									if e_date_review <= to_month and month1_review!="":
										flag=1

										#ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
										ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
										ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
										ind.QC_Date_rev= ind.QC_Date_rev + Quality_date_review + "\n"
										ind.Comm_cal_Date_rev= ind.Comm_cal_Date_rev + comm_cal_date_review + "\n"
										ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
										ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
										ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
										data = {}

							for i in range(0,num_month_rebal):
								year = int(month[:4])
								if(month1_rebal!=""):
									if(month1_rebal>12):
										month1_rebal = month1_rebal%12
									if(month1_rebal<cal_month_rebal):
										year = year+1
								if eff_rule_rebal:
									date1_rebal = EffectiveDate_Calculate(month1_rebal,year,int(eff_rule_rebal),holidays)
									eff_day_rebal = int(date1_rebal[3:5])
									eff_month_rebal = int(date1_rebal[0:2])
									eff_year_rebal = int(date1_rebal[6:8])
									dt_obj = datetime.strptime(date1_rebal,'%m/%d/%y')
									e_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
									e_y_rebal = int(e_date_rebal[:4])
									e_date_rebal = date(e_y_rebal,eff_month_rebal,eff_day_rebal)

								if (int(fre_rule_rebal)==6):
									freeze_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.freeze_day_rebal))
								else:
									freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(fre_rule_rebal),holidays)
									freeze_day_rebal = int(freeze_date_rebal[3:5])
									freeze_month_rebal = int(freeze_date_rebal[0:2])
									freeze_year_rebal = int(freeze_date_rebal[6:8])

								if (int(ann_rule_rebal)==3):
									announce_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.announce_day_rebal))
								else:
									if(int(ann_rule_rebal)==1):
										announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(ann_rule_rebal),holidays)
									else:
										announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(ann_rule_rebal),holidays)

								if (int(client_comm_rule_rebal)==3):
									client_comm_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.clientcomm_day_rebal))
								else:
									if(int(client_comm_rule_rebal)==1):
										client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(client_comm_rule_rebal),holidays)
									else:
										client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(client_comm_rule_rebal),holidays)
								if(int(qc_rule_rebal)==3):
									Quality_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.qc_day_rebal))
								else:
									Quality_date_rebal = freeze_date_rebal
								if(comm_cal_rule_rebal=="2"):
									comm_cal_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.comm_cal_agent_day_rebal))
								else:
									comm_cal_date_rebal = client_comm_date_rebal


								if(month1_rebal!=""):
									cal_month_rebal = month1_rebal
									month1_rebal= month1_rebal+gap_rebal
								if month!="":
									if from_month <= e_date_rebal and e_date_rebal <= to_month and month1_rebal!="":
										flag=1
										ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
										ind.QC_Date_re = ind.QC_Date_re  + Quality_date_rebal + "\n"
										ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
										ind.Comm_cal_Date_re= ind.Comm_cal_Date_re + comm_cal_date_rebal + "\n"
										ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
										ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
										data = {}
								else:
									if e_date_rebal <= to_month and month1_rebal!="":
										flag=1
										ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
										ind.QC_Date_re = ind.QC_Date_re  + Quality_date_rebal + "\n"
										ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
										ind.Comm_cal_Date_re= ind.Comm_cal_Date_re + comm_cal_date_rebal + "\n"
										ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
										ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
										data = {}



						elif(sel_type1=="reco" and month1!=""):
							for i in range(0,num_month):
								year = int(month[:4])
								if(month1>12):
									month1 = month1%12;
								if(month1<cal_month):
									year = year+1
								if eff_rule:
									#print(month1)
									#print(year)
									#print(eff_rule)
									#pdb.set_trace()
									date1 = EffectiveDate_Calculate(month1,year,int(eff_rule),holidays)
									eff_day = int(date1[3:5])
									eff_month = int(date1[0:2])
									eff_year = int(date1[6:8])
									dt_obj = datetime.strptime(date1,'%m/%d/%y')
									e_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									e_y = int(e_date[:4])
									e_date = date(e_y,eff_month,eff_day)
								if sel_rule2:
									if(int(sel_rule2)==7):
										sel_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.sel_cycle_2_day_recon))
									else:
										sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,int(sel_rule2),holidays)
									sel_day = int(sel_date[3:5])
									sel_month = int(sel_date[0:2])
									sel_year = int(sel_date[6:8])
									dt_obj = datetime.strptime(sel_date,'%m/%d/%y')
									s_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									s_y = int(s_date[:4])
									s_date = date(s_y,sel_month,sel_day)

								if fre_rule:
									if(int(fre_rule)==6):
										freeze_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.freeze_day_recon))
									else:
										freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,int(fre_rule),holidays)
									freeze_day = int(freeze_date[3:5])
									freeze_month = int(freeze_date[0:2])
									freeze_year = int(freeze_date[6:8])
									dt_obj = datetime.strptime(freeze_date,'%m/%d/%y')
									f_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									f_y = int(f_date[:4])
									f_date = date(f_y,freeze_month,freeze_day)
								if(int(ann_rule)==3):
										announce_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.announce_day_recon))
								else:
								     if(int(ann_rule)==1):
									      announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(ann_rule),holidays)
								     else:
									      announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(ann_rule),holidays)
								announce_day = int(announce_date[3:5])
								announce_month = int(announce_date[0:2])
								announce_year = int(announce_date[6:8])
								dt_obj = datetime.strptime(announce_date,'%m/%d/%y')
								a_date = datetime.strftime(dt_obj, "%Y-%m-%d")
								a_y = int(a_date[:4])
								a_date = date(a_y,announce_month,announce_day)

								if(int(client_comm_rule)==3):
										client_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.clientcomm_day_recon))
								else:
									if(int(client_comm_rule)==1):
										client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(client_comm_rule),holidays)
									else:
										client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(client_comm_rule),holidays)
								Client_comm_day = int(client_comm_date[3:5])
								Client_comm_month = int(client_comm_date[0:2])
								Client_comm_year = int(client_comm_date[6:8])
								dt_obj = datetime.strptime(client_comm_date,'%m/%d/%y')
								cl_date = datetime.strftime(dt_obj, "%Y-%m-%d")
								cl_y = int(cl_date[:4])
								cl_date = date(cl_y,Client_comm_month,Client_comm_day)
								if(Prelim_comm_rule!=""):
									if(int(Prelim_comm_rule)==2):
										Prelim_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.prelim_day_recon))
									else:
										Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year,holidays)
									Prelim_comm_day = int(Prelim_comm_date[3:5])
									Prelim_comm_month = int(Prelim_comm_date[0:2])
									Prelim_comm_year = int(Prelim_comm_date[6:8])
									dt_obj = datetime.strptime(Prelim_comm_date,'%m/%d/%y')
									p_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									p_y = int(p_date[:4])
									p_date = date(p_y,Prelim_comm_month,Prelim_comm_day)
								else:
									Prelim_comm_date=""


								if(int(comp_rule)==2):
									Completion_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.comp_cycle_1_day_recon))
								else:
									Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year,holidays)
								Completion_month = int(Completion_date[0:2])
								Completion_day = int(Completion_date[3:5])
								dt_obj = datetime.strptime(Completion_date,'%m/%d/%y')
								c_date = datetime.strftime(dt_obj, "%Y-%m-%d")
								c_y = int(c_date[:4])
								c_date = date(c_y,Completion_month,Completion_day)
								"""
								if(icom_rule=="3"):
									Committee_comm_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.indcommittee_day_recon))
								elif(icom_rule=="2"):
									Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year,holidays)
								else:
									Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year,holidays)
								Committee_comm_month = int(Committee_comm_date[0:2])
								Committee_comm_day = int(Committee_comm_date[3:5])
								Committee_comm_year = int(Committee_comm_date[6:8])
								"""
								dt_obj = datetime.strptime(freeze_date,'%m/%d/%y')
								com_date = datetime.strftime(dt_obj, "%Y-%m-%d")
								com_y = int(com_date[:4])
								com_date = date(com_y,freeze_month,freeze_day)
								if(sel_rule==2):
									selection_date1 = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.sel_cycle_1_day_recon))
									sel_1_month = int(selection_date1[0:2])
									sel_1_day = int(selection_date1[3:5])
									dt_obj = datetime.strptime(selection_date1,'%m/%d/%y')
									s1_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									s1_y = int(s1_date[:4])
									sel_1_date = date(s1_y, sel_1_month, sel_1_day)
								elif(sel_rule==1):
									selection_date1 = selection_date_1_Calculate(sel_day, sel_month, sel_year,holidays)
									sel_1_month = int(selection_date1[0:2])
									sel_1_day = int(selection_date1[3:5])
									dt_obj = datetime.strptime(selection_date1,'%m/%d/%y')
									s1_date = datetime.strftime(dt_obj, "%Y-%m-%d")
									s1_y = int(s1_date[:4])
									sel_1_date = date(s1_y, sel_1_month, sel_1_day)
								else:
									selection_date1 =""


								if(int(qc_rule)==3):
									Quality_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.qc_day_recon))
								else:
									Quality_date = QualityDate_Calculate(freeze_day,freeze_month,freeze_year,holidays)
								qc_month = int(Quality_date[0:2])
								qc_day = int(Quality_date[3:5])
								dt_obj = datetime.strptime(Quality_date,'%m/%d/%y')
								qc_date = datetime.strftime(dt_obj, "%Y-%m-%d")
								qc_y = int(qc_date[:4])
								qc_date = date(qc_y, qc_month, qc_day)
								if(int(comm_cal_rule)==2):
									comm_cal_ag_date = priorday_calculate(eff_day,eff_month,eff_year,holidays,int(priordays.comm_cal_agent_day_recon))
								else:
									comm_cal_ag_date = client_comm_date
								comm_cal_month = int(comm_cal_ag_date[0:2])
								comm_cal_day = int(comm_cal_ag_date[3:5])
								dt_obj = datetime.strptime(comm_cal_ag_date,'%m/%d/%y')
								comm_cal_date = datetime.strftime(dt_obj, "%Y-%m-%d")
								comm_cal_y = int(comm_cal_date[:4])
								comm_cal_date = date(comm_cal_y, comm_cal_month, comm_cal_day)

								cal_month = month1
								month1= month1+gap
								if month!="":
									if((from_month <= e_date and e_date <= to_month)  or (from_month<= s_date and s_date <= to_month )or (from_month<= f_date and f_date <= to_month)
										or (from_month <= comm_cal_date and comm_cal_date <= to_month)
										or (from_month<=a_date and a_date <= to_month) or (from_month<=cl_date and cl_date <= to_month) or (from_month<=qc_date and qc_date <= to_month)
									 or (from_month<=c_date and c_date <= to_month) or (from_month<=com_date and com_date <= to_month) ):
										flag=1
										#print(to_month)
										#or (from_month <= sel_1_date and sel_1_date <= to_month ) 
										#or (from_month <= p_date and p_date <= to_month)
										ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
										ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
										ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
										#ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
										ind.Comm_cal_Date_rec= ind.Comm_cal_Date_rec + comm_cal_ag_date + "\n"
										ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
										ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
										ind.QC_Date_rec = ind.QC_Date_rec + Quality_date + "\n"
										ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
										ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
										ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
										data = {}
								else:
									if(e_date <= to_month or s_date <= to_month or f_date <= to_month
										or a_date <= to_month or cl_date <= to_month or qc_date <= to_month or comm_cal_date <= to_month or c_date <= to_month or sel_1_date <= to_month or com_date <= to_month ):
										flag=1
										#or p_date <= to_month
										ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
										ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
										ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
										#ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
										ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
										ind.Comm_cal_Date_rec= ind.Comm_cal_Date_rec + "\n"
										ind.QC_Date_rec = ind.QC_Date_rec + Quality_date + "\n"
										ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
										ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
										ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
										ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
										data = {}



						elif(sel_type1=="reb" and month1_rebal!=""):
							for i in range(0,num_month_rebal):
								year = int(month[:4])
								if(month1_rebal>12):
									month1_rebal = month1_rebal%12;
								if(month1_rebal<cal_month_rebal):
									year = year+1

								if eff_rule_rebal:
										date1_rebal = EffectiveDate_Calculate(month1_rebal,year,int(eff_rule_rebal),holidays)
										eff_day_rebal = int(date1_rebal[3:5])
										eff_month_rebal = int(date1_rebal[0:2])
										eff_year_rebal = int(date1_rebal[6:8])
										dt_obj = datetime.strptime(date1_rebal,'%m/%d/%y')
										e_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
										e_y_rebal = int(e_date_rebal[:4])
										e_date_rebal = date(e_y_rebal,eff_month_rebal,eff_day_rebal)
										if (int(fre_rule_rebal)==6):
											freeze_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.freeze_day_rebal))
										else:
											freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(fre_rule_rebal),holidays)
										freeze_day_rebal = int(freeze_date_rebal[3:5])
										freeze_month_rebal = int(freeze_date_rebal[0:2])
										freeze_year_rebal = int(freeze_date_rebal[6:8])
										dt_obj = datetime.strptime(freeze_date_rebal,'%m/%d/%y')
										f_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
										f_y_rebal = int(f_date_rebal[:4])
										f_date_rebal = date(f_y_rebal,freeze_month_rebal,freeze_day_rebal)


								if (int(ann_rule_rebal)==3):
									announce_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.announce_day_rebal))
								else:
									if(int(ann_rule_rebal)==1):
										announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(ann_rule_rebal),holidays)
									else:
										announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(ann_rule_rebal),holidays)
								announce_day_rebal = int(announce_date_rebal[3:5])
								announce_month_rebal = int(announce_date_rebal[0:2])
								announce_year_rebal = int(announce_date_rebal[6:8])
								dt_obj = datetime.strptime(announce_date_rebal,'%m/%d/%y')
								a_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
								a_y_rebal = int(a_date_rebal[:4])
								a_date_rebal = date(a_y_rebal,announce_month_rebal,announce_day_rebal)
								if (int(client_comm_rule_rebal)==3):
									client_comm_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.clientcomm_day_rebal))
								else:
									if(int(client_comm_rule_rebal)==1):
										client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(client_comm_rule_rebal),holidays)
									else:
										client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(client_comm_rule_rebal),holidays)
								Client_comm_day_rebal = int(client_comm_date_rebal[3:5])
								Client_comm_month_rebal = int(client_comm_date_rebal[0:2])
								Client_comm_year_rebal = int(client_comm_date_rebal[6:8])
								dt_obj = datetime.strptime(client_comm_date_rebal,'%m/%d/%y')
								cl_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
								cl_y_rebal = int(cl_date_rebal[:4])
								cl_date_rebal = date(cl_y_rebal,Client_comm_month_rebal,Client_comm_day_rebal)

								if(int(qc_rule_rebal)==3):
									Quality_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.qc_day_rebal))
								else:
									Quality_date_rebal = freeze_date_rebal
								qc_month_rebal = int(Quality_date_rebal[0:2])
								qc_day_rebal = int(Quality_date_rebal[3:5])
								dt_obj = datetime.strptime(Quality_date_rebal,'%m/%d/%y')
								qc_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
								qc_y_rebal = int(qc_date_rebal[:4])
								qc_date_rebal = date(qc_y_rebal, qc_month_rebal, qc_day_rebal)

								if(int(comm_cal_rule_rebal)==2):
									comm_cal_ag_date_rebal = priorday_calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,holidays,int(priordays.comm_cal_agent_day_rebal))
								else:
									comm_cal_ag_date_rebal = client_comm_date_rebal
								comm_cal_month_rebal = int(comm_cal_ag_date_rebal[0:2])
								comm_cal_day_rebal = int(comm_cal_ag_date_rebal[3:5])
								dt_obj = datetime.strptime(comm_cal_ag_date_rebal,'%m/%d/%y')
								comm_cal_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
								comm_cal_y_rebal = int(comm_cal_date_rebal[:4])
								comm_cal_date_rebal = date(comm_cal_y_rebal, comm_cal_month_rebal, comm_cal_day_rebal)


								cal_month_rebal = month1_rebal
								month1_rebal= month1_rebal+gap_rebal
								if month!="":
									if((from_month <= e_date_rebal and e_date_rebal <= to_month) or
										(from_month<= f_date_rebal and f_date_rebal <= to_month) or (from_month<= qc_date_rebal and qc_date_rebal <= to_month)
										or (from_month<=a_date_rebal and a_date_rebal <= to_month) or (from_month<=cl_date_rebal and cl_date_rebal <= to_month)
									  or (from_month<= comm_cal_date_rebal and comm_cal_date_rebal <= to_month)):
										ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
										ind.QC_Date_re = ind.QC_Date_re  + Quality_date_rebal + "\n"
										ind.Comm_cal_Date_re= ind.Comm_cal_Date_re + comm_cal_ag_date_rebal + "\n"
										ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
										ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
										ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
										data = {}
										flag=1
								else:
									if(e_date_rebal <= to_month or comm_cal_date_rebal <= to_month or f_date_rebal <= to_month or a_date_rebal <= to_month or cl_date_rebal <= to_month
									 or qc_date_rebal <= to_month):
										ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
										ind.QC_Date_re = ind.QC_Date_re  + Quality_date_rebal + "\n"
										ind.Comm_cal_Date_re= ind.Comm_cal_Date_re + comm_cal_ag_date_rebal + "\n"
										ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
										ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
										ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
										data = {}
										flag=1


						elif(sel_type1=="rev" and month1_review!=""):
							for i in range(0,num_month_review):
								year = int(month[:4])
								if(month1_review>12):
									month1_review = month1_review%12;
								if(month1_review<cal_month_review):
									year = year+1
								if eff_rule_review:
									date1_review = EffectiveDate_Calculate(month1_review,year,int(eff_rule_review),holidays)
									eff_day_review = int(date1_review[3:5])
									eff_month_review = int(date1_review[0:2])
									eff_year_review = int(date1_review[6:8])
									dt_obj = datetime.strptime(date1_review,'%m/%d/%y')
									e_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
									e_y_review = int(e_date_review[:4])
									e_date_review = date(e_y_review,eff_month_review,eff_day_review)

								if(int(sel_rule2_review)==7):
								    sel_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.sel_cycle_2_day_review))
								else:
									sel_date_review = SelectionDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(sel_rule2_review),holidays)
								sel_day_review = int(sel_date_review[3:5])
								sel_month_review = int(sel_date_review[0:2])
								sel_year_review = int(sel_date_review[6:8])
								dt_obj = datetime.strptime(sel_date_review,'%m/%d/%y')
								s_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								s_y_review = int(s_date_review[:4])
								s_date_review = date(s_y_review,sel_month_review,sel_day_review)
								
								if (int(fre_rule_review)==6):
									freeze_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.freeze_day_review))
								else:
									freeze_date_review = FreezeDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(fre_rule_review),holidays)
								
								freeze_day_review = int(freeze_date_review[3:5])
								freeze_month_review = int(freeze_date_review[0:2])
								freeze_year_review = int(freeze_date_review[6:8])
								dt_obj = datetime.strptime(freeze_date_review,'%m/%d/%y')
								f_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								f_y_review = int(f_date_review[:4])
								f_date_review = date(f_y_review,freeze_month_review,freeze_day_review)

								if (int(ann_rule_review)==3):
									announce_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.announce_day_review))
								elif(int(ann_rule_review)==1):
									announce_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(ann_rule_review),holidays)
								else:
									announce_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(ann_rule_review),holidays)
								
								announce_day_review = int(announce_date_review[3:5])
								announce_month_review = int(announce_date_review[0:2])
								announce_year_review = int(announce_date_review[6:8])
								dt_obj = datetime.strptime(announce_date_review,'%m/%d/%y')
								a_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								a_y_review = int(a_date_review[:4])
								a_date_review = date(a_y_review,announce_month_review,announce_day_review)
								if (int(client_comm_rule_review)==3):
									client_comm_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.clientcomm_day_review))
								elif(int(client_comm_rule_review)==1):
									client_comm_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(client_comm_rule_review),holidays)
								else:
									client_comm_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(client_comm_rule_review),holidays)
								Client_comm_day_review = int(client_comm_date_review[3:5])
								Client_comm_month_review = int(client_comm_date_review[0:2])
								Client_comm_year_review = int(client_comm_date_review[6:8])
								dt_obj = datetime.strptime(client_comm_date_review,'%m/%d/%y')
								cl_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								cl_y_review = int(cl_date_review[:4])
								cl_date_review = date(cl_y_review,Client_comm_month_review,Client_comm_day_review)
								"""
								if(Prelim_comm_rule_review==""):
									if (icom_rule_review=="3"):
										Committee_comm_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.indcommittee_day_review))
									else:
										Committee_comm_date_review = Committe_comm_date_Calculate(Client_comm_day_review,Client_comm_month_review,Client_comm_year_review,holidays)
								else:
										Committee_comm_date_review = Committe_comm_date_Calculate(Prelim_comm_day_review,Prelim_comm_month_review,Prelim_comm_year_review,holidays)

								Committee_comm_month_review = int(Committee_comm_date_review[0:2])
								Committee_comm_year_review = int(Committee_comm_date_review[6:8])
								Committee_comm_day_review = int(Committee_comm_date_review[3:5])
								"""
								dt_obj = datetime.strptime(freeze_date_review,'%m/%d/%y')
								com_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								com_y_review = int(com_date_review[:4])
								com_date_review = date(com_y_review,freeze_month_review,freeze_day_review)
								if(int(qc_rule_review)==3):
									Quality_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.qc_day_review))
								else:
									Quality_date_review = QualityDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,holidays)
								qc_month_review = int(Quality_date_review[0:2])
								qc_day_review = int(Quality_date_review[3:5])
								dt_obj = datetime.strptime(Quality_date_review,'%m/%d/%y')
								qc_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								qc_y_review = int(qc_date_review[:4])
								qc_date_review = date(qc_y_review, qc_month_review, qc_day_review)

								if(int(comm_cal_rule_review)==2):
									comm_cal_ag_date_review = priorday_calculate(eff_day_review,eff_month_review,eff_year_review,holidays,int(priordays.comm_cal_agent_day_review))
								else:
									comm_cal_ag_date_review = client_comm_date_review
								comm_cal_month_review = int(comm_cal_ag_date_review[0:2])
								comm_cal_day_review = int(comm_cal_ag_date_review[3:5])
								dt_obj = datetime.strptime(comm_cal_ag_date_review,'%m/%d/%y')
								comm_cal_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
								comm_cal_y_review = int(comm_cal_date_review[:4])
								comm_cal_date_review = date(comm_cal_y_review, comm_cal_month_review, comm_cal_day_review)
								cal_month_review = month1_review
								month1_review= month1_review+gap_review
								if month!="":
									if((from_month <= e_date_review and e_date_review <= to_month) or (from_month<= qc_date_review and qc_date_review <= to_month) or
										(from_month<= f_date_review and f_date_review <= to_month) or (from_month<= s_date_review and s_date_review <= to_month)
										or (from_month<=a_date_review and a_date_review <= to_month) or (from_month<=cl_date_review and cl_date_review <= to_month)
									 or (from_month<=com_date_review and com_date_review <= to_month) or (from_month<=comm_cal_date_review and comm_cal_date_review <= to_month) ):

										#ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
										ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
										ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
										ind.QC_Date_rev= ind.QC_Date_rev + Quality_date_review + "\n"
										ind.Comm_cal_Date_rev= ind.Comm_cal_Date_rev + comm_cal_ag_date_review + "\n"
										ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
										ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
										ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
										data = {}
										flag=1
								else:
									if(e_date_review <= to_month or f_date_review <= to_month or s_date_review <= to_month or a_date_review <= to_month or (from_month<=cl_date_review and cl_date_review <= to_month)
									 or com_date_review <= to_month or qc_date_review <= to_month or comm_cal_date_review<= to_month):

										#ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
										ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
										ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
										ind.QC_Date_rev= ind.QC_Date_rev + Quality_date_review + "\n"
										ind.Comm_cal_Date_rev= ind.Comm_cal_Date_rev + comm_cal_ag_date_review + "\n"
										ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
										ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
										ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
										data = {}
										flag=1



						else:
							pass

						if(flag==1):
							informa.append(ind)	


			'''page = request.GET.get('page',1)
			paginator = Paginator(informa, 10)
			try:
				informa = paginator.page(page)
			except PageNotAnInteger:
				informa = paginator.page(1)
			except EmptyPage:
				informa = paginator.page(paginator.num_pages)'''
			print(len(informa))
			now = dt.datetime.now()
			context = {
				'info' : informa,
				'client_name' : client_name,
				'client_code' : cname1,
				'from_MY' : month,
				'to_MY' : month2,
				'sel_type1' : sel_type1,
				'curr_year' : now.year,
				'index_name' : indexName,


			}



	else:
		'''page = request.GET.get('page',1)
		paginator = Paginator(info, 10)
		try:
			info = paginator.page(page)
		except PageNotAnInteger:
			info = paginator.page(1)
		except EmptyPage:
			info = paginator.page(paginator.num_pages)'''

		context = {
			'info' : info,
			'client_name' : client_name,

		}
	templateName = "reportgenerate.html"
	return render(request, templateName, context)


def logout(request):
	try:
		del request.session['username']
	except:
		pass
	templateName = "login.html"
	context = {"message":"You are logged out."}
	return render(request, templateName, context)


def error_404_view(request, exception):
	templateName = "404.html"
	context = {}
	return render(request, templateName, context)

def error_500_view(request, exception):
	templateName = "500.html"
	context = {}
	return render(request, templateName, context)

def export_index_to_xlsx(request):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    Regindex_queryset = Registerindex.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the column titles and widths
    columns = [
        ('ID', 8),
        ('Index_Name', 40),
        ('Ident_ISIN', 80),
        ('Ident_Bloomberg', 15),
        ('Ind_Ver', 15),
        ('Ind_Ver_ID', 15),
    ]

    # Iterate through movie categories
    for category_index, category in enumerate(Regindex_queryset):
        # Create a worksheet/tab with the title of the category
        worksheet = workbook.create_sheet(
            title=category.Index_Name,
            index=category_index,
        )
        # Define the background color of the header cells

        fill = PatternFill(
            #start_color=category.html_color,
            #end_color=category.html_color,
            fill_type="solid",
        )
        row_num = 1

        # Assign values, styles, and formatting for each cell in the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            #cell.fill = fill
            # set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

        # Iterate through all movies of a category
        for movie in Registerindex.objects.all():
            row_num += 1

            # Define data and formats for each cell in the row
            row = [
                (movie.pk, 'Normal'),
                (movie.Index_Name, 'Normal'),
                (movie.Ident_ISIN, 'Normal'),
                #(timedelta(minutes=movie.length_in_minutes), 'Normal'),
				(movie.Ident_Bloomberg, 'Normal'),
                (movie.Ind_Ver, 'Normal'),
                (movie.Ind_Ver_ID, 'Normal'),
            ]

            # Assign values, styles, and formatting for each cell in the row
            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                '''if cell_format == 'Currency':
                    cell.number_format = '#,##0.00 €'
                if col_num == 4:
                    cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment
				'''

        # freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # set tab color
        #worksheet.sheet_properties.tabColor = category.html_color

    workbook.save(response)
    return response


def next_weekday(d, weekday):
    days_ahead = weekday - d.weekday()
    if days_ahead <= 0: # Target day already happened this week
        days_ahead += 7
    return d + timedelta(days_ahead)

from django.template.loader import render_to_string
from django.utils.html import strip_tags
def mail_index(request):
	client_name = Calendarlist.objects.filter(category = 'Client_Name')
	info = Registerindex.objects.all().order_by('id').reverse()
	inst = USTradingCalendar()
	y1 = date.today().year
	months = dict(January=1, February=2, March=3, April = 4, May=5, June=6, July=7, August=8,September=9, October=10, November=11, December=12)
	holidays = inst.holidays(dt.datetime(y1-1, 9, 30), dt.datetime(y1, 12, 31))
	informa=[]
	months = dict(January=1, February=2, March=3, April = 4, May=5, June=6, July=7, August=8,September=9, October=10, November=11, December=12)
	for ind in info:
		eff_rule=""
		eff_rule_review="";
		eff_rule_rebal="";
		month1="";
		eff_mon=0;
		month1_rebal="";
		eff_mon_rebal=0;
		month1_review="";
		eff_mon_review=0;
		num_month=0;
		num_month_rebal=0;
		num_month_review=0;

		rulelist = Ruleslist.objects.get(index_id=ind.id)
		if(rulelist.eff_rule!=""):
			eff_rule = rulelist.eff_rule
			sel_rule = rulelist.sel_rule1
			sel_rule2 = rulelist.sel_rule2
			fre_rule = rulelist.freeze_rule
			ann_rule = rulelist.announce_rule
			client_comm_rule = rulelist.clientcomm_rule
			Prelim_comm_rule = rulelist.prelim_rule
			month1 = months[ind.reconst_month]
			time = ind.Reconstitution
			cal_month = month1
			if(time=="Annual"):
				num_month=1
				gap=12
			elif(time=="Monthly"):
				num_month =12
				gap=1
			elif(time=="Quarterly"):
				num_month=4
				gap=3
			else:
				num_month=2
				gap=6
			month1 = int(month1)
		if(rulelist.eff_rule_rebal!=""):
			eff_rule_rebal = rulelist.eff_rule_rebal
			sel_rule_rebal = rulelist.sel_rule1_rebal
			sel_rule2_rebal = rulelist.sel_rule2_rebal
			fre_rule_rebal = rulelist.freeze_rule_rebal
			ann_rule_rebal = rulelist.announce_rule_rebal
			client_comm_rule_rebal = rulelist.clientcomm_rule_rebal
			Prelim_comm_rule_rebal = rulelist.prelim_rule_rebal
			month1_rebal = months[ind.rebalance_month]

			month1_rebal = int(month1_rebal)

			time_rebal = ind.Rebalance
			cal_month_rebal = month1_rebal
			if(time_rebal=="Annual"):
				num_month_rebal=1
				gap_rebal=12
			elif(time_rebal=="Monthly"):
				num_month_rebal =12
				gap_rebal=1
			elif(time_rebal=="Quarterly"):
				num_month_rebal=4
				gap_rebal=3
			else:
				num_month_rebal=2
				gap_rebal=6

		if(rulelist.eff_rule_review!=""):
			eff_rule_review = rulelist.eff_rule_review
			sel_rule2_review = rulelist.sel_rule2_review
			fre_rule_review = rulelist.freeze_rule_review
			ann_rule_review = rulelist.announce_rule_review
			client_comm_rule_review = rulelist.clientcomm_rule_review
			Prelim_comm_rule_review = rulelist.prelim_rule_review
			month1_review = months[ind.review_month]
			month1_review = int(month1_review)

			time_review = ind.Review
			cal_month_review = month1_review
			if(time_review=="Annual"):
				num_month_review=1
				gap_review=12
			elif(time_review=="Monthly"):
				num_month_review =12
				gap_review=1
			elif(time_review=="Quarterly"):
				num_month_review=4
				gap_review=3
			else:
				num_month_review=2
				gap_review=6
		d = date.today()
		next_monday = next_weekday(d, 0)
		next_friday = next_weekday(next_monday, 4)

		ind.Selection_Date_Cycle_1_rec="";
		ind.Completion_Date_rec ="";
		ind.Selection_Date_Cycle_2_rec="";
		ind.Ind_Cmte_Comm_Date_rec=""
		ind.Prelim_Comm_Date_rec=""
		ind.Weights_Share_Freeze_Date_rec=""
		ind.Public_Announcement_Date_rec=""
		ind.Client_Comm_Date_rec=""
		ind.Effective_Date_rec=""
		ind.Ind_Cmte_Comm_Date_re=""
		ind.Prelim_Comm_Date_re=""
		ind.Weights_Share_Freeze_Date_re=""
		ind.Public_Announcement_Date_re=""
		ind.Client_Comm_Date_re=""
		ind.Effective_Date_re=""
		ind.Selection_Date_Cycle_2_rev="";
		ind.Ind_Cmte_Comm_Date_rev=""
		ind.Prelim_Comm_Date_rev=""
		ind.Weights_Share_Freeze_Date_rev=""
		ind.Public_Announcement_Date_rev=""
		ind.Client_Comm_Date_rev=""
		ind.Effective_Date_rev=""
		flag=0

		for i in range(0,num_month):
			year = y1
			if(month1>12):
				month1 = month1%12;
			if(month1<cal_month):
				year = year+1
			if eff_rule:
				date1 = EffectiveDate_Calculate(month1,year,int(eff_rule),holidays)
				eff_day = int(date1[3:5])
				eff_month = int(date1[0:2])
				eff_year = int(date1[6:8])
				dt_obj = datetime.strptime(date1,'%m/%d/%y')
				e_date = datetime.strftime(dt_obj, "%Y-%m-%d")
				e_y = int(e_date[:4])
				e_date = date(e_y,eff_month,eff_day)
			if sel_rule2:
				sel_date = SelectionDate_Calculate(eff_day,eff_month,eff_year,int(sel_rule2),holidays)
				sel_day = int(sel_date[3:5])
				sel_month = int(sel_date[0:2])
				sel_year = int(sel_date[6:8])
				dt_obj = datetime.strptime(sel_date,'%m/%d/%y')
				s_date = datetime.strftime(dt_obj, "%Y-%m-%d")
				s_y = int(s_date[:4])
				s_date = date(s_y,sel_month,sel_day)

			if fre_rule:
				freeze_date = FreezeDate_Calculate(eff_day,eff_month,eff_year,int(fre_rule),holidays)
				freeze_day = int(freeze_date[3:5])
				freeze_month = int(freeze_date[0:2])
				freeze_year = int(freeze_date[6:8])
				dt_obj = datetime.strptime(freeze_date,'%m/%d/%y')
				f_date = datetime.strftime(dt_obj, "%Y-%m-%d")
				f_y = int(f_date[:4])
				f_date = date(f_y,freeze_month,freeze_day)
			if(int(ann_rule)==1):
				announce_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(ann_rule),holidays)
			else:
				announce_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(ann_rule),holidays)
			announce_day = int(announce_date[3:5])
			announce_month = int(announce_date[0:2])
			announce_year = int(announce_date[6:8])
			dt_obj = datetime.strptime(announce_date,'%m/%d/%y')
			a_date = datetime.strftime(dt_obj, "%Y-%m-%d")
			a_y = int(a_date[:4])
			a_date = date(a_y,announce_month,announce_day)

			if(int(client_comm_rule)==1):
				client_comm_date = AnnouncementDate_Calculate(eff_day,eff_month,eff_year,int(client_comm_rule),holidays)
			else:
				client_comm_date = AnnouncementDate_Calculate(freeze_day,freeze_month,freeze_year,int(client_comm_rule),holidays)
			Client_comm_day = int(client_comm_date[3:5])
			Client_comm_month = int(client_comm_date[0:2])
			Client_comm_year = int(client_comm_date[6:8])
			dt_obj = datetime.strptime(client_comm_date,'%m/%d/%y')
			cl_date = datetime.strftime(dt_obj, "%Y-%m-%d")
			cl_y = int(cl_date[:4])
			cl_date = date(cl_y,Client_comm_month,Client_comm_day)
			Prelim_comm_date = PreliminaryCommDate_Calculate(announce_day,announce_month,announce_year,holidays)
			Prelim_comm_day = int(Prelim_comm_date[3:5])
			Prelim_comm_month = int(Prelim_comm_date[0:2])
			Prelim_comm_year = int(Prelim_comm_date[6:8])
			dt_obj = datetime.strptime(Prelim_comm_date,'%m/%d/%y')
			p_date = datetime.strftime(dt_obj, "%Y-%m-%d")
			p_y = int(p_date[:4])
			p_date = date(p_y,Prelim_comm_month,Prelim_comm_day)

			Completion_date = CompletionDate_Calculate(sel_day,sel_month,sel_year,holidays)
			Completion_month = int(Completion_date[0:2])
			Completion_day = int(Completion_date[3:5])
			dt_obj = datetime.strptime(Completion_date,'%m/%d/%y')
			c_date = datetime.strftime(dt_obj, "%Y-%m-%d")
			c_y = int(c_date[:4])
			c_date = date(c_y,Completion_month,Completion_day)

			if(Prelim_comm_rule==""):
				Committee_comm_date = Committe_comm_date_Calculate(Client_comm_day,Client_comm_month,Client_comm_year,holidays)
			else:
				Committee_comm_date = Committe_comm_date_Calculate(Prelim_comm_day,Prelim_comm_month,Prelim_comm_year,holidays)
			Committee_comm_month = int(Committee_comm_date[0:2])
			Committee_comm_day = int(Committee_comm_date[3:5])
			dt_obj = datetime.strptime(Committee_comm_date,'%m/%d/%y')
			com_date = datetime.strftime(dt_obj, "%Y-%m-%d")
			com_y = int(com_date[:4])
			com_date = date(com_y,Committee_comm_month,Committee_comm_day)
			selection_date1 = selection_date_1_Calculate(sel_day,sel_month,sel_year,holidays)
			sel_1_month = int(selection_date1[0:2])
			sel_1_day = int(selection_date1[3:5])
			dt_obj = datetime.strptime(selection_date1,'%m/%d/%y')
			s1_date = datetime.strftime(dt_obj, "%Y-%m-%d")
			s1_y = int(s1_date[:4])
			sel_1_date = date(s1_y, sel_1_month, sel_1_day)

			cal_month = month1
			month1= month1+gap

			if(next_monday <= e_date and e_date <= next_friday):

				flag=1
				ind.Effective_Date_rec = ind.Effective_Date_rec + date1 + "\n"
			if(next_monday <= s_date and s_date <= next_friday):
				flag=1
				ind.Selection_Date_Cycle_2_rec = ind.Selection_Date_Cycle_2_rec + sel_date + "\n"
			if(next_monday <= f_date and f_date <= next_friday):
				flag=1
				ind.Weights_Share_Freeze_Date_rec = ind.Weights_Share_Freeze_Date_rec + freeze_date + "\n"
			if(next_monday <= a_date and a_date <= next_friday):
				flag=1
				ind.Public_Announcement_Date_rec = ind.Public_Announcement_Date_rec + announce_date + "\n"
			if(next_monday <= cl_date and cl_date <= next_friday):
				flag=1
				ind.Client_Comm_Date_rec = ind.Client_Comm_Date_rec + client_comm_date + "\n"
			"""
			if(next_monday <= p_date and p_date <= next_friday):
				flag=1
				ind.Prelim_Comm_Date_rec = ind.Prelim_Comm_Date_rec + Prelim_comm_date + "\n"
			"""
			if(next_monday <= c_date and c_date <= next_friday):
				flag=1
				ind.Completion_Date_rec = ind.Completion_Date_rec + Completion_date + "\n"
			if(next_monday <= com_date and com_date <= next_friday):
				flag=1
				ind.Ind_Cmte_Comm_Date_rec = ind.Ind_Cmte_Comm_Date_rec + Committee_comm_date + "\n"
			if(next_monday <= sel_1_date and sel_1_date <= next_friday):
				flag=1
				ind.Selection_Date_Cycle_1_rec = ind.Selection_Date_Cycle_1_rec + selection_date1 + "\n"
		for i in range(0,num_month_rebal):
			year = y1
			if(month1_rebal>12):
				month1_rebal = month1_rebal%12;
			if(month1_rebal<cal_month_rebal):
				year = year+1

			if eff_rule_rebal:
				date1_rebal = EffectiveDate_Calculate(month1_rebal,year,int(eff_rule_rebal),holidays)
				eff_day_rebal = int(date1_rebal[3:5])
				eff_month_rebal = int(date1_rebal[0:2])
				eff_year_rebal = int(date1_rebal[6:8])
				dt_obj = datetime.strptime(date1_rebal,'%m/%d/%y')
				e_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
				e_y_rebal = int(e_date_rebal[:4])
				e_date_rebal = date(e_y_rebal,eff_month_rebal,eff_day_rebal)
				freeze_date_rebal = FreezeDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(fre_rule_rebal),holidays)
				freeze_day_rebal = int(freeze_date_rebal[3:5])
				freeze_month_rebal = int(freeze_date_rebal[0:2])
				freeze_year_rebal = int(freeze_date_rebal[6:8])
				dt_obj = datetime.strptime(freeze_date_rebal,'%m/%d/%y')
				f_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
				f_y_rebal = int(f_date_rebal[:4])
				f_date_rebal = date(f_y_rebal,freeze_month_rebal,freeze_day_rebal)


			if(int(ann_rule_rebal)==1):
				announce_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(ann_rule_rebal),holidays)
			else:
				announce_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(ann_rule_rebal),holidays)
			announce_day_rebal = int(announce_date_rebal[3:5])
			announce_month_rebal = int(announce_date_rebal[0:2])
			announce_year_rebal = int(announce_date_rebal[6:8])
			dt_obj = datetime.strptime(announce_date_rebal,'%m/%d/%y')
			a_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
			a_y_rebal = int(a_date_rebal[:4])
			a_date_rebal = date(a_y_rebal,announce_month_rebal,announce_day_rebal)
			if(int(client_comm_rule_rebal)==1):
				client_comm_date_rebal = AnnouncementDate_Calculate(eff_day_rebal,eff_month_rebal,eff_year_rebal,int(client_comm_rule_rebal),holidays)
			else:
				client_comm_date_rebal = AnnouncementDate_Calculate(freeze_day_rebal,freeze_month_rebal,freeze_year_rebal,int(client_comm_rule_rebal),holidays)
			Client_comm_day_rebal = int(client_comm_date_rebal[3:5])
			Client_comm_month_rebal = int(client_comm_date_rebal[0:2])
			Client_comm_year_rebal = int(client_comm_date_rebal[6:8])
			dt_obj = datetime.strptime(client_comm_date_rebal,'%m/%d/%y')
			cl_date_rebal = datetime.strftime(dt_obj, "%Y-%m-%d")
			cl_y_rebal = int(cl_date_rebal[:4])
			cl_date_rebal = date(cl_y_rebal,Client_comm_month_rebal,Client_comm_day_rebal)


			cal_month_rebal = month1_rebal
			month1_rebal= month1_rebal+gap_rebal
			if(next_monday <= e_date_rebal and e_date_rebal <= next_friday):
				flag=1
				ind.Effective_Date_re = ind.Effective_Date_re + date1_rebal + "\n"
			if(next_monday <= f_date_rebal and f_date_rebal <= next_friday):
				flag=1
				ind.Weights_Share_Freeze_Date_re = ind.Weights_Share_Freeze_Date_re + freeze_date_rebal + "\n"
			if(next_monday <= cl_date_rebal and cl_date_rebal <= next_friday):
				flag=1
				ind.Client_Comm_Date_re = ind.Client_Comm_Date_re + client_comm_date_rebal + "\n"
			if(next_monday <= a_date_rebal and a_date_rebal <= next_friday):
				flag=1
				ind.Public_Announcement_Date_re = ind.Public_Announcement_Date_re + announce_date_rebal + "\n"
		for i in range(0,num_month_review):
			year = y1
			if(month1_review>12):
				month1_review = month1_review%12;
			if(month1_review<cal_month_review):
				year = year+1
			if eff_rule_review:
				date1_review = EffectiveDate_Calculate(month1_review,year,int(eff_rule_review),holidays)
				eff_day_review = int(date1_review[3:5])
				eff_month_review = int(date1_review[0:2])
				eff_year_review = int(date1_review[6:8])
				dt_obj = datetime.strptime(date1_review,'%m/%d/%y')
				e_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
				e_y_review = int(e_date_review[:4])
				e_date_review = date(e_y_review,eff_month_review,eff_day_review)

			if sel_rule2_review:
				sel_date_review = SelectionDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(sel_rule2_review),holidays)
				sel_day_review = int(sel_date_review[3:5])
				sel_month_review = int(sel_date_review[0:2])
				sel_year_review = int(sel_date_review[6:8])
				dt_obj = datetime.strptime(sel_date_review,'%m/%d/%y')
				s_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
				s_y_review = int(s_date_review[:4])
				s_date_review = date(s_y_review,sel_month_review,sel_day_review)
			if fre_rule_review:
				freeze_date_review = FreezeDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(fre_rule_review),holidays)
				freeze_day_review = int(freeze_date_review[3:5])
				freeze_month_review = int(freeze_date_review[0:2])
				freeze_year_review = int(freeze_date_review[6:8])
				dt_obj = datetime.strptime(freeze_date_review,'%m/%d/%y')
				f_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
				f_y_review = int(f_date_review[:4])
				f_date_review = date(f_y_review,freeze_month_review,freeze_day_review)

			if(int(ann_rule_review)==1):
				announce_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(ann_rule_review),holidays)
			else:
				announce_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(ann_rule_review),holidays)
			announce_day_review = int(announce_date_review[3:5])
			announce_month_review = int(announce_date_review[0:2])
			announce_year_review = int(announce_date_review[6:8])
			dt_obj = datetime.strptime(announce_date_review,'%m/%d/%y')
			a_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
			a_y_review = int(a_date_review[:4])
			a_date_review = date(a_y_review,announce_month_review,announce_day_review)
			if(int(client_comm_rule_review)==1):
				client_comm_date_review = AnnouncementDate_Calculate(eff_day_review,eff_month_review,eff_year_review,int(client_comm_rule_review),holidays)
			else:
				client_comm_date_review = AnnouncementDate_Calculate(freeze_day_review,freeze_month_review,freeze_year_review,int(client_comm_rule_review),holidays)
			Client_comm_day_review = int(client_comm_date_review[3:5])
			Client_comm_month_review = int(client_comm_date_review[0:2])
			Client_comm_year_review = int(client_comm_date_review[6:8])
			dt_obj = datetime.strptime(client_comm_date_review,'%m/%d/%y')
			cl_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
			cl_y_review = int(cl_date_review[:4])
			cl_date_review = date(cl_y_review,Client_comm_month_review,Client_comm_day_review)
			Prelim_comm_date_review = PreliminaryCommDate_Calculate(announce_day_review,announce_month_review,announce_year_review,holidays)
			Prelim_comm_day_review = int(Prelim_comm_date_review[3:5])
			Prelim_comm_month_review = int(Prelim_comm_date_review[0:2])
			Prelim_comm_year_review = int(Prelim_comm_date_review[6:8])
			dt_obj = datetime.strptime(Prelim_comm_date_review,'%m/%d/%y')
			p_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
			p_y_review = int(p_date_review[:4])
			p_date_review = date(p_y_review,Prelim_comm_month_review,Prelim_comm_day_review)
			
			"""
			if(Prelim_comm_rule_review==""):
				Committee_comm_date_review = Committe_comm_date_Calculate(Client_comm_day_review,Client_comm_month_review,Client_comm_year_review,holidays)
			else:
				Committee_comm_date_review = Committe_comm_date_Calculate(Prelim_comm_day_review,Prelim_comm_month_review,Prelim_comm_year_review,holidays)

			Committee_comm_month_review = int(Committee_comm_date_review[0:2])
			Committee_comm_year_review = int(Committee_comm_date_review[6:8])
			Committee_comm_day_review = int(Committee_comm_date_review[3:5])
			"""
			dt_obj = datetime.strptime(freeze_date_review,'%m/%d/%y')
			com_date_review = datetime.strftime(dt_obj, "%Y-%m-%d")
			com_y_review = int(com_date_review[:4])
			com_date_review = date(com_y_review,freeze_month_review,freeze_day_review)
			cal_month_review = month1_review
			month1_review= month1_review+gap_review
			if(next_monday <= e_date_review and e_date_review <= next_friday):
				flag=1
				ind.Effective_Date_rev = ind.Effective_Date_rev + date1_review + "\n"
			if(next_monday <= f_date_review and f_date_review <= next_friday):
				flag=1
				ind.Weights_Share_Freeze_Date_rev = ind.Weights_Share_Freeze_Date_rev + freeze_date_review + "\n"
			if(next_monday <= s_date_review and s_date_review <= next_friday):
				flag=1
				ind.Selection_Date_Cycle_2_rev = ind.Selection_Date_Cycle_2_rev + sel_date_review + "\n"
			if(next_monday <= a_date_review and a_date_review <= next_friday):
				flag=1
				ind.Public_Announcement_Date_rev = ind.Public_Announcement_Date_rev + announce_date_review + "\n"
			if(next_monday <= cl_date_review and cl_date_review <= next_friday):
				flag=1
				ind.Client_Comm_Date_rev = ind.Client_Comm_Date_rev + client_comm_date_review + "\n"
			if(next_monday <= p_date_review and p_date_review <= next_friday):
				flag=1
				ind.Prelim_Comm_Date_rev = ind.Prelim_Comm_Date_rev + Prelim_comm_date_review + "\n"
			"""
			if(next_monday <= com_date_review and com_date_review <= next_friday):
				flag=1
				ind.Ind_Cmte_Comm_Date_rev = ind.Ind_Cmte_Comm_Date_rev + Committee_comm_date_review + "\n"
			"""


		if(flag==1):
			informa.append(ind)
			body = 'Subject: Email testing'+'\n' + '\nHello, \n\n Email testing for calendar automation\n' + '\nHave a nice day!'
	context = {
				'info' : informa,
	            'client_name' : client_name,}
	'''templateName = "mail.html"
	return render(request, templateName, context)'''

	html_content = render_to_string('mail.html', context) # render with dynamic value
	text_content = strip_tags(html_content)

	body = 'Subject: Email testing'+'\n' + '\nHello, \n\n Email testing for calendar automation\n' + '\nHave a nice day!'
	send_mail(text_content)


